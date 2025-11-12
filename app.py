from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_from_directory, send_file, session, Response
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_socketio import SocketIO, emit, join_room
from flask_mail import Mail, Message
from werkzeug.utils import secure_filename
from functools import wraps
import os
from datetime import datetime, date, timedelta
import re
import threading
import time
import random
from werkzeug.security import generate_password_hash, check_password_hash
from models import db, User, PaymentRequest, AuditLog, Notification, PaidNotification, RecurringPaymentSchedule, LateInstallment, InstallmentEditHistory, RequestType, Branch, BranchAlias, FinanceAdminNote, ChequeBook, ChequeSerial
from config import Config
import json
from playwright.sync_api import sync_playwright
from io import BytesIO
import base64

# --- Maintenance mode storage (instance file) ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_INSTANCE_DIR = os.path.join(BASE_DIR, 'instance')
MAINTENANCE_FILE_PATH = os.path.join(DEFAULT_INSTANCE_DIR, 'maintenance.json')

def read_maintenance_state():
    try:
        with open(MAINTENANCE_FILE_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {"enabled": False, "message": "The system is undergoing maintenance. Please try again later."}

def write_maintenance_state(enabled: bool, message: str = None):
    state = read_maintenance_state()
    state['enabled'] = bool(enabled)
    if message is not None:
        state['message'] = message
    os.makedirs(os.path.dirname(MAINTENANCE_FILE_PATH), exist_ok=True)
    with open(MAINTENANCE_FILE_PATH, 'w', encoding='utf-8') as f:
        json.dump(state, f)

# Initialize Flask app
app = Flask(__name__)
app.config.from_object(Config)

# Initialize maintenance path now that app is created
try:
    MAINTENANCE_FILE_PATH = os.path.join(app.instance_path, 'maintenance.json')
except Exception:
    # Fallback to default path if app.instance_path not available
    MAINTENANCE_FILE_PATH = os.path.join(DEFAULT_INSTANCE_DIR, 'maintenance.json')

# Make timedelta available in templates
from datetime import timedelta
app.jinja_env.globals.update(timedelta=timedelta)

# Timezone conversion function
def utc_to_local(utc_datetime):
    """Convert UTC datetime to Oman time (UTC+4)"""
    if utc_datetime is None:
        return None
    return utc_datetime + timedelta(hours=4)

# Make timezone conversion available in templates
app.jinja_env.globals.update(utc_to_local=utc_to_local)

# Add JSON filter for templates
import json
@app.template_filter('from_json')
def from_json_filter(value):
    """Parse JSON string to Python object"""
    try:
        return json.loads(value) if value else []
    except (json.JSONDecodeError, TypeError):
        return value if value else []

def format_recurring_schedule(interval, payment_schedule=None):
    """Format recurring interval into human-readable text"""
    try:
        parts = interval.split(':')
        frequency = parts[0]
        interval_value = int(parts[1])
        
        # Handle the new single date format
        if frequency == 'monthly' and len(parts) > 2 and parts[2] == 'date':
            # Extract the date (format: YYYY-MM-DD)
            date_str = parts[3]
            
            try:
                from datetime import datetime
                date_obj = datetime.strptime(date_str, '%Y-%m-%d')
                formatted_date = date_obj.strftime('%B %d, %Y')
            except ValueError:
                formatted_date = date_str
            
            # Create the base schedule text
            if interval_value == 1:
                schedule_text = f"Every month starting on {formatted_date}"
            else:
                schedule_text = f"Every {interval_value} months starting on {formatted_date}"
            
            # Check for end date
            end_date_index = parts.index('end') if 'end' in parts else -1
            if end_date_index != -1 and end_date_index + 1 < len(parts):
                end_date = parts[end_date_index + 1]
                try:
                    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
                    formatted_end_date = end_date_obj.strftime('%B %d, %Y')
                    schedule_text += f" until {formatted_end_date}"
                except ValueError:
                    schedule_text += f" until {end_date}"
            
            # Add payment schedule information if available
            if payment_schedule:
                try:
                    import json
                    schedule_data = json.loads(payment_schedule)
                    if schedule_data:
                        payment_details = []
                        for payment in schedule_data:
                            date = payment.get('date', '')
                            amount = payment.get('amount', 0)
                            if date and amount > 0:
                                payment_details.append(f"{date}: {amount:.3f} OMR")
                        
                        if payment_details:
                            schedule_text += f"\n\nPayment Schedule:\n" + "\n".join(payment_details)
                except (json.JSONDecodeError, TypeError):
                    pass
            
            return schedule_text
        
        # Handle the old format with specific days (for backward compatibility)
        elif frequency == 'monthly' and len(parts) > 2 and parts[2] == 'days':
            # Extract specific days and dates
            days = parts[3].split(',')
            year = parts[4] if len(parts) > 4 else None
            month = parts[5] if len(parts) > 5 else None
            
            # Format the starting dates using the exact same logic as the preview
            if year and month:
                month_names = ['January', 'February', 'March', 'April', 'May', 'June', 
                             'July', 'August', 'September', 'October', 'November', 'December']
                # The stored month is 1-based (1-12), so convert to 0-based for array indexing
                month_int = int(month) - 1  # Convert 1-based to 0-based
                month_name = month_names[month_int] if 0 <= month_int <= 11 else str(month)
                starting_dates = [f"{month_name} {day}, {year}" for day in days]
                starting_text = ", ".join(starting_dates)
            else:
                starting_text = f"days {', '.join(days)}"
            
            # Create the base schedule text using the exact same format as the preview
            if interval_value == 1:
                schedule_text = f"Every month starting on {starting_text}"
            else:
                schedule_text = f"Every {interval_value} months starting on {starting_text}"
            
            # Check for end date
            end_date_index = parts.index('end') if 'end' in parts else -1
            if end_date_index != -1 and end_date_index + 1 < len(parts):
                end_date = parts[end_date_index + 1]
                try:
                    from datetime import datetime
                    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
                    formatted_end_date = end_date_obj.strftime('%B %d, %Y')
                    schedule_text += f" until {formatted_end_date}"
                except ValueError:
                    schedule_text += f" until {end_date}"
            
            # Add payment schedule information if available
            if payment_schedule:
                try:
                    import json
                    schedule_data = json.loads(payment_schedule)
                    if schedule_data:
                        payment_details = []
                        for payment in schedule_data:
                            date = payment.get('date', '')
                            amount = payment.get('amount', 0)
                            if date and amount > 0:
                                payment_details.append(f"{date}: {amount:.3f} OMR")
                        
                        if payment_details:
                            schedule_text += f"\n\nPayment Schedule:\n" + "\n".join(payment_details)
                except (json.JSONDecodeError, TypeError):
                    pass
            
            return schedule_text
        
        # Handle legacy formats
        if frequency == 'daily':
            if interval_value == 1:
                return "Daily"
            else:
                return f"Every {interval_value} days"
        elif frequency == 'weekly':
            if interval_value == 1:
                return "Weekly"
            else:
                return f"Every {interval_value} weeks"
        elif frequency == 'monthly':
            if interval_value == 1:
                return "Monthly"
            else:
                return f"Every {interval_value} months"
        elif frequency == 'quarterly':
            if interval_value == 1:
                return "Quarterly"
            else:
                return f"Every {interval_value} quarters"
        elif frequency == 'yearly':
            if interval_value == 1:
                return "Yearly"
            else:
                return f"Every {interval_value} years"
    except (ValueError, IndexError):
        return "Invalid schedule"

# Make the function available in templates
app.jinja_env.globals.update(format_recurring_schedule=format_recurring_schedule)

# Initialize database
db.init_app(app)

# Initialize Flask-Mail
mail = Mail(app)

# Initialize SocketIO for real-time updates
socketio = SocketIO(app, cors_allowed_origins="*", async_mode='threading')

@socketio.on('connect')
def handle_connect():
    """Handle client connection"""
    print(f'Client connected: {request.sid}')

@socketio.on('disconnect')
def handle_disconnect():
    """Handle client disconnection"""
    print(f'Client disconnected: {request.sid}')

@socketio.on('join_room')
def handle_join_room(data):
    """Handle client joining a room"""
    room = data.get('room')
    if room and current_user.is_authenticated:
        if room == 'finance_admin' and current_user.role in ['Finance Staff', 'Finance Admin']:
            join_room(room)
            print(f'User {current_user.username} joined room: {room}')
        elif room == 'all_users':
            join_room(room)
            print(f'User {current_user.username} joined room: {room}')

# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Create upload folder if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['CHEQUE_UPLOAD_FOLDER'], exist_ok=True)


@login_manager.user_loader
def load_user(user_id):
    """Load user for Flask-Login"""
    return User.query.get(int(user_id))

# --- Idle timeout enforcement ---
@app.before_request
def enforce_idle_timeout():
    try:
        # Skip static files, socket traffic, and login/logout routes
        if request.path.startswith('/static') or request.path.startswith('/socket.io'):
            return None
        if request.endpoint in ['login', 'logout', 'check_tab_session']:
            return None
        # Allow maintenance endpoints and health checks to pass through
        if request.path.startswith('/maintenance') or request.path.startswith('/health'):
            return None

        if current_user.is_authenticated:
            now_ts = datetime.utcnow().timestamp()
            last_activity = session.get('last_activity')
            session_start = session.get('session_start')
            tab_session_id = session.get('tab_session_id')
            
            # If session doesn't have last_activity, session_start, or tab_session_id, it's a stale session
            # This can happen if Flask-Login restored a user from an old session cookie or a new tab was opened
            # Log them out immediately and clear all session data
            if last_activity is None or session_start is None or tab_session_id is None:
                try:
                    logout_user()
                except Exception:
                    pass
                # Clear all session data including Flask-Login's user_id
                session.clear()
                flash('Your session has expired. Please log in again.', 'warning')
                return redirect(url_for('login'))
            
            idle_seconds = now_ts - float(last_activity)
            # 10 minutes = 600 seconds
            if idle_seconds > 600:
                # Invalidate session and force logout
                try:
                    logout_user()
                except Exception:
                    pass
                session.clear()
                flash('You were logged out due to inactivity.', 'warning')
                return redirect(url_for('login'))
            # Update last activity timestamp for active user
            session['last_activity'] = now_ts
    except Exception:
        # Fail-open: do not block the request on any error here
        return None

# --- Maintenance gate ---
@app.before_request
def maintenance_gate():
    try:
        state = read_maintenance_state()
        if not state.get('enabled'):
            return None

        # Allow static and socket endpoints
        if request.path.startswith('/static') or request.path.startswith('/socket.io'):
            return None

        # Allow maintenance endpoints so IT can toggle and for public status
        if request.path.startswith('/maintenance'):
            return None

        # Allow health checks if any
        if request.path.startswith('/health'):
            return None

        # CRITICAL: Allow login-related routes so IT users can log in to disable maintenance
        # This prevents the lockout scenario where maintenance is enabled and no one can log in
        if request.path in ['/login', '/logout', '/validate_credentials', '/verify_pin']:
            return None

        # Allow IT department users to proceed
        if current_user.is_authenticated and (getattr(current_user, 'department', None) == 'IT'):
            return None

        # For everyone else, show maintenance page
        message = state.get('message') or 'The system is undergoing maintenance. Please try again later.'
        return render_template('maintenance.html', message=message), 503
    except Exception:
        # Fail-open on any error to avoid locking out IT unintentionally
        return None

# --- Maintenance endpoints (IT only) ---
def it_department_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('login'))
        if getattr(current_user, 'department', None) != 'IT':
            flash('Only IT department can perform this action.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return wrapper

@app.route('/maintenance/status')
@login_required
@it_department_required
def maintenance_status():
    return jsonify(read_maintenance_state())

@app.route('/maintenance/public_status')
def maintenance_public_status():
    state = read_maintenance_state()
    return jsonify({ 'enabled': bool(state.get('enabled')) })

@app.route('/maintenance/enable', methods=['POST'])
@login_required
@it_department_required
def maintenance_enable():
    message = request.form.get('message') or request.json.get('message') if request.is_json else None
    write_maintenance_state(True, message)
    log_action('Enabled maintenance mode')
    try:
        socketio.emit('maintenance_update', { 'enabled': True, 'message': message or '' }, broadcast=True)
    except Exception:
        pass
    flash('Maintenance mode enabled.', 'warning')
    return redirect(request.referrer or url_for('dashboard'))

@app.route('/maintenance/disable', methods=['POST'])
@login_required
@it_department_required
def maintenance_disable():
    write_maintenance_state(False)
    log_action('Disabled maintenance mode')
    try:
        socketio.emit('maintenance_update', { 'enabled': False }, broadcast=True)
    except Exception:
        pass
    flash('Maintenance mode disabled.', 'success')
    return redirect(request.referrer or url_for('dashboard'))


def role_required(*roles):
    """Decorator to check if user has required role"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            if current_user.role not in roles:
                flash('You do not have permission to access this page.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator


def log_action(action):
    """Helper function to log user actions"""
    if current_user.is_authenticated:
        log = AuditLog(
            user_id=current_user.user_id, 
            action=action,
            username_snapshot=current_user.username
        )
        db.session.add(log)
        db.session.commit()


def get_status_priority_order():
    """
    Returns SQLAlchemy case expression for ordering payment requests by status priority.
    Priority order:
    1. Pending Manager Approval
    2. Pending Finance Approval
    3. Proof Pending
    4. Proof Sent
    5. Proof Rejected
    6. Recurring
    7. Completed
    8. Rejected (Rejected by Manager, Rejected by Finance, Proof Rejected)
    """
    return db.case(
        (PaymentRequest.status == 'Pending Manager Approval', 1),
        (PaymentRequest.status == 'Pending Finance Approval', 2),
        (PaymentRequest.status == 'Proof Pending', 3),
        (PaymentRequest.status == 'Proof Sent', 4),
        (PaymentRequest.status == 'Proof Rejected', 5),
        (PaymentRequest.status == 'Recurring', 6),
        (PaymentRequest.status == 'Completed', 7),
        (PaymentRequest.status.in_(['Rejected by Manager', 'Rejected by Finance', 'Proof Rejected']), 8),
        else_=99  # Any other status goes to the end
    )


def get_all_tab_datetime_order():
    """
    Secondary ordering for the All Requests tab by per-status recency (most recent first):
    - Completed: finance_approval_end_time → completion_date → approval_date → updated_at → created_at
    - Pending Finance Approval: finance_approval_start_time → manager_approval_end_time → updated_at → created_at
    - Pending Manager Approval: manager_approval_start_time → updated_at → created_at
    - Proof Pending: updated_at → finance_approval_end_time → created_at
    - Proof Sent: updated_at → created_at
    - Proof Rejected: finance_rejection_date → updated_at → created_at
    - Recurring: updated_at → created_at
    - Rejected by Manager: manager_rejection_date → updated_at → created_at
    - Rejected by Finance: finance_rejection_date → updated_at → created_at
    - Default: created_at
    """
    return db.case(
        (
            PaymentRequest.status == 'Completed',
            db.func.coalesce(
                PaymentRequest.finance_approval_end_time,
                PaymentRequest.completion_date,
                PaymentRequest.approval_date,
                PaymentRequest.updated_at,
                PaymentRequest.created_at
            )
        ),
        (
            PaymentRequest.status == 'Pending Finance Approval',
            db.func.coalesce(
                PaymentRequest.finance_approval_start_time,
                PaymentRequest.manager_approval_end_time,
                PaymentRequest.updated_at,
                PaymentRequest.created_at
            )
        ),
        (
            PaymentRequest.status == 'Pending Manager Approval',
            db.func.coalesce(
                PaymentRequest.manager_approval_start_time,
                PaymentRequest.updated_at,
                PaymentRequest.created_at
            )
        ),
        (
            PaymentRequest.status == 'Proof Pending',
            db.func.coalesce(
                PaymentRequest.updated_at,
                PaymentRequest.finance_approval_end_time,
                PaymentRequest.created_at
            )
        ),
        (
            PaymentRequest.status == 'Proof Sent',
            db.func.coalesce(
                PaymentRequest.updated_at,
                PaymentRequest.created_at
            )
        ),
        (
            PaymentRequest.status == 'Proof Rejected',
            db.func.coalesce(
                PaymentRequest.finance_rejection_date,
                PaymentRequest.updated_at,
                PaymentRequest.created_at
            )
        ),
        (
            PaymentRequest.status == 'Recurring',
            db.func.coalesce(
                PaymentRequest.updated_at,
                PaymentRequest.created_at
            )
        ),
        (
            PaymentRequest.status == 'Rejected by Manager',
            db.func.coalesce(
                PaymentRequest.manager_rejection_date,
                PaymentRequest.updated_at,
                PaymentRequest.created_at
            )
        ),
        (
            PaymentRequest.status == 'Rejected by Finance',
            db.func.coalesce(
                PaymentRequest.finance_rejection_date,
                PaymentRequest.updated_at,
                PaymentRequest.created_at
            )
        ),
        else_=PaymentRequest.created_at
    ).desc()


def get_completed_datetime_order():
    """Ordering for completed requests: finance_approval_end_time → completion_date → approval_date → updated_at → created_at (desc)."""
    return db.func.coalesce(
        PaymentRequest.finance_approval_end_time,
        PaymentRequest.completion_date,
        PaymentRequest.approval_date,
        PaymentRequest.updated_at,
        PaymentRequest.created_at
    ).desc()


def get_rejected_datetime_order():
    """Ordering for rejected requests: pick manager/finance rejection date → updated_at → created_at (desc)."""
    return db.func.coalesce(
        PaymentRequest.finance_rejection_date,
        PaymentRequest.manager_rejection_date,
        PaymentRequest.updated_at,
        PaymentRequest.created_at
    ).desc()


def get_recurring_datetime_order():
    """Ordering for recurring requests: updated_at → created_at (desc)."""
    return db.func.coalesce(
        PaymentRequest.updated_at,
        PaymentRequest.created_at
    ).desc()


def send_pin_email(user_email, user_name, pin):
    """Send PIN to user's email"""
    try:
        msg = Message(
            subject='Your Login PIN - Payment Request System',
            sender=app.config['MAIL_DEFAULT_SENDER'],
            recipients=[user_email]
        )
        
        msg.html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body {{
                    font-family: Arial, sans-serif;
                    line-height: 1.6;
                    color: #333;
                }}
                .container {{
                    max-width: 600px;
                    margin: 0 auto;
                    padding: 20px;
                    background-color: #f9f9f9;
                }}
                .header {{
                    background: linear-gradient(135deg, #2c3e50 0%, #1a252f 100%);
                    color: white;
                    padding: 30px;
                    text-align: center;
                    border-radius: 10px 10px 0 0;
                }}
                .content {{
                    background-color: white;
                    padding: 30px;
                    border-radius: 0 0 10px 10px;
                }}
                .pin-box {{
                    background-color: #f0f0f0;
                    border: 2px solid #2c3e50;
                    border-radius: 10px;
                    padding: 20px;
                    text-align: center;
                    margin: 20px 0;
                }}
                .pin {{
                    font-size: 36px;
                    font-weight: bold;
                    color: #2c3e50;
                    letter-spacing: 10px;
                }}
                .warning {{
                    background-color: #fff3cd;
                    border-left: 4px solid #ffc107;
                    padding: 15px;
                    margin: 20px 0;
                }}
                .footer {{
                    text-align: center;
                    color: #666;
                    margin-top: 20px;
                    font-size: 12px;
                }}
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h1>🔐 Login PIN</h1>
                    <p>Payment Request Management System</p>
                </div>
                <div class="content">
                    <p>Hello <strong>{user_name}</strong>,</p>
                    <p>Your login PIN has been generated. Please use this PIN to complete your login:</p>
                    
                    <div class="pin-box">
                        <div class="pin">{pin}</div>
                    </div>
                    
                    <div class="warning">
                        <strong>⚠️ Important:</strong>
                        <ul>
                            <li>This PIN is valid for <strong>2 minutes</strong></li>
                            <li>Do not share this PIN with anyone</li>
                            <li>If you didn't request this PIN, please contact IT immediately</li>
                        </ul>
                    </div>
                    
                    <p>If you're having trouble logging in, please contact the IT department for assistance.</p>
                </div>
                <div class="footer">
                    <p>This is an automated message. Please do not reply to this email.</p>
                    <p>&copy; 2024 Payment Request Management System</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        msg.body = f"""
        Hello {user_name},

        Your login PIN has been generated. Please use this PIN to complete your login:

        PIN: {pin}

        IMPORTANT:
        - This PIN is valid for 2 minutes
        - Do not share this PIN with anyone
        - If you didn't request this PIN, please contact IT immediately

        If you're having trouble logging in, please contact the IT department for assistance.

        This is an automated message. Please do not reply to this email.
        
        © 2024 Payment Request Management System
        """
        
        mail.send(msg)
        return True, "PIN sent successfully"
    except Exception as e:
        app.logger.error(f"Failed to send PIN email: {str(e)}")
        return False, f"Failed to send email: {str(e)}"


def create_notification(user_id, title, message, notification_type, request_id=None):
    """Helper function to create notifications"""
    print(f"DEBUG: Creating notification for user_id: {user_id}")
    print(f"   - title: {title}")
    print(f"   - message: {message}")
    print(f"   - notification_type: {notification_type}")
    print(f"   - request_id: {request_id}")
    
    notification = Notification(
        user_id=user_id,
        title=title,
        message=message,
        notification_type=notification_type,
        request_id=request_id
    )
    db.session.add(notification)
    db.session.commit()
    
    print(f"DEBUG: Notification created successfully with ID: {notification.notification_id}")
    return notification

def get_authorized_manager_approvers(request):
    """Get all users who are authorized to approve this request at the manager stage.
    This mirrors the authorization logic in manager_approve_request."""
    authorized_users = []
    
    # If temporary manager is assigned, only that user is authorized
    if request.temporary_manager_id:
        temp_manager = User.query.get(request.temporary_manager_id)
        if temp_manager:
            authorized_users.append(temp_manager)
            print(f"DEBUG: Found temporary manager: {temp_manager.name} (ID: {temp_manager.user_id})")
        return authorized_users
    
    # No temporary manager, use standard authorization checks
    
    # Hard rule: Requests submitted by GM/CEO/Operation Manager can ONLY be approved by Abdalaziz
    # Abdalaziz is their assigned manager and should receive notifications
    if request.user.role in ['GM', 'CEO', 'Operation Manager']:
        abdalaziz = User.query.filter_by(name='Abdalaziz Al-Brashdi').first()
        if abdalaziz:
            authorized_users.append(abdalaziz)
            print(f"DEBUG: Added Abdalaziz for GM/CEO/Operation Manager submitter (assigned manager)")
    else:
        # General Manager - can approve all requests (except GM/CEO/Operation Manager)
        gm_users = User.query.filter_by(role='GM').all()
        for user in gm_users:
            if user.user_id != request.user_id:  # Can't approve own request
                authorized_users.append(user)
                print(f"DEBUG: Added GM: {user.name}")
        
        # Operation Manager - can approve all requests (except GM/CEO/Operation Manager)
        op_manager_users = User.query.filter_by(role='Operation Manager').all()
        for user in op_manager_users:
            if user.user_id != request.user_id:  # Can't approve own request
                authorized_users.append(user)
                print(f"DEBUG: Added Operation Manager: {user.name}")
        
        # Check if submitter has a manager_id (direct manager relationship)
        if request.user.manager_id:
            manager = User.query.get(request.user.manager_id)
            if manager and manager.user_id != request.user_id:
                if manager not in authorized_users:
                    authorized_users.append(manager)
                    print(f"DEBUG: Added direct manager: {manager.name} (ID: {manager.user_id})")
        
        # Special case: General Manager can approve Department Manager requests
        if request.user.role == 'Department Manager':
            gm_users = User.query.filter_by(role='GM').all()
            for user in gm_users:
                if user not in authorized_users:
                    authorized_users.append(user)
                    print(f"DEBUG: Added GM for Department Manager request: {user.name}")
            
            # Operation Manager can also approve Department Manager requests
            op_manager_users = User.query.filter_by(role='Operation Manager').all()
            for user in op_manager_users:
                if user not in authorized_users:
                    authorized_users.append(user)
                    print(f"DEBUG: Added Operation Manager for Department Manager request: {user.name}")
        
        # Special case: Abdalaziz can approve GM, CEO, Finance Staff, and Operation Manager requests
        # (But GM/CEO/Operation Manager requests are handled above, so this covers Finance Staff)
        # Abdalaziz is the assigned manager for Finance Staff and should receive notifications
        if request.user.role == 'Finance Staff':
            abdalaziz = User.query.filter_by(name='Abdalaziz Al-Brashdi').first()
            if abdalaziz and abdalaziz not in authorized_users:
                authorized_users.append(abdalaziz)
                print(f"DEBUG: Added Abdalaziz for Finance Staff request (assigned manager)")
        
        # Special case: Operation Manager can approve Operation department and Project requests
        if (request.user.department == 'Operation' or request.user.department == 'Project') and \
           request.user.role != 'Operation Manager':
            op_manager_users = User.query.filter_by(role='Operation Manager').all()
            for user in op_manager_users:
                if user.user_id != request.user_id and user not in authorized_users:
                    authorized_users.append(user)
                    print(f"DEBUG: Added Operation Manager for Operation/Project department: {user.name}")
        
        # Special case: Finance Admin can approve Finance department requests
        if request.user.department == 'Finance' and request.user.role != 'Finance Admin':
            finance_admins = User.query.filter_by(role='Finance Admin').all()
            for user in finance_admins:
                if user.user_id != request.user_id and user not in authorized_users:
                    authorized_users.append(user)
                    print(f"DEBUG: Added Finance Admin for Finance department: {user.name}")
        
        # Special case: Department Manager can approve same department requests
        dept_managers = User.query.filter_by(
            role='Department Manager',
            department=request.user.department
        ).all()
        for user in dept_managers:
            if user.user_id != request.user_id and user not in authorized_users:
                authorized_users.append(user)
                print(f"DEBUG: Added Department Manager for same department: {user.name}")
    
    # Remove duplicates (in case a user was added multiple times)
    seen = set()
    unique_authorized = []
    for user in authorized_users:
        if user.user_id not in seen:
            seen.add(user.user_id)
            unique_authorized.append(user)
    
    print(f"DEBUG: Total authorized manager approvers: {len(unique_authorized)}")
    return unique_authorized

def notify_users_by_role(request, notification_type, title, message, request_id=None):
    """Notify users based on RBAC notification permissions"""
    
    # Handle user management notifications (no request object)
    if notification_type in ['user_created', 'user_updated', 'user_deleted']:
        it_staff_users = User.query.filter_by(role='IT Staff').all()
        for user in it_staff_users:
            create_notification(user.user_id, title, message, notification_type, request_id)
        return
    
    # Get the requestor's role and department
    requestor_role = request.user.role
    requestor_department = request.department
    
    print(f"DEBUG: notify_users_by_role called")
    print(f"   - notification_type: {notification_type}")
    print(f"   - requestor_role: {requestor_role}")
    print(f"   - requestor_department: {requestor_department}")
    print(f"   - title: {title}")
    
    # Finance Admin and Finance Staff - get notifications when requests reach Pending Finance Approval
    if notification_type in ['ready_for_finance_review']:
        finance_users = User.query.filter(User.role.in_(['Finance Staff', 'Finance Admin'])).all()
        for user in finance_users:
            create_notification(user.user_id, title, message, notification_type, request_id)
    
    # Handle new_submission notifications
    elif notification_type == 'new_submission':
        print(f"DEBUG: Processing new_submission for role: {requestor_role}, department: {requestor_department}")
        
        # First, notify the submitting user (request creator)
        submitter_notification_title = "Payment Request Submitted"
        request_type_name = getattr(request, 'request_type', 'payment request')
        if hasattr(request, 'amount') and request.amount is not None:
            try:
                amount_str = f"{float(request.amount):.3f}"
            except (ValueError, TypeError):
                amount_str = str(request.amount)
        else:
            amount_str = 'N/A'
        submitter_notification_message = f"Your {request_type_name} request for OMR {amount_str} has been submitted successfully and is awaiting manager approval."
        create_notification(request.user_id, submitter_notification_title, submitter_notification_message, notification_type, request_id)
        print(f"DEBUG: Notified submitting user: {request.user.name} (ID: {request.user_id})")
        
        # Get all authorized manager approvers using the helper function
        authorized_approvers = get_authorized_manager_approvers(request)
        
        # Notify all authorized manager approvers
        for approver in authorized_approvers:
            create_notification(approver.user_id, title, message, notification_type, request_id)
            print(f"DEBUG: Notified authorized approver: {approver.name} (ID: {approver.user_id}, Role: {approver.role})")
        
        # General Manager - receives notifications from ALL requests (regardless of role/department)
        gm_users = User.query.filter_by(role='GM').all()
        for user in gm_users:
            # Only notify if not already in authorized_approvers (avoid duplicates)
            if user not in authorized_approvers:
                create_notification(user.user_id, title, message, notification_type, request_id)
                print(f"DEBUG: Notified GM {user.username} about new_submission")
        
        # Operation Manager - receives notifications from ALL requests (regardless of role/department)
        op_manager_users = User.query.filter_by(role='Operation Manager').all()
        for user in op_manager_users:
            # Only notify if not already in authorized_approvers (avoid duplicates)
            if user not in authorized_approvers:
                create_notification(user.user_id, title, message, notification_type, request_id)
                print(f"DEBUG: Notified Operation Manager {user.username} about new_submission")
        
        # IT Department Manager - only from IT Staff submissions (explicitly notify even if not in authorized_approvers)
        if requestor_role == 'IT Staff' or requestor_department == 'IT':
            it_manager_users = User.query.filter_by(role='Department Manager', department='IT').all()
            for user in it_manager_users:
                # Only notify if not already in authorized_approvers (avoid duplicates)
                if user not in authorized_approvers:
                    create_notification(user.user_id, title, message, notification_type, request_id)
                    print(f"DEBUG: Notified IT Department Manager {user.username} about IT Staff request")
        
        print(f"DEBUG: Total notifications sent for new_submission: 1 to submitter + {len(authorized_approvers)} to authorized approvers + GM/Operation Manager/IT Manager")
    
    # Requestor - for updates on their own requests
    elif notification_type in ['request_rejected', 'request_approved', 'proof_uploaded', 'status_changed']:
        create_notification(request.user_id, title, message, notification_type, request_id)
        
        # For proof_uploaded, also notify Finance Admin
        if notification_type == 'proof_uploaded':
            print(f"DEBUG: Also notifying Finance Admin about proof upload")
            finance_users = User.query.filter(User.role.in_(['Finance Staff', 'Finance Admin'])).all()
            for user in finance_users:
                create_notification(
                    user_id=user.user_id,
                    title=title,
                    message=message,
                    notification_type=notification_type,
                    request_id=request_id
                )
                print(f"DEBUG: Notified Finance user {user.username} about proof upload")
        
        # For request_approved, also notify Finance Admin if they didn't approve it
        elif notification_type == 'request_approved':
            print(f"DEBUG: Also notifying Finance Admin about request approval")
            finance_users = User.query.filter(User.role.in_(['Finance Staff', 'Finance Admin'])).all()
            for user in finance_users:
                create_notification(
                    user_id=user.user_id,
                    title=title,
                    message=message,
                    notification_type=notification_type,
                    request_id=request_id
                )
                print(f"DEBUG: Notified Finance user {user.username} about request approval")
    
    # Emit real-time notification to all users after creating database notifications
    try:
        socketio.emit('new_notification', {
            'title': title,
            'message': message,
            'type': notification_type,
            'request_id': request_id
        }, room='all_users')
        
        # Also emit a general update event to trigger notification count updates
        socketio.emit('notification_update', {
            'action': 'new_notification',
            'type': notification_type
        }, room='all_users')
        
        print(f"DEBUG: WebSocket events emitted for {notification_type}")
    except Exception as e:
        print(f"Error emitting WebSocket notification: {e}")

def notify_recurring_payment_due(request_id, user_id, title, message):
    """Notify specific user about recurring payment due"""
    create_notification(user_id, title, message, 'recurring_due', request_id)


def notify_system_wide(title, message, notification_type):
    """Notify all users who should receive system-wide notifications"""
    # Roles that receive system-wide notifications
    system_roles = ['Finance Admin', 'Finance Staff', 'GM', 'CEO', 'Operation Manager', 'IT Staff', 'Department Manager']
    
    for role in system_roles:
        users = User.query.filter_by(role=role).all()
        for user in users:
            create_notification(user.user_id, title, message, notification_type)

def create_recurring_payment_schedule(request_id, total_amount, payment_schedule_data):
    """Create a recurring payment schedule with variable amounts"""
    try:
        # Clear any existing schedule for this request
        RecurringPaymentSchedule.query.filter_by(request_id=request_id).delete()
        
        # Create new schedule entries
        for i, payment_data in enumerate(payment_schedule_data, 1):
            schedule_entry = RecurringPaymentSchedule(
                request_id=request_id,
                payment_date=datetime.strptime(payment_data['date'], '%Y-%m-%d').date(),
                amount=payment_data['amount'],
                payment_order=i
            )
            db.session.add(schedule_entry)
        
        db.session.commit()
        return True
    except Exception as e:
        db.session.rollback()
        print(f"Error creating payment schedule: {e}")
        return False

def get_payment_schedule(request_id):
    """Get the payment schedule for a specific request"""
    schedules = RecurringPaymentSchedule.query.filter_by(request_id=request_id).order_by(RecurringPaymentSchedule.payment_order).all()
    return [schedule.to_dict() for schedule in schedules]

def check_recurring_payments_due():
    """Check for recurring payments due today and create notifications"""
    today = date.today()
    
    # Get all recurring payment requests with Recurring status
    recurring_requests = PaymentRequest.query.filter_by(
        recurring='Recurring',
        status='Recurring'
    ).all()
    
    for request in recurring_requests:
        # Check if this request has a custom payment schedule
        payment_schedules = RecurringPaymentSchedule.query.filter_by(
            request_id=request.request_id,
            is_paid=False
        ).order_by(RecurringPaymentSchedule.payment_order).all()
        
        if payment_schedules:
            # Handle variable amount recurring payments
            for schedule in payment_schedules:
                if schedule.payment_date == today:
                    # Check if payment was already marked as paid today
                    paid_today = PaidNotification.query.filter_by(
                        request_id=request.request_id,
                        paid_date=today
                    ).first()
                    
                    if paid_today:
                        continue  # Skip this payment - already marked as paid today
                    
                    # Check if notification already exists for today
                    start_of_day = datetime.combine(today, datetime.min.time())
                    end_of_day = datetime.combine(today, datetime.max.time())
                    
                    existing_notification = Notification.query.filter_by(
                        request_id=request.request_id,
                        notification_type='recurring_due'
                    ).filter(
                        Notification.created_at >= start_of_day,
                        Notification.created_at <= end_of_day
                    ).first()
                    
                    if not existing_notification:
                        # Create notifications for all admin and project users
                        admin_users = User.query.filter(User.role.in_(['Finance Admin', 'Project Staff'])).all()
                        for user in admin_users:
                            create_notification(
                                user_id=user.user_id,
                                title="Recurring Payment Due",
                                message=f'Recurring payment due today for {request.request_type} - {request.purpose} (Amount: {schedule.amount} OMR)',
                                notification_type='recurring_due',
                                request_id=request.request_id
                            )
                        
                        # Log the action
                        log_action(f"Recurring payment due notification created for request #{request.request_id} - Payment {schedule.payment_order}")
        else:
            # Handle traditional recurring payments (single amount)
            if is_payment_due_today(request, today):
                # Check if payment was already marked as paid today
                paid_today = PaidNotification.query.filter_by(
                    request_id=request.request_id,
                    paid_date=today
                ).first()
                
                if paid_today:
                    continue  # Skip this payment - already marked as paid today
                
                # Check if notification already exists for today (more robust check)
                start_of_day = datetime.combine(today, datetime.min.time())
                end_of_day = datetime.combine(today, datetime.max.time())
                
                existing_notification = Notification.query.filter_by(
                    request_id=request.request_id,
                    notification_type='recurring_due'
                ).filter(
                    Notification.created_at >= start_of_day,
                    Notification.created_at <= end_of_day
                ).first()
                
                if not existing_notification:
                    # Create notification for all admin and project users
                    admin_users = User.query.filter(User.role.in_(['Finance Admin', 'Project Staff'])).all()
                    for user in admin_users:
                        create_notification(
                            user_id=user.user_id,
                            title="Recurring Payment Due",
                            message=f'Recurring payment due today for {request.request_type} - {request.purpose}',
                            notification_type='recurring_due',
                            request_id=request.request_id
                        )
                    
                    # Log the action
                    log_action(f"Recurring payment due notification created for request #{request.request_id}")


def get_overdue_requests_count():
    """Get count of overdue finance approval requests"""
    try:
        current_time = datetime.utcnow()
        
        # Get ALL requests that have started finance approval timing but haven't ended
        # Finance approval is still moving if finance_approval_start_time is set
        # and finance_approval_end_time is None
        # Exclude archived requests
        pending_requests = PaymentRequest.query.filter(
            PaymentRequest.finance_approval_start_time.isnot(None),
            PaymentRequest.finance_approval_end_time.is_(None),  # Finance approval still moving
            PaymentRequest.is_archived == False  # Exclude archived requests
        ).all()
        
        overdue_count = 0
        for request in pending_requests:
            # EXCLUDE requests with status "Completed" or "Recurring"
            # These statuses mean the request is no longer in finance approval stage
            if request.status in ['Completed', 'Recurring']:
                continue
            
            # Calculate time elapsed since finance approval started
            time_elapsed = current_time - request.finance_approval_start_time
            
            # Determine alert thresholds based on urgency
            if request.is_urgent:
                # Urgent requests: 2 hours or more
                alert_threshold = timedelta(hours=2)
            else:
                # Non-urgent requests: 24 hours or more (includes days)
                alert_threshold = timedelta(hours=24)
            
            # Check if this request is overdue
            # If time elapsed >= threshold, it MUST be counted
            # This includes requests showing "days" since 1 day = 24 hours
            if time_elapsed >= alert_threshold:
                overdue_count += 1
        
        return overdue_count
        
    except Exception as e:
        print(f"Error getting overdue requests count: {e}")
        import traceback
        traceback.print_exc()
        return 0


def get_overdue_requests():
    """Get all overdue finance approval requests"""
    try:
        current_time = datetime.utcnow()
        
        # Get ALL requests that have started finance approval timing but haven't ended
        # Finance approval is still moving if finance_approval_start_time is set
        # and finance_approval_end_time is None
        # Exclude archived requests
        pending_requests = PaymentRequest.query.filter(
            PaymentRequest.finance_approval_start_time.isnot(None),
            PaymentRequest.finance_approval_end_time.is_(None),  # Finance approval still moving
            PaymentRequest.is_archived == False  # Exclude archived requests
        ).all()
        
        overdue_requests = []
        for request in pending_requests:
            # EXCLUDE requests with status "Completed" or "Recurring"
            # These statuses mean the request is no longer in finance approval stage
            if request.status in ['Completed', 'Recurring']:
                continue
            
            # Calculate time elapsed since finance approval started
            time_elapsed = current_time - request.finance_approval_start_time
            
            # Determine alert thresholds based on urgency
            if request.is_urgent:
                # Urgent requests: 2 hours or more
                alert_threshold = timedelta(hours=2)
            else:
                # Non-urgent requests: 24 hours or more (includes days)
                alert_threshold = timedelta(hours=24)
            
            # Check if this request is overdue
            # If time elapsed >= threshold, it MUST appear on overdue page
            # This includes requests showing "days" since 1 day = 24 hours
            if time_elapsed >= alert_threshold:
                # Format time elapsed for display (show days, hours, and minutes)
                total_seconds = int(time_elapsed.total_seconds())
                days = total_seconds // 86400
                hours = (total_seconds % 86400) // 3600
                minutes = (total_seconds % 3600) // 60
                
                # Build time display string with appropriate units
                time_parts = []
                if days > 0:
                    time_parts.append(f"{days} day{'s' if days != 1 else ''}")
                if hours > 0:
                    time_parts.append(f"{hours} hour{'s' if hours != 1 else ''}")
                if minutes > 0 or len(time_parts) == 0:  # Show minutes if no days/hours, or if there are minutes
                    time_parts.append(f"{minutes} minute{'s' if minutes != 1 else ''}")
                
                time_display = " and ".join(time_parts)
                
                overdue_requests.append({
                    'request': request,
                    'time_elapsed': time_elapsed,
                    'time_display': time_display,
                    'threshold': alert_threshold
                })
        
        # Sort by time elapsed (most overdue first)
        overdue_requests.sort(key=lambda x: x['time_elapsed'], reverse=True)
        
        return overdue_requests
        
    except Exception as e:
        print(f"Error getting overdue requests: {e}")
        import traceback
        traceback.print_exc()
        return []


def check_finance_approval_timing_alerts():
    """Check for finance approval timing alerts and send notifications"""
    try:
        current_time = datetime.utcnow()
        
        # Get all requests that are pending finance approval and have started timing
        # Exclude archived requests
        pending_requests = PaymentRequest.query.filter(
            PaymentRequest.status == 'Pending Finance Approval',
            PaymentRequest.finance_approval_start_time.isnot(None),
            PaymentRequest.finance_approval_end_time.is_(None),  # Not yet completed
            PaymentRequest.is_archived == False  # Exclude archived requests
        ).all()
        
        for request in pending_requests:
            # Calculate time elapsed since finance approval started
            time_elapsed = current_time - request.finance_approval_start_time
            
            # Determine alert thresholds based on urgency
            if request.is_urgent:
                # Urgent requests: 2 hours
                alert_threshold = timedelta(hours=2)
                recurring_threshold = timedelta(hours=2)
            else:
                # Non-urgent requests: 24 hours
                alert_threshold = timedelta(hours=24)
                recurring_threshold = timedelta(hours=24)
            
            # Check if we should send an alert
            should_send_alert = False
            alert_type = None
            
            if time_elapsed >= alert_threshold:
                # Check if this is the first alert or a recurring alert
                # Look for existing timing alerts for this request
                existing_alerts = Notification.query.filter(
                    Notification.request_id == request.request_id,
                    Notification.notification_type.in_(['finance_approval_timing_alert', 'finance_approval_timing_recurring'])
                ).order_by(Notification.created_at.desc()).all()
                
                if not existing_alerts:
                    # First alert
                    should_send_alert = True
                    alert_type = 'finance_approval_timing_alert'
                else:
                    # Check if enough time has passed for a recurring alert
                    last_alert = existing_alerts[0]
                    time_since_last_alert = current_time - last_alert.created_at
                    
                    if time_since_last_alert >= recurring_threshold:
                        should_send_alert = True
                        alert_type = 'finance_approval_timing_recurring'
            
            if should_send_alert:
                # Format time elapsed for display
                hours = int(time_elapsed.total_seconds() // 3600)
                minutes = int((time_elapsed.total_seconds() % 3600) // 60)
                
                if hours > 0:
                    time_display = f"{hours} hour{'s' if hours != 1 else ''} and {minutes} minute{'s' if minutes != 1 else ''}"
                else:
                    time_display = f"{minutes} minute{'s' if minutes != 1 else ''}"
                
                # Create alert message
                urgency_text = "URGENT" if request.is_urgent else "NON-URGENT"
                threshold_text = "2 hours" if request.is_urgent else "24 hours"
                
                if alert_type == 'finance_approval_timing_alert':
                    title = f"Finance Approval Overdue - {urgency_text} Request"
                    message = f"Payment request #{request.request_id} has been pending finance approval for {time_display} (limit: {threshold_text}). Please take action immediately."
                else:  # finance_approval_timing_recurring
                    title = f"Finance Approval Still Overdue - {urgency_text} Request"
                    message = f"Payment request #{request.request_id} is still pending finance approval after {time_display} (limit: {threshold_text}). This is a recurring alert - please take action immediately."
                
                # Send notification to all Finance Admin users
                finance_admin_users = User.query.filter_by(role='Finance Admin').all()
                for user in finance_admin_users:
                    create_notification(
                        user_id=user.user_id,
                        title=title,
                        message=message,
                        notification_type=alert_type,
                        request_id=request.request_id
                    )
                
                # Log the action
                log_action(f"Finance approval timing alert sent for request #{request.request_id} - {time_display} elapsed")
                
                print(f"Sent {alert_type} for request #{request.request_id} - {time_display} elapsed")
        
        print(f"Checked {len(pending_requests)} pending finance approval requests for timing alerts")
        
    except Exception as e:
        print(f"Error checking finance approval timing alerts: {e}")


def background_scheduler():
    """Background scheduler that runs timing checks every hour"""
    while True:
        try:
            with app.app_context():
                check_finance_approval_timing_alerts()
                check_recurring_payments_due()
        except Exception as e:
            print(f"Error in background scheduler: {e}")
        
        # Sleep for 1 hour (3600 seconds)
        time.sleep(3600)


def is_payment_due_today(request, today):
    """Check if a recurring payment is due today based on its configuration"""
    if not request.recurring_interval:
        return False
    
    try:
        # Parse the recurring interval configuration
        parts = request.recurring_interval.split(':')
        frequency = parts[0]
        interval = int(parts[1])
        
        if frequency == 'daily':
            # Daily payments - check if it's been the right number of days AND time has passed
            days_since_created = (today - request.date).days
            
            # Parse the time from the recurring interval
            time_parts = request.recurring_interval.split(':')
            if len(time_parts) >= 5 and time_parts[2] == 'time':
                try:
                    scheduled_hour = int(time_parts[3])
                    scheduled_minute = int(time_parts[4])
                    
                    # Use local time instead of UTC
                    from datetime import datetime
                    current_time = datetime.now()
                    scheduled_time = datetime.combine(today, datetime.min.time().replace(hour=scheduled_hour, minute=scheduled_minute))
                    
                    
                    # Check if it's the right day and time has passed
                    if interval == 1:
                        # Every day - due if it's the same day or later AND time has passed
                        if days_since_created >= 0:
                            return current_time >= scheduled_time
                    else:
                        # Every X days - due if it's been X days since creation AND time has passed
                        if days_since_created > 0 and days_since_created % interval == 0:
                            return current_time >= scheduled_time
                except (ValueError, IndexError):
                    # If time parsing fails, fall back to day-only logic
                    pass
            
            # Fallback to day-only logic if time parsing fails
            if interval == 1:
                return days_since_created >= 0
            else:
                return days_since_created > 0 and days_since_created % interval == 0
            
        elif frequency == 'weekly':
            # Weekly payments - check if it's the right day of week
            if len(parts) > 2 and parts[2] == 'days':
                selected_days = parts[3].split(',')
                weekday_names = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']
                today_weekday = weekday_names[today.weekday()]
                return today_weekday in selected_days
            else:
                # Default to same day of week as created
                return today.weekday() == request.date.weekday()
                
        elif frequency == 'monthly':
            # Monthly payments - check if it's the right day of month and enough months have passed
            if len(parts) > 2 and parts[2] == 'days':
                selected_days = [int(d) for d in parts[3].split(',')]
                if today.day in selected_days:
                    # Check if enough months have passed since creation
                    months_since_created = (today.year - request.date.year) * 12 + (today.month - request.date.month)
                    # Always require at least 1 month to pass (never due on creation day)
                    return months_since_created > 0 and months_since_created % interval == 0
            else:
                # Default to same day of month as created
                if today.day == request.date.day:
                    months_since_created = (today.year - request.date.year) * 12 + (today.month - request.date.month)
                    # Always require at least 1 month to pass (never due on creation day)
                    return months_since_created > 0 and months_since_created % interval == 0
            return False
                
        elif frequency == 'quarterly':
            # Quarterly payments - check if it's the right month and day
            if len(parts) > 2 and parts[2] == 'months':
                # Convert month names to numbers
                month_names = ['January', 'February', 'March', 'April', 'May', 'June',
                             'July', 'August', 'September', 'October', 'November', 'December']
                selected_month_names = [m.strip() for m in parts[3].split(',')]
                selected_months = [month_names.index(m) + 1 for m in selected_month_names if m in month_names]
                
                if today.month in selected_months:
                    if len(parts) > 4 and parts[4] == 'days':
                        selected_days = [int(d) for d in parts[5].split(',')]
                        if today.day in selected_days:
                            # Check if enough quarters have passed
                            months_since_created = (today.year - request.date.year) * 12 + (today.month - request.date.month)
                            return months_since_created > 0 and months_since_created % (interval * 3) == 0
                    else:
                        if today.day == request.date.day:
                            months_since_created = (today.year - request.date.year) * 12 + (today.month - request.date.month)
                            return months_since_created > 0 and months_since_created % (interval * 3) == 0
            return False
            
        elif frequency == 'yearly':
            # Yearly payments - check if it's the right month and day
            if len(parts) > 2 and parts[2] == 'months':
                # Convert month names to numbers
                month_names = ['January', 'February', 'March', 'April', 'May', 'June',
                             'July', 'August', 'September', 'October', 'November', 'December']
                selected_month_names = [m.strip() for m in parts[3].split(',')]
                selected_months = [month_names.index(m) + 1 for m in selected_month_names if m in month_names]
                
                if today.month in selected_months:
                    if len(parts) > 4 and parts[4] == 'days':
                        selected_days = [int(d) for d in parts[5].split(',')]
                        if today.day in selected_days:
                            # Check if enough years have passed
                            years_since_created = today.year - request.date.year
                            return years_since_created > 0 and years_since_created % interval == 0
                    else:
                        if today.day == request.date.day:
                            years_since_created = today.year - request.date.year
                            return years_since_created > 0 and years_since_created % interval == 0
            return False
            
    except (ValueError, IndexError, AttributeError):
        # If parsing fails, don't create notification
        return False
    
    return False


def calculate_finance_approval_duration(request):
    """Calculate and set finance approval duration if not already set"""
    if (request.finance_approval_start_time and 
        request.finance_approval_end_time and 
        not request.finance_approval_duration_minutes):
        duration = request.finance_approval_end_time - request.finance_approval_start_time
        # Store duration in seconds for more precision (consistent with template logic)
        request.finance_approval_duration_minutes = int(duration.total_seconds())
        return True
    return False

def check_recurring_payment_completion(request_id):
    """Check if all installments for a recurring payment are paid and mark as completed"""
    try:
        # Get the payment request
        req = PaymentRequest.query.get(request_id)
        if not req or req.status != 'Recurring' or req.recurring != 'Recurring':
            return False
        
        # Get all installments for this request
        installments = RecurringPaymentSchedule.query.filter_by(request_id=request_id).all()
        
        if not installments:
            return False
        
        # Check if all installments are paid
        all_paid = all(installment.is_paid for installment in installments)
        
        if all_paid:
            # Calculate total paid amount
            total_paid = sum(float(installment.amount) for installment in installments)
            total_requested = float(req.amount)
            
            # Check if there's no remaining amount (allowing for small floating point differences)
            remaining_amount = total_requested - total_paid
            if abs(remaining_amount) < 0.001:  # Less than 0.001 OMR difference
                # Mark the request as completed
                req.status = 'Completed'
                req.completion_date = datetime.utcnow().date()
                req.approval_date = datetime.utcnow().date()  # Set approval_date when status becomes Completed
                req.updated_at = datetime.utcnow()
                
                # End finance approval timing when completed
                if req.finance_approval_start_time and not req.finance_approval_end_time:
                    current_time = datetime.utcnow()
                    req.finance_approval_end_time = current_time
                    duration = current_time - req.finance_approval_start_time
                    req.finance_approval_duration_minutes = int(duration.total_seconds())
                
                db.session.commit()
                
                # Create completion notification
                create_notification(
                    user_id=req.user_id,
                    title="Recurring Payment Completed",
                    message=f'All installments for recurring payment request #{request_id} have been paid. Request marked as completed.',
                    notification_type="recurring_completed",
                    request_id=request_id
                )
                
                # Notify Auditing Staff and Auditing Department Manager
                auditing_users = User.query.filter(
                    db.and_(
                        User.department == 'Auditing',
                        User.role.in_(['Auditing Staff', 'Department Manager'])
                    )
                ).all()
                for auditing_user in auditing_users:
                    create_notification(
                        user_id=auditing_user.user_id,
                        title="Recurring Payment Completed",
                        message=f'All installments for recurring payment request #{request_id} from {req.department} department have been paid. Request marked as completed.',
                        notification_type="recurring_completed",
                        request_id=request_id
                    )
                
                # Log the action
                log_action(f"Recurring payment request #{request_id} automatically marked as completed - all installments paid")
                
                # Emit real-time update
                socketio.emit('request_updated', {
                    'request_id': request_id,
                    'status': 'Completed',
                    'recurring': True,
                    'completed': True
                })
                
                return True
        
        return False
        
    except Exception as e:
        print(f"Error checking recurring payment completion: {e}")
        return False


def get_notifications_for_user(user, limit=5, page=None, per_page=None):
    """Get notifications for a user based on RBAC.
    - If limit is provided, return limited results (for dropdowns)
    - If page and per_page are provided, return paginated results
    - Otherwise return all results
    """

    query = None

    if user.role == 'Project Staff':
        # Project Staff: Updates on their own requests + recurring payment due on their own requests only
        query = Notification.query.filter(
            db.and_(
                Notification.user_id == user.user_id,
                db.or_(
                    Notification.notification_type == 'recurring_due',
                    Notification.notification_type.in_([
                        'request_rejected', 'request_approved', 'proof_uploaded', 'proof_rejected',
                        'status_changed', 'proof_required', 'recurring_approved', 'request_completed',
                        'installment_paid', 'finance_note_added', 'one_time_payment_scheduled'
                    ])
                )
            )
        ).order_by(Notification.created_at.desc())

    elif user.role in ['Finance Staff', 'Finance Admin']:
        # Finance roles: New submissions when requests reach Pending Finance Approval + proof uploaded + recurring payment due + system-wide
        # Finance Staff additionally get updates on their own requests
        if user.role == 'Finance Staff':
            query = Notification.query.filter(
                db.and_(
                    Notification.user_id == user.user_id,
                    db.or_(
                        Notification.notification_type.in_([
                            'ready_for_finance_review', 'proof_uploaded', 'recurring_due', 'installment_edited',
                            'finance_approval_timing_alert', 'finance_approval_timing_recurring',
                            'system_maintenance', 'system_update', 'security_alert', 'system_error',
                            'admin_announcement'
                        ]),
                        Notification.notification_type.in_([
                            'request_rejected', 'request_approved', 'proof_uploaded', 'proof_rejected',
                            'status_changed', 'proof_required', 'recurring_approved', 'request_completed',
                            'installment_paid', 'finance_note_added', 'one_time_payment_scheduled'
                        ])
                    )
                )
            ).order_by(Notification.created_at.desc())
        elif user.name == 'Abdalaziz Al-Brashdi':
            # Abdalaziz gets finance notifications + new_submission from GM/Operation Manager/Finance Staff + updates on Finance Staff, GM, and Operation Manager requests
            query = Notification.query.filter(
                db.and_(
                    Notification.user_id == user.user_id,
                    db.or_(
                        Notification.notification_type.in_([
                            'ready_for_finance_review', 'proof_uploaded', 'recurring_due', 'installment_edited',
                            'finance_approval_timing_alert', 'finance_approval_timing_recurring',
                            'system_maintenance', 'system_update', 'security_alert', 'system_error',
                            'admin_announcement'
                        ]),
                        Notification.notification_type.in_([
                            'request_rejected', 'request_approved', 'proof_uploaded', 'proof_rejected',
                            'status_changed', 'proof_required', 'recurring_approved', 'request_completed',
                            'installment_paid', 'finance_note_added', 'one_time_payment_scheduled'
                        ]),
                        # new_submission only for GM, Operation Manager, Finance Staff, or CEO (his assigned managers)
                        db.and_(
                            Notification.notification_type == 'new_submission',
                            Notification.request_id.isnot(None),
                            db.exists().where(
                                db.and_(
                                    PaymentRequest.request_id == Notification.request_id,
                                    PaymentRequest.user_id == User.user_id,
                                    User.role.in_(['GM', 'CEO', 'Operation Manager', 'Finance Staff'])
                                )
                            )
                        )
                    )
                )
            ).order_by(Notification.created_at.desc())
        else:  # Other Finance Admin
            query = Notification.query.filter(
                db.and_(
                    Notification.user_id == user.user_id,
                    Notification.notification_type.in_([
                        'ready_for_finance_review', 'proof_uploaded', 'recurring_due', 'installment_edited',
                        'finance_approval_timing_alert', 'finance_approval_timing_recurring',
                        'system_maintenance', 'system_update', 'security_alert', 'system_error',
                        'admin_announcement', 'one_time_payment_scheduled'
                    ])
                )
            ).order_by(Notification.created_at.desc())

    elif user.role in ['GM', 'CEO']:
        # GM: New submissions from ALL requests (all roles/departments) + updates on their own requests + system-wide + temporary manager assignments
        query = Notification.query.filter(
            db.and_(
                Notification.user_id == user.user_id,
                db.or_(
                    Notification.notification_type == 'new_submission',
                    Notification.notification_type.in_([
                        'request_rejected', 'request_approved', 'proof_uploaded', 'proof_rejected',
                        'status_changed', 'proof_required', 'recurring_approved', 'request_completed',
                        'installment_paid', 'finance_note_added', 'one_time_payment_scheduled'
                    ]),
                    Notification.notification_type.in_([
                        'system_maintenance', 'system_update', 'security_alert', 'system_error',
                        'admin_announcement'
                    ]),
                    Notification.notification_type == 'temporary_manager_assignment'
                )
            )
        ).order_by(Notification.created_at.desc())

    elif user.role == 'Operation Manager':
        # Operation Manager: New submissions from ALL requests (all roles/departments) + updates on their own requests + system-wide
        query = Notification.query.filter(
            db.and_(
                Notification.user_id == user.user_id,
                db.or_(
                    Notification.notification_type == 'new_submission',
                    Notification.notification_type.in_([
                        'request_rejected', 'request_approved', 'proof_uploaded', 'proof_rejected',
                        'status_changed', 'proof_required', 'recurring_approved', 'request_completed',
                        'installment_paid', 'one_time_payment_scheduled'
                    ]),
                    Notification.notification_type.in_([
                        'system_maintenance', 'system_update', 'security_alert', 'system_error',
                        'admin_announcement'
                    ]),
                    Notification.notification_type == 'temporary_manager_assignment'
                )
            )
        ).order_by(Notification.created_at.desc())

    elif user.role == 'Department Manager' and user.department == 'IT':
        # IT Department Manager: New submissions from IT Staff only + updates on their own requests + system-wide + user management + temporary manager assignments
        query = Notification.query.filter(
            db.and_(
                Notification.user_id == user.user_id,
                db.or_(
                    # new_submission only if it's from IT Staff (join with PaymentRequest to check)
                    db.and_(
                        Notification.notification_type == 'new_submission',
                        Notification.request_id.isnot(None),
                        db.or_(
                            # Request created by IT Staff (check user role)
                            db.exists().where(
                                db.and_(
                                    PaymentRequest.request_id == Notification.request_id,
                                    PaymentRequest.department == 'IT',
                                    User.user_id == PaymentRequest.user_id,
                                    User.role == 'IT Staff'
                                )
                            ),
                            # Or request from IT department with IT Staff role in message (fallback for existing notifications)
                            Notification.message.contains('IT')
                        )
                    ),
                    Notification.notification_type.in_([
                        'request_rejected', 'request_approved', 'proof_uploaded', 'status_changed',
                        'proof_required', 'recurring_approved', 'request_completed', 'installment_paid',
                        'user_created', 'user_updated', 'user_deleted', 'finance_note_added',
                        'request_archived', 'request_restored', 'request_permanently_deleted', 'one_time_payment_scheduled'
                    ]),
                    Notification.notification_type.in_([
                        'system_maintenance', 'system_update', 'security_alert', 'system_error',
                        'admin_announcement'
                    ]),
                    Notification.notification_type == 'temporary_manager_assignment'
                )
            )
        ).order_by(Notification.created_at.desc())

    elif user.role == 'IT Staff':
        # IT Staff: Updates on their own requests + system-wide + user management + request archives
        query = Notification.query.filter(
            db.and_(
                Notification.user_id == user.user_id,
                db.or_(
                    Notification.notification_type.in_([
                        'request_rejected', 'request_approved', 'proof_uploaded', 'status_changed',
                        'proof_required', 'recurring_approved', 'request_completed', 'installment_paid',
                        'user_created', 'user_updated', 'user_deleted', 'finance_note_added',
                        'request_archived', 'request_restored', 'request_permanently_deleted', 'one_time_payment_scheduled'
                    ]),
                    Notification.notification_type.in_([
                        'system_maintenance', 'system_update', 'security_alert', 'system_error',
                        'admin_announcement'
                    ])
                )
            )
        ).order_by(Notification.created_at.desc())

    elif user.role == 'Department Manager':
        # Other Department Managers: New submissions from their own department staff only + recurring payment due for their department + updates on their own requests + temporary manager assignments
        print(f"DEBUG: Getting notifications for Department Manager {user.username} from {user.department}")

        # Get all notifications for this user first
        all_user_notifications = Notification.query.filter_by(user_id=user.user_id).all()
        print(f"DEBUG: Total notifications for user {user.username}: {len(all_user_notifications)}")
        for notif in all_user_notifications:
            print(f"DEBUG: Notification {notif.notification_id}: {notif.notification_type} - {notif.title}")

        query = Notification.query.filter(
            db.and_(
                Notification.user_id == user.user_id,
                db.or_(
                    Notification.notification_type == 'new_submission',
                    Notification.notification_type == 'recurring_due',
                    Notification.notification_type.in_([
                        'request_rejected', 'request_approved', 'proof_uploaded', 'proof_rejected',
                        'status_changed', 'proof_required', 'recurring_approved', 'request_completed',
                        'installment_paid', 'finance_note_added', 'one_time_payment_scheduled'
                    ]),
                    Notification.notification_type == 'temporary_manager_assignment'
                )
            )
        ).order_by(Notification.created_at.desc())

    else:
        # Department Staff: Updates on their own requests only + recurring payment due for their own requests
        query = Notification.query.filter(
            db.and_(
                Notification.user_id == user.user_id,
                Notification.notification_type.in_([
                    'request_rejected', 'request_approved', 'proof_uploaded', 'proof_rejected',
                    'status_changed', 'recurring_due', 'proof_required', 'recurring_approved',
                    'request_completed', 'installment_paid', 'finance_note_added', 'one_time_payment_scheduled'
                ])
            )
        ).order_by(Notification.created_at.desc())

    # Handle pagination, limit, or return all
    if page is not None and per_page is not None:
        return query.paginate(page=page, per_page=per_page, error_out=False)
    elif limit:
        return query.limit(limit).all()
    return query.all()

def get_unread_count_for_user(user):
    """Get unread notification count for a user based on their role per RBAC"""
    
    if user.role == 'Project Staff':
        # Project Staff: Updates on their own requests + recurring payment due on their own requests only
        return Notification.query.filter(
            db.and_(
                Notification.user_id == user.user_id,
                Notification.is_read == False,
                db.or_(
                    Notification.notification_type == 'recurring_due',
                    Notification.notification_type.in_(['request_rejected', 'request_approved', 'proof_uploaded', 'proof_rejected', 'status_changed', 'proof_required', 'recurring_approved', 'request_completed', 'installment_paid'])
                )
            )
        ).count()
    
    elif user.role in ['Finance Staff', 'Finance Admin']:
        # Finance roles: New submissions when requests reach Pending Finance Approval + proof uploaded + recurring payment due + system-wide
        # Finance Staff additionally get updates on their own requests
        if user.role == 'Finance Staff':
            return Notification.query.filter(
                db.and_(
                    Notification.user_id == user.user_id,
                    Notification.is_read == False,
                    db.or_(
                        Notification.notification_type.in_(['ready_for_finance_review', 'proof_uploaded', 'recurring_due', 'installment_edited', 'finance_approval_timing_alert', 'finance_approval_timing_recurring', 'system_maintenance', 'system_update', 'security_alert', 'system_error', 'admin_announcement']),
                        Notification.notification_type.in_(['request_rejected', 'request_approved', 'proof_uploaded', 'proof_rejected', 'status_changed', 'proof_required', 'recurring_approved', 'request_completed', 'installment_paid'])
                    )
                )
            ).count()
        elif user.name == 'Abdalaziz Al-Brashdi':
            # Abdalaziz gets finance notifications + new_submission from GM/Operation Manager/Finance Staff + updates on Finance Staff, GM, and Operation Manager requests
            return Notification.query.filter(
                db.and_(
                    Notification.user_id == user.user_id,
                    Notification.is_read == False,
                    db.or_(
                        Notification.notification_type.in_(['ready_for_finance_review', 'proof_uploaded', 'recurring_due', 'installment_edited', 'finance_approval_timing_alert', 'finance_approval_timing_recurring', 'system_maintenance', 'system_update', 'security_alert', 'system_error', 'admin_announcement']),
                        Notification.notification_type.in_(['request_rejected', 'request_approved', 'proof_uploaded', 'proof_rejected', 'status_changed', 'proof_required', 'recurring_approved', 'request_completed', 'installment_paid']),
                        # new_submission only for GM, Operation Manager, Finance Staff, or CEO (his assigned managers)
                        db.and_(
                            Notification.notification_type == 'new_submission',
                            Notification.request_id.isnot(None),
                            db.exists().where(
                                db.and_(
                                    PaymentRequest.request_id == Notification.request_id,
                                    PaymentRequest.user_id == User.user_id,
                                    User.role.in_(['GM', 'CEO', 'Operation Manager', 'Finance Staff'])
                                )
                            )
                        )
                    )
                )
            ).count()
        else:  # Other Finance Admin
            return Notification.query.filter(
                db.and_(
                    Notification.user_id == user.user_id,
                    Notification.is_read == False,
                    db.or_(
                        Notification.notification_type.in_(['ready_for_finance_review', 'proof_uploaded', 'recurring_due', 'installment_edited', 'finance_approval_timing_alert', 'finance_approval_timing_recurring', 'system_maintenance', 'system_update', 'security_alert', 'system_error', 'admin_announcement']),
                        Notification.notification_type.in_(['request_rejected', 'request_approved', 'proof_uploaded', 'proof_rejected', 'status_changed', 'proof_required', 'recurring_approved', 'request_completed', 'installment_paid'])
                    )
                )
            ).count()
    
    elif user.role in ['GM', 'CEO']:
        # GM: New submissions from ALL requests (all roles/departments) + updates on their own requests + system-wide
        return Notification.query.filter(
            db.and_(
                Notification.user_id == user.user_id,
                Notification.is_read == False,
                db.or_(
                    Notification.notification_type == 'new_submission',
                    Notification.notification_type.in_(['request_rejected', 'request_approved', 'proof_uploaded', 'proof_rejected', 'status_changed', 'proof_required', 'recurring_approved', 'request_completed', 'installment_paid', 'finance_note_added', 'one_time_payment_scheduled']),
                    Notification.notification_type.in_(['system_maintenance', 'system_update', 'security_alert', 'system_error', 'admin_announcement']),
                    Notification.notification_type == 'temporary_manager_assignment'
                )
            )
        ).count()
    
    elif user.role == 'Operation Manager':
        # Operation Manager: New submissions from ALL requests (all roles/departments) + updates on their own requests + system-wide
        return Notification.query.filter(
            db.and_(
                Notification.user_id == user.user_id,
                Notification.is_read == False,
                db.or_(
                    Notification.notification_type == 'new_submission',
                    Notification.notification_type.in_(['request_rejected', 'request_approved', 'proof_uploaded', 'proof_rejected', 'status_changed', 'proof_required', 'recurring_approved', 'request_completed', 'installment_paid']),
                    Notification.notification_type.in_(['system_maintenance', 'system_update', 'security_alert', 'system_error', 'admin_announcement']),
                    Notification.notification_type == 'temporary_manager_assignment'
                )
            )
        ).count()
    
    elif user.role == 'Department Manager' and user.department == 'IT':
        # IT Department Manager: New submissions from IT Staff only + updates on their own requests + system-wide + user management + temporary manager assignments
        return Notification.query.filter(
            db.and_(
                Notification.user_id == user.user_id,
                Notification.is_read == False,
                db.or_(
                    # new_submission only if it's from IT Staff (join with PaymentRequest to check)
                    db.and_(
                        Notification.notification_type == 'new_submission',
                        Notification.request_id.isnot(None),
                        db.or_(
                            # Request created by IT Staff (check user role)
                            db.exists().where(
                                db.and_(
                                    PaymentRequest.request_id == Notification.request_id,
                                    PaymentRequest.department == 'IT',
                                    User.user_id == PaymentRequest.user_id,
                                    User.role == 'IT Staff'
                                )
                            ),
                            # Or request from IT department (fallback for existing notifications)
                            Notification.message.contains('IT')
                        )
                    ),
                    Notification.notification_type.in_(['request_rejected', 'request_approved', 'proof_uploaded', 'status_changed', 'proof_required', 'recurring_approved', 'request_completed', 'installment_paid', 'user_created', 'user_updated', 'user_deleted', 'finance_note_added', 'request_archived', 'request_restored', 'request_permanently_deleted', 'one_time_payment_scheduled']),
                    Notification.notification_type.in_(['system_maintenance', 'system_update', 'security_alert', 'system_error', 'admin_announcement']),
                    Notification.notification_type == 'temporary_manager_assignment'
                )
            )
        ).count()
    
    elif user.role == 'IT Staff':
        # IT Staff: Updates on their own requests + system-wide + user management + request archives
        return Notification.query.filter(
            db.and_(
                Notification.user_id == user.user_id,
                Notification.is_read == False,
                db.or_(
                    Notification.notification_type.in_(['request_rejected', 'request_approved', 'proof_uploaded', 'status_changed', 'proof_required', 'recurring_approved', 'request_completed', 'installment_paid', 'user_created', 'user_updated', 'user_deleted', 'finance_note_added', 'request_archived', 'request_restored', 'request_permanently_deleted', 'one_time_payment_scheduled']),
                    Notification.notification_type.in_(['system_maintenance', 'system_update', 'security_alert', 'system_error', 'admin_announcement'])
                )
            )
        ).count()
    
    elif user.role == 'Department Manager':
        # Other Department Managers: New submissions from their own department staff only + recurring payment due for their department + updates on their own requests
        return Notification.query.filter(
            db.and_(
                Notification.user_id == user.user_id,
                Notification.is_read == False,
                db.or_(
                    Notification.notification_type == 'new_submission',  # Simplified - same as get_notifications_for_user
                    Notification.notification_type == 'recurring_due',
                    Notification.notification_type.in_(['request_rejected', 'request_approved', 'proof_uploaded', 'proof_rejected', 'status_changed', 'proof_required', 'recurring_approved', 'request_completed', 'installment_paid']),
                    Notification.notification_type == 'temporary_manager_assignment'
                )
            )
        ).count()
    
    else:
        # Department Staff: Updates on their own requests only + recurring payment due for their own requests
        return Notification.query.filter(
            db.and_(
                Notification.user_id == user.user_id,
                Notification.is_read == False,
                Notification.notification_type.in_(['request_rejected', 'request_approved', 'proof_uploaded', 'proof_rejected', 'status_changed', 'recurring_due', 'proof_required', 'recurring_approved', 'request_completed', 'installment_paid', 'one_time_payment_scheduled'])
            )
        ).count()

def notify_finance_and_admin(title, message, notification_type, request_id=None):
    """Notify Finance and Admin users about new submissions"""
    # Get all Finance and Admin users
    finance_admin_users = User.query.filter(User.role.in_(['Finance Staff', 'Finance Admin'])).all()
    
    for user in finance_admin_users:
        create_notification(
            user_id=user.user_id,
            title=title,
            message=message,
            notification_type=notification_type,
            request_id=request_id
        )
    
    # Emit real-time notification to all users
    socketio.emit('new_notification', {
        'title': title,
        'message': message,
        'type': notification_type,
        'request_id': request_id
    }, room='all_users')
    
    # Also emit a general update event to trigger notification count updates
    socketio.emit('notification_update', {
        'action': 'new_notification',
        'type': notification_type
    }, room='all_users')
    
    # Emit to specific role rooms based on notification type
    if notification_type == 'ready_for_finance_review':
        socketio.emit('notification_update', {
            'action': 'new_notification',
            'type': notification_type
        }, room='finance_admin')
    elif notification_type == 'new_submission':
        # Emit to appropriate role rooms based on requestor role
        if request and hasattr(request, 'user'):
            requestor_role = request.user.role
            if requestor_role == 'Department Manager':
                socketio.emit('notification_update', {
                    'action': 'new_notification',
                    'type': notification_type
                }, room='gm')
            elif requestor_role == 'Operation Staff':
                socketio.emit('notification_update', {
                    'action': 'new_notification',
                    'type': notification_type
                }, room='operation_manager')
            elif requestor_role == 'IT Staff':
                socketio.emit('notification_update', {
                    'action': 'new_notification',
                    'type': notification_type
                }, room='it_staff')
            elif requestor_role.endswith(' Staff'):
                socketio.emit('notification_update', {
                    'action': 'new_notification',
                    'type': notification_type
                }, room='department_managers')


def emit_request_update_to_all_rooms(event_name, data):
    """Helper function to emit request updates to all relevant rooms"""
    try:
        # Emit to all users
        socketio.emit(event_name, data, room='all_users')
        
        # Emit to specific role rooms
        socketio.emit(event_name, data, room='finance_admin')
        socketio.emit(event_name, data, room='department_staff')
        socketio.emit(event_name, data, room='department_managers')
        socketio.emit(event_name, data, room='gm')
        socketio.emit(event_name, data, room='operation_manager')
        socketio.emit(event_name, data, room='it_staff')
        socketio.emit(event_name, data, room='project_staff')
        
        print(f"DEBUG: Emitted {event_name} to all relevant rooms")
    except Exception as e:
        print(f"Error emitting {event_name} to all rooms: {e}")

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


# ==================== AUTHENTICATION ROUTES ====================

@app.route('/')
def index():
    """Redirect to login or dashboard"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/verify_pin', methods=['POST'])
def verify_pin():
    """Verify PIN via AJAX - returns JSON response"""
    try:
        data = request.get_json()
        username = data.get('username')
        pin = data.get('pin')
        
        if not username or not pin:
            return jsonify({
                'success': False,
                'message': 'Username and PIN are required.'
            })
        
        user = User.query.filter_by(username=username).first()
        
        if not user:
            return jsonify({
                'success': False,
                'message': 'User not found.'
            })
        
        # Check if account is locked
        if user.is_account_locked():
            return jsonify({
                'success': False,
                'message': 'Your account has been locked due to too many failed login attempts. Please contact IT Staff to unlock your account.',
                'account_locked': True
            })
        
        # Check if PIN exists in session
        stored_pin = session.get('temp_login_pin')
        stored_username = session.get('temp_pin_username')
        pin_expires = session.get('temp_pin_expires')
        
        if not stored_pin or stored_username != username:
            return jsonify({
                'success': False,
                'message': 'No PIN generated or session expired. Please try logging in again.',
                'session_expired': True
            })
        
        # Check if PIN is expired
        if datetime.utcnow().timestamp() > pin_expires:
            # Clear expired session data
            session.pop('temp_login_pin', None)
            session.pop('temp_pin_username', None)
            session.pop('temp_pin_expires', None)
            return jsonify({
                'success': False,
                'message': 'PIN has expired. Please request a new one by logging in again.',
                'pin_expired': True
            })
        
        # Verify PIN
        if check_password_hash(stored_pin, pin):
            # Clear session PIN data
            session.pop('temp_login_pin', None)
            session.pop('temp_pin_username', None)
            session.pop('temp_pin_expires', None)
            
            # Reset failed login attempts on successful login
            user.reset_failed_login()
            login_user(user, remember=False)  # Session-only: expires when browser closes
            # Set activity timestamp and session start time (session is non-permanent, expires on browser close)
            # Generate unique tab session ID for tab-specific session tracking
            try:
                session.permanent = False  # Session cookie expires when browser closes
                now_ts = datetime.utcnow().timestamp()
                session['last_activity'] = now_ts
                session['session_start'] = now_ts  # Track when this session started
                # Generate unique tab session ID
                tab_session_id = f"{random.randint(100000, 999999)}{now_ts}{random.randint(100000, 999999)}"
                session['tab_session_id'] = tab_session_id
            except Exception:
                pass
            log_action(f"User {username} logged in successfully with email PIN")
            
            return jsonify({
                'success': True,
                'message': f'Welcome back, {user.name}!',
                'redirect_url': url_for('dashboard'),
                'tab_session_id': session.get('tab_session_id')
            })
        else:
            # Invalid PIN - increment failed attempts
            user.increment_failed_login()
            remaining_attempts = 5 - user.failed_login_attempts
            
            if user.is_account_locked():
                # Clear session data
                session.pop('temp_login_pin', None)
                session.pop('temp_pin_username', None)
                session.pop('temp_pin_expires', None)
                log_action(f"Account locked due to failed PIN attempts: {username}")
                return jsonify({
                    'success': False,
                    'message': 'Too many failed login attempts. Your account has been locked. Please contact IT Staff to unlock your account.',
                    'account_locked': True
                })
            else:
                log_action(f"Failed PIN attempt for user: {username} ({remaining_attempts} attempts remaining)")
                return jsonify({
                    'success': False,
                    'message': f'Wrong PIN. You have {remaining_attempts} attempt(s) remaining before your account is locked.',
                    'remaining_attempts': remaining_attempts
                })
    
    except Exception as e:
        app.logger.error(f"PIN verification error: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': 'An error occurred. Please try again.'
        })


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page"""
    # Ensure a default IT user exists if database is empty
    try:
        if User.query.count() == 0:
            default_it = User(
                name='Default IT',
                username='it@system.local',
                department='IT',
                role='IT Staff',
                email='it@system.local'
            )
            default_it.set_password('admin123')
            db.session.add(default_it)
            db.session.commit()
            app.logger.info('Default IT user created: it@system.local / admin123')
    except Exception as _e:
        # If DB not ready or other issue, continue to login page without blocking
        pass
    # If user is authenticated but session doesn't have last_activity or session_start, it's a stale session
    # Clear it and force re-login
    if current_user.is_authenticated:
        if 'last_activity' not in session or 'session_start' not in session:
            try:
                logout_user()
                # Clear only the session keys we care about, not the entire session
                # This prevents issues with Flask's session initialization
                session.pop('last_activity', None)
                session.pop('session_start', None)
                session.pop('user_id', None)  # Flask-Login's session key
                session.pop('_permanent', None)  # Flask's permanent session flag
            except Exception as e:
                app.logger.error(f"Error clearing stale session: {str(e)}", exc_info=True)
            flash('Your session has expired. Please log in again.', 'info')
        else:
            return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        pin = request.form.get('pin')
        
        # Debug logging
        app.logger.info(f"Login attempt: username={username}, pin_provided={bool(pin)}")
        
        user = User.query.filter_by(username=username).first()
        
        # Debug user found
        app.logger.info(f"User found: {user is not None}, user_id: {user.user_id if user else 'None'}")
        
        # Check if account is locked
        if user and user.is_account_locked():
            flash('Your account has been locked due to too many failed login attempts. Please contact IT Staff to unlock your account.', 'danger')
            log_action(f"Login attempt for locked account: {username}")
            return render_template('login.html')
        
        # Verify password and temporary PIN
        if user and user.check_password(password):
            # Special case: IT system account and test admin bypass PIN requirement
            if username in ['it@system.local', 'testadmin@maagroup.om']:
                app.logger.info(f"System account bypass triggered for {username}")
                # Reset failed login attempts on successful login
                user.reset_failed_login()
                login_user(user, remember=False)  # Session-only: expires when browser closes
                # Set activity timestamp and session start time (session is non-permanent, expires on browser close)
                # Generate unique tab session ID for tab-specific session tracking
                try:
                    session.permanent = False  # Session cookie expires when browser closes
                    now_ts = datetime.utcnow().timestamp()
                    session['last_activity'] = now_ts
                    session['session_start'] = now_ts  # Track when this session started
                    # Generate unique tab session ID
                    tab_session_id = f"{random.randint(100000, 999999)}{now_ts}{random.randint(100000, 999999)}"
                    session['tab_session_id'] = tab_session_id
                except Exception:
                    pass
                app.logger.info(f"System user logged in successfully, redirecting to dashboard")
                log_action(f"System account {username} logged in successfully (PIN bypassed)")
                flash(f'Welcome back, {user.name}!', 'success')
                return redirect(url_for('dashboard'))
            
            # Check if PIN exists in session
            stored_pin = session.get('temp_login_pin')
            stored_username = session.get('temp_pin_username')
            pin_expires = session.get('temp_pin_expires')
            
            if not stored_pin or stored_username != username:
                flash('No PIN generated or session expired. Please try logging in again.', 'danger')
                return render_template('login.html')
            
            # Check if PIN is expired
            if datetime.utcnow().timestamp() > pin_expires:
                # Clear expired session data
                session.pop('temp_login_pin', None)
                session.pop('temp_pin_username', None)
                session.pop('temp_pin_expires', None)
                flash('PIN has expired. Please request a new one by logging in again.', 'danger')
                return render_template('login.html')
            
            # Verify PIN
            if check_password_hash(stored_pin, pin):
                # Clear session PIN data
                session.pop('temp_login_pin', None)
                session.pop('temp_pin_username', None)
                session.pop('temp_pin_expires', None)
                
                # Reset failed login attempts on successful login
                user.reset_failed_login()
                login_user(user, remember=False)  # Session-only: expires when browser closes
                # Set activity timestamp and session start time (session is non-permanent, expires on browser close)
                # Generate unique tab session ID for tab-specific session tracking
                try:
                    session.permanent = False  # Session cookie expires when browser closes
                    now_ts = datetime.utcnow().timestamp()
                    session['last_activity'] = now_ts
                    session['session_start'] = now_ts  # Track when this session started
                    # Generate unique tab session ID
                    tab_session_id = f"{random.randint(100000, 999999)}{now_ts}{random.randint(100000, 999999)}"
                    session['tab_session_id'] = tab_session_id
                except Exception:
                    pass
                log_action(f"User {username} logged in successfully with email PIN")
                flash(f'Welcome back, {user.name}!', 'success')
                return redirect(url_for('dashboard'))
            else:
                # Invalid PIN - increment failed attempts but stay in modal
                user.increment_failed_login()
                remaining_attempts = 5 - user.failed_login_attempts
                
                if user.is_account_locked():
                    # Clear session data and redirect to login with error
                    session.pop('temp_login_pin', None)
                    session.pop('temp_pin_username', None)
                    session.pop('temp_pin_expires', None)
                    flash(f'Too many failed login attempts. Your account has been locked. Please contact IT Staff to unlock your account.', 'danger')
                    log_action(f"Account locked due to failed PIN attempts: {username}")
                    return render_template('login.html')
                else:
                    # Return error message that will be handled by JavaScript
                    flash(f'Invalid PIN. You have {remaining_attempts} attempt(s) remaining before your account is locked.', 'danger')
                    log_action(f"Failed PIN attempt for user: {username} ({remaining_attempts} attempts remaining)")
                    return render_template('login.html')
        else:
            # Invalid password
            if user:
                user.increment_failed_login()
                remaining_attempts = 5 - user.failed_login_attempts
                
                if user.is_account_locked():
                    flash(f'Too many failed login attempts. Your account has been locked. Please contact IT Staff to unlock your account.', 'danger')
                    log_action(f"Account locked due to failed login attempts: {username}")
                elif remaining_attempts > 0:
                    flash(f'Invalid email address or password. You have {remaining_attempts} attempt(s) remaining before your account is locked.', 'danger')
                    log_action(f"Failed login attempt for user: {username} ({remaining_attempts} attempts remaining)")
            else:
                flash('Invalid email address or password', 'danger')
                log_action(f"Failed login attempt for non-existent user: {username}")
    
    return render_template('login.html')


@app.route('/validate_credentials', methods=['POST'])
def validate_credentials():
    """Validate email and password before showing PIN modal"""
    try:
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        
        if not username or not password:
            return jsonify({
                'success': False,
                'message': 'Email and password are required.'
            })
        
        user = User.query.filter_by(username=username).first()
        
        # Check if account is locked
        if user and user.is_account_locked():
            return jsonify({
                'success': False,
                'message': 'Your account has been locked due to too many failed login attempts. Please contact IT Staff to unlock your account.'
            })
        
        # Verify password only (not PIN yet)
        if user and user.check_password(password):
            # Special case: IT system account and test admin bypass PIN requirement
            if username in ['it@system.local', 'testadmin@maagroup.om']:
                return jsonify({
                    'success': True,
                    'message': 'System account - PIN bypassed. Redirecting to dashboard.',
                    'bypass_pin': True
                })
            
            # Check if user has email
            if not user.email:
                return jsonify({
                    'success': False,
                    'message': 'Your account does not have an email address set. Please contact IT Staff.'
                })
            
            # Generate a random 4-digit PIN
            pin = str(random.randint(1000, 9999))
            
            # Ensure session is properly initialized before storing data
            try:
                # Store PIN in session (not database)
                session['temp_login_pin'] = generate_password_hash(pin)
                session['temp_pin_username'] = username
                session['temp_pin_expires'] = (datetime.utcnow() + timedelta(minutes=app.config.get('PIN_EXPIRY_MINUTES', 5))).timestamp()
                # Mark session as modified to ensure it's saved
                session.modified = True
            except Exception as session_error:
                app.logger.error(f"Error storing PIN in session: {str(session_error)}", exc_info=True)
                return jsonify({
                    'success': False,
                    'message': 'Failed to initialize login session. Please try again.'
                })
            
            # Send PIN via email
            success, message = send_pin_email(user.email, user.name, pin)
            
            if success:
                log_action(f"Login PIN sent to {username}")
                return jsonify({
                    'success': True,
                    'message': f'A 4-digit PIN has been sent to your email ({user.email}). Please check your email and enter the PIN to continue.'
                })
            else:
                # Clear session data if email fails
                session.pop('temp_login_pin', None)
                session.pop('temp_pin_username', None)
                session.pop('temp_pin_expires', None)
                return jsonify({
                    'success': False,
                    'message': f'Failed to send PIN to your email. Please try again or contact IT Staff. Error: {message}'
                })
        else:
            # Invalid credentials - increment failed attempts
            if user:
                user.increment_failed_login()
                remaining_attempts = 5 - user.failed_login_attempts
                
                if user.is_account_locked():
                    log_action(f"Account locked due to failed login attempts: {username}")
                    return jsonify({
                        'success': False,
                        'message': 'Too many failed login attempts. Your account has been locked. Please contact IT Staff to unlock your account.'
                    })
                elif remaining_attempts > 0:
                    log_action(f"Failed login attempt for user: {username} ({remaining_attempts} attempts remaining)")
                    return jsonify({
                        'success': False,
                        'message': f'Invalid email address or password. You have {remaining_attempts} attempt(s) remaining before your account is locked.'
                    })
            else:
                log_action(f"Failed login attempt for non-existent user: {username}")
                return jsonify({
                    'success': False,
                    'message': 'Invalid email address or password.'
                })
    
    except Exception as e:
        app.logger.error(f"Validate credentials error: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': 'An error occurred. Please try again.'
        })


@app.route('/check_tab_session', methods=['POST'])
def check_tab_session():
    """Check if the client's tab session ID matches the server's session"""
    try:
        if not current_user.is_authenticated:
            return jsonify({'valid': False, 'message': 'Not authenticated'})
        
        data = request.get_json()
        client_tab_session_id = data.get('tab_session_id')
        server_tab_session_id = session.get('tab_session_id')
        
        if not server_tab_session_id:
            # Server doesn't have a tab session ID - invalidate
            try:
                logout_user()
            except Exception:
                pass
            session.clear()
            return jsonify({'valid': False, 'message': 'Session expired'})
        
        if client_tab_session_id != server_tab_session_id:
            # Tab session IDs don't match - this is a different tab, invalidate
            try:
                logout_user()
            except Exception:
                pass
            session.clear()
            return jsonify({'valid': False, 'message': 'Tab session mismatch'})
        
        return jsonify({'valid': True, 'tab_session_id': server_tab_session_id})
    except Exception as e:
        app.logger.error(f"Tab session check error: {str(e)}", exc_info=True)
        return jsonify({'valid': False, 'message': 'Error checking session'})


@app.route('/logout')
def logout():
    """Logout current user - allow logout even if not authenticated (for tab session cleanup)"""
    try:
        if current_user.is_authenticated:
            log_action(f"User {current_user.username} logged out")
            logout_user()
    except Exception:
        pass
    # Clear all session data including tab_session_id
    session.clear()
    flash('You have been logged out successfully.', 'info')
    return redirect(url_for('login'))


# ==================== DASHBOARD ROUTES ====================

@app.route('/dashboard')
@login_required
def dashboard():
    """Main dashboard - routes to appropriate dashboard based on role"""
    role = current_user.role
    
    if role == 'Finance Admin':
        return redirect(url_for('admin_dashboard'))
    elif role == 'Finance Staff':
        return redirect(url_for('finance_dashboard'))
    elif role == 'GM':
        return redirect(url_for('gm_dashboard'))
    elif role == 'CEO':
        return redirect(url_for('ceo_dashboard'))
    elif role == 'IT Staff':
        return redirect(url_for('it_dashboard'))
    elif role == 'Department Manager':
        # Route department managers to their specific dashboards
        # Debug: Print the actual department value
        print(f"DEBUG: Department Manager department = '{current_user.department}'")
        print(f"DEBUG: Department type = {type(current_user.department)}")
        print(f"DEBUG: Department length = {len(current_user.department) if current_user.department else 'None'}")
        
        if current_user.department == 'IT':
            return redirect(url_for('it_dashboard'))
        elif current_user.department in ['Project', 'project', 'PROJECT']:
            print("DEBUG: Redirecting to project dashboard")
            return redirect(url_for('project_dashboard'))
        else:
            print(f"DEBUG: Department '{current_user.department}' not matched, going to department dashboard")
            return redirect(url_for('department_dashboard'))
    elif role == 'Project Staff':
        return redirect(url_for('project_dashboard'))
    elif role == 'Operation Manager':
        return redirect(url_for('operation_dashboard'))
    elif role in ['HR Staff', 'Purchasing Staff', 'PR Staff', 'Auditing Staff', 
                  'Customer Service Staff', 'Marketing Staff', 'Operation Staff', 
                  'Quality Control Staff', 'Research and Development Staff', 
                  'Office Staff', 'Maintenance Staff', 'Procurement Staff', 'Logistic Staff']:
        return redirect(url_for('department_dashboard'))
    else:
        # Fallback for any unrecognized roles
        flash('Your role is not properly configured. Please contact IT.', 'warning')
        return redirect(url_for('department_dashboard'))


@app.route('/department/dashboard')
@login_required
@role_required(
    # Department-specific Staff roles (excluding Finance Staff, Project Staff, and IT Staff who have their own dashboards)
    'HR Staff', 'Purchasing Staff', 'PR Staff', 'Auditing Staff',
    'Customer Service Staff', 'Marketing Staff', 'Operation Staff', 'Quality Control Staff',
    'Research and Development Staff', 'Office Staff', 'Maintenance Staff', 'Procurement Staff', 'Logistic Staff',
    # Other roles that can access this dashboard
    'Department Manager', 'Operation Manager'
)
def department_dashboard():
    """Dashboard for department users, finance, and project users"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    search_query = request.args.get('search', None)
    status_filter = request.args.get('status', None)
    tab = request.args.get('tab', 'all')
    # Enforce visibility: only Auditing department users can access 'my_requests' tab
    if tab == 'my_requests' and (not current_user.department or current_user.department != 'Auditing'):
        tab = 'all'
    urgent_filter = request.args.get('urgent', None)
    
    # Validate per_page to prevent abuse
    if per_page not in [10, 20, 50, 100]:
        per_page = 10
    
    # For Department Managers and Operation Managers, show requests from their departments
    if current_user.role in ['Department Manager', 'Operation Manager']:
        # Get requests from their department(s) (including completed/paid ones)
        if current_user.role == 'Operation Manager':
            # Operation Manager can see ALL departments (exclude archived)
            base_query = PaymentRequest.query.filter(PaymentRequest.is_archived == False)
        elif current_user.department == 'Auditing':
            # Auditing Department Manager can see:
            # 1. All their own requests (regardless of status or department)
            # 2. All requests from Auditing department (as department manager)
            # 3. Completed and Recurring requests from OTHER departments (view-only)
            # 4. Any request where they are the temporary manager
            # Exclude archived requests
            base_query = PaymentRequest.query.filter(
                db.or_(
                    PaymentRequest.user_id == current_user.user_id,
                    PaymentRequest.department == 'Auditing',
                    PaymentRequest.temporary_manager_id == current_user.user_id,
                    db.and_(
                        PaymentRequest.department != 'Auditing',
                        PaymentRequest.status.in_(['Completed', 'Recurring'])
                    )
                ),
                PaymentRequest.is_archived == False
            )
        else:
            # Other Department Managers can see ALL their department's requests
            # plus any request where they are the temporary manager
            # Exclude archived requests
            base_query = PaymentRequest.query.filter(
                db.or_(
                    PaymentRequest.department == current_user.department,
                    PaymentRequest.temporary_manager_id == current_user.user_id
                ),
                PaymentRequest.is_archived == False
            )
    elif current_user.department == 'Auditing' and current_user.role == 'Auditing Staff':
        # Auditing Staff can see:
        # 1. All their own requests (regardless of status or department)
        # 2. Completed and Recurring requests from OTHER departments (view-only)
        # Exclude archived requests
        base_query = PaymentRequest.query.filter(
            db.or_(
                PaymentRequest.user_id == current_user.user_id,
                db.and_(
                    PaymentRequest.department != 'Auditing',
                    PaymentRequest.status.in_(['Completed', 'Recurring'])
                )
            ),
            PaymentRequest.is_archived == False
        )
    else:
        # For regular users, show their own requests (exclude archived)
        base_query = PaymentRequest.query.filter(
            PaymentRequest.user_id == current_user.user_id,
            PaymentRequest.is_archived == False
        )
    
    # Exclude CEO-submitted requests for non-authorized roles (visibility hardening)
    if current_user.role not in ['Finance Admin', 'GM', 'Operation Manager']:
        base_query = base_query.filter(~PaymentRequest.user.has(User.role == 'CEO'))

    # Apply urgent filter if provided (before tab filtering)
    if urgent_filter == 'urgent':
        base_query = base_query.filter(PaymentRequest.is_urgent == True)
    elif urgent_filter == 'not_urgent':
        base_query = base_query.filter(PaymentRequest.is_urgent == False)
    
    # Apply search filter if provided (before tab filtering)
    if search_query:
        try:
            # Try to convert to integer for exact match
            search_id = int(search_query)
            base_query = base_query.filter(PaymentRequest.request_id == search_id)
        except ValueError:
            # If not a number, search ONLY by requestor name
            search_term = f'%{search_query}%'
            base_query = base_query.filter(
                PaymentRequest.requestor_name.ilike(search_term)
            )
    
    # Apply tab-based filtering
    if tab == 'completed':
        query = base_query.filter(PaymentRequest.status == 'Completed')
    elif tab == 'rejected':
        query = base_query.filter(PaymentRequest.status.in_(['Rejected by Manager', 'Rejected by Finance', 'Proof Rejected']))
    elif tab == 'recurring':
        query = base_query.filter(PaymentRequest.status == 'Recurring')
    elif tab == 'my_requests':
        # For 'my_requests' tab, show only the current user's requests
        query = base_query.filter(PaymentRequest.user_id == current_user.user_id)
    elif tab == 'all':
        # 'all' tab - show ALL requests, but apply status filter if provided
        query = base_query
        if status_filter:
            query = query.filter(PaymentRequest.status == status_filter)
    else:  # default - show ALL requests
        query = base_query
        if status_filter:
            query = query.filter(PaymentRequest.status == status_filter)
    
    # Get separate queries for each tab content
    # For Auditing department users (Staff and Department Manager), include their own requests + other department's Completed/Recurring
    if current_user.department == 'Auditing' and (current_user.role == 'Auditing Staff' or current_user.role == 'Department Manager'):
        if current_user.role == 'Department Manager':
            # Auditing Department Manager can see:
            # 1. All their own requests
            # 2. All Auditing department requests (as manager)
            # 3. Completed/Recurring from other departments
            # Exclude archived requests
            completed_query = PaymentRequest.query.filter(
                db.or_(
                    PaymentRequest.user_id == current_user.user_id,
                    PaymentRequest.department == 'Auditing',
                    db.and_(
                        PaymentRequest.department != 'Auditing',
                        PaymentRequest.status == 'Completed'
                    )
                ),
                PaymentRequest.is_archived == False
            )
            recurring_query = PaymentRequest.query.filter(
                db.or_(
                    PaymentRequest.user_id == current_user.user_id,
                    PaymentRequest.department == 'Auditing',
                    db.and_(
                        PaymentRequest.department != 'Auditing',
                        PaymentRequest.status == 'Recurring'
                    )
                ),
                PaymentRequest.is_archived == False
            )
            # Rejected query - their own rejected + Auditing department rejected (exclude archived)
            rejected_query = PaymentRequest.query.filter(
                db.or_(
                    db.and_(
                        PaymentRequest.user_id == current_user.user_id,
                        PaymentRequest.status.in_(['Rejected by Manager', 'Rejected by Finance', 'Proof Rejected'])
                    ),
                    db.and_(
                        PaymentRequest.department == 'Auditing',
                        PaymentRequest.status.in_(['Rejected by Manager', 'Rejected by Finance', 'Proof Rejected'])
                    )
                ),
                PaymentRequest.is_archived == False
            )
        else:
            # Auditing Staff (exclude archived)
            completed_query = PaymentRequest.query.filter(
                db.or_(
                    PaymentRequest.user_id == current_user.user_id,
                    db.and_(
                        PaymentRequest.department != 'Auditing',
                        PaymentRequest.status == 'Completed'
                    )
                ),
                PaymentRequest.is_archived == False
            )
            recurring_query = PaymentRequest.query.filter(
                db.or_(
                    PaymentRequest.user_id == current_user.user_id,
                    db.and_(
                        PaymentRequest.department != 'Auditing',
                        PaymentRequest.status == 'Recurring'
                    )
                ),
                PaymentRequest.is_archived == False
            )
            # Rejected query - only their own rejected requests (not other departments' rejected) (exclude archived)
            rejected_query = PaymentRequest.query.filter(
                db.and_(
                    PaymentRequest.user_id == current_user.user_id,
                    PaymentRequest.status.in_(['Rejected by Manager', 'Rejected by Finance', 'Proof Rejected']),
                    PaymentRequest.is_archived == False
                )
            )
    else:
        completed_query = base_query.filter(PaymentRequest.status == 'Completed')
        rejected_query = base_query.filter(PaymentRequest.status.in_(['Rejected by Manager', 'Rejected by Finance', 'Proof Rejected']))
        recurring_query = base_query.filter(PaymentRequest.status == 'Recurring')
    
    # Apply urgent filter to separate queries
    if urgent_filter == 'urgent':
        completed_query = completed_query.filter(PaymentRequest.is_urgent == True)
        rejected_query = rejected_query.filter(PaymentRequest.is_urgent == True)
        recurring_query = recurring_query.filter(PaymentRequest.is_urgent == True)
    elif urgent_filter == 'not_urgent':
        completed_query = completed_query.filter(PaymentRequest.is_urgent == False)
        rejected_query = rejected_query.filter(PaymentRequest.is_urgent == False)
        recurring_query = recurring_query.filter(PaymentRequest.is_urgent == False)
    
    # Apply search filter to separate queries
    if search_query:
        try:
            search_id = int(search_query)
            completed_query = completed_query.filter(PaymentRequest.request_id == search_id)
            rejected_query = rejected_query.filter(PaymentRequest.request_id == search_id)
            recurring_query = recurring_query.filter(PaymentRequest.request_id == search_id)
        except ValueError:
            search_term = f'%{search_query}%'
            completed_query = completed_query.filter(
                db.or_(
                    PaymentRequest.requestor_name.ilike(search_term),
                    PaymentRequest.purpose.ilike(search_term),
                    PaymentRequest.account_name.ilike(search_term)
                )
            )
            rejected_query = rejected_query.filter(
                db.or_(
                    PaymentRequest.requestor_name.ilike(search_term),
                    PaymentRequest.purpose.ilike(search_term),
                    PaymentRequest.account_name.ilike(search_term)
                )
            )
            recurring_query = recurring_query.filter(
                db.or_(
                    PaymentRequest.requestor_name.ilike(search_term),
                    PaymentRequest.purpose.ilike(search_term),
                    PaymentRequest.account_name.ilike(search_term)
                )
            )
    
    # Get data for each tab
    completed_requests = completed_query.order_by(get_completed_datetime_order()).all()
    rejected_requests = rejected_query.order_by(get_rejected_datetime_order()).all()
    recurring_requests = recurring_query.order_by(PaymentRequest.created_at.desc()).all()
    
    # Paginate the main query
    # For 'all' tab, sort by status priority then by date (Completed by completion_date, others by created_at)
    if tab == 'all':
        requests_pagination = query.order_by(
            get_status_priority_order(),
            get_all_tab_datetime_order()
        ).paginate(
            page=page, per_page=per_page, error_out=False
        )
    elif tab == 'completed':
        # Completed tab sorted by approval_date (most recent first)
        requests_pagination = query.order_by(get_completed_datetime_order()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    else:
        requests_pagination = query.order_by(PaymentRequest.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    
    # Get notifications for department managers and regular users
    notifications = get_notifications_for_user(current_user)
    unread_count = get_unread_count_for_user(current_user)
    
    # Get user's own requests for the My Requests tab (exclude archived)
    my_requests_query = PaymentRequest.query.filter(
        PaymentRequest.user_id == current_user.user_id,
        PaymentRequest.is_archived == False
    )
    my_requests_pagination = my_requests_query.order_by(PaymentRequest.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return render_template('department_dashboard.html', 
                         requests=requests_pagination.items, 
                         pagination=requests_pagination,
                         user=current_user,
                         notifications=notifications,
                         unread_count=unread_count,
                         status_filter=status_filter,
                         search_query=search_query,
                         completed_requests=completed_requests,
                         rejected_requests=rejected_requests,
                         recurring_requests=recurring_requests,
                         urgent_filter=urgent_filter,
                         my_requests=my_requests_pagination.items,
                         active_tab=tab)


@app.route('/admin/dashboard')
@login_required
@role_required('Finance Admin')
def admin_dashboard():
    """Dashboard for admin - shows all requests with optional status filtering"""
    # Check for recurring payments due today
    check_recurring_payments_due()
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    status_filter = request.args.get('status', None)
    department_filter = request.args.get('department', None)
    search_query = request.args.get('search', None)
    urgent_filter = request.args.get('urgent', None)
    tab = request.args.get('tab', 'all')  # 'all' tab shows all requests
    
    # Validate per_page to prevent abuse
    if per_page not in [10, 20, 50, 100]:
        per_page = 10
    
    # Build query with optional status, department, and search filters
    # Apply finance status filtering for all tabs (including "All Requests" tab)
    # Finance Admin can see finance-related statuses + Pending Manager Approval from Finance department only (or their own requests)
    finance_statuses = ['Pending Finance Approval', 'Proof Pending', 'Proof Sent', 'Proof Rejected', 'Recurring', 'Completed', 'Rejected by Finance']

    # For Abdalaziz, also include Pending Manager Approval and Rejected by Manager
    # for Finance department, his own requests, temporary manager assignments,
    # and requests submitted by Finance Staff, GM, Operation Manager, and CEO
    if current_user.name == 'Abdalaziz Al-Brashdi':
        special_submitter_roles = ['Finance Staff', 'GM', 'Operation Manager', 'CEO']
        query = PaymentRequest.query.filter(
            db.or_(
                PaymentRequest.status.in_(finance_statuses),
                # Pending Manager Approval from Finance department
                db.and_(
                    PaymentRequest.status == 'Pending Manager Approval',
                    PaymentRequest.department == 'Finance'
                ),
                # Rejected by Manager from Finance department
                db.and_(
                    PaymentRequest.status == 'Rejected by Manager',
                    PaymentRequest.department == 'Finance'
                ),
                # Always include current user's own requests regardless of status
                PaymentRequest.user_id == current_user.user_id,
                # Include requests where the current user is temporarily assigned as manager
                db.and_(
                    PaymentRequest.status == 'Pending Manager Approval',
                    PaymentRequest.temporary_manager_id == current_user.user_id
                ),
                # Include PMA/Rejected-by-Manager for specific submitter roles (GM/CEO/etc.)
                db.and_(
                    PaymentRequest.status.in_(['Pending Manager Approval', 'Rejected by Manager']),
                    PaymentRequest.user.has(User.role.in_(special_submitter_roles))
                )
            ),
            PaymentRequest.is_archived == False
        )
    else:
        # Other Finance Admins only see finance-related statuses plus temporary assignments awaiting manager approval
        query = PaymentRequest.query.filter(
            db.or_(
                PaymentRequest.status.in_(finance_statuses),
                # Always include current user's own requests regardless of status
                PaymentRequest.user_id == current_user.user_id,
                db.and_(
                    PaymentRequest.status == 'Pending Manager Approval',
                    PaymentRequest.temporary_manager_id == current_user.user_id
                )
            ),
            PaymentRequest.is_archived == False
        )
    
    # Apply department filter (before tab filtering)
    if department_filter:
        query = query.filter(PaymentRequest.department == department_filter)
    
    # Apply search filter (before tab filtering)
    if search_query:
        # Search by request ID or requestor name
        try:
            # Try to convert to integer for exact match
            search_id = int(search_query)
            query = query.filter(PaymentRequest.request_id == search_id)
        except ValueError:
            # If not a number, search ONLY by requestor name
            search_term = f'%{search_query}%'
            query = query.filter(
                PaymentRequest.requestor_name.ilike(search_term)
            )
    
    # Apply urgent filter (before tab filtering)
    if urgent_filter:
        if urgent_filter == 'urgent':
            query = query.filter(PaymentRequest.is_urgent == True)
        elif urgent_filter == 'not_urgent':
            query = query.filter(PaymentRequest.is_urgent == False)
    
    # Apply tab-based filtering
    if tab == 'completed':
        query = query.filter(PaymentRequest.status == 'Completed')
    elif tab == 'rejected':
        query = query.filter(db.or_(
            PaymentRequest.status == 'Rejected by Manager',
            PaymentRequest.status == 'Rejected by Finance'
        ))
    elif tab == 'recurring':
        query = query.filter(PaymentRequest.status == 'Recurring')
    elif tab == 'my_requests':
        # For 'my_requests' tab, show only the current user's requests
        query = query.filter(PaymentRequest.user_id == current_user.user_id)
    elif tab == 'all':
        # 'all' tab - apply status filter if provided (only on all tab)
        if status_filter:
            query = query.filter(PaymentRequest.status == status_filter)
    
    # Get paginated requests
    # For 'all' tab, sort by status priority then by date (Completed by completion_date, others by created_at)
    if tab == 'all':
        requests_pagination = query.order_by(
            get_status_priority_order(),
            get_all_tab_datetime_order()
        ).paginate(
            page=page, per_page=per_page, error_out=False
        )
    elif tab == 'completed':
        requests_pagination = query.order_by(get_completed_datetime_order()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    else:
        requests_pagination = query.order_by(PaymentRequest.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    
    notifications = get_notifications_for_user(current_user)
    unread_count = get_unread_count_for_user(current_user)
    
    # Get overdue requests count
    overdue_count = get_overdue_requests_count()
    
    # Get user's own requests for the My Requests tab (exclude archived)
    my_requests_query = PaymentRequest.query.filter(
        PaymentRequest.user_id == current_user.user_id,
        PaymentRequest.is_archived == False
    )
    my_requests_pagination = my_requests_query.order_by(PaymentRequest.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return render_template('admin_dashboard.html', 
                         requests=requests_pagination.items, 
                         my_requests=my_requests_pagination.items,
                         pagination=requests_pagination,
                         user=current_user, 
                         notifications=notifications, 
                         unread_count=unread_count,
                         overdue_count=overdue_count,
                         status_filter=status_filter,
                         department_filter=department_filter,
                         search_query=search_query,
                         urgent_filter=urgent_filter,
                         active_tab=tab)


@app.route('/finance/dashboard')
@login_required
@role_required('Finance Staff')
def finance_dashboard():
    """Dashboard for finance - can view all reports and submit requests"""
    # Check for recurring payments due today
    check_recurring_payments_due()
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    department_filter = request.args.get('department', None)
    status_filter = request.args.get('status', None)
    search_query = request.args.get('search', None)
    urgent_filter = request.args.get('urgent', None)
    tab = request.args.get('tab', 'all')  # 'all' tab shows all requests
    
    # Validate per_page to prevent abuse
    if per_page not in [10, 20, 50, 100]:
        per_page = 10
    
    # Build query with optional department and search filters
    # Apply finance status filtering for all tabs (including "All Requests" tab)
    # Finance Staff can see finance-related statuses + their own requests with Pending Manager Approval + their own requests with Rejected by Manager
    # Abdalaziz can see finance-related statuses + Pending Manager Approval + Rejected by Manager for Finance department only (or his own requests)
    finance_statuses = ['Pending Finance Approval', 'Proof Pending', 'Proof Sent', 'Proof Rejected', 'Recurring', 'Completed', 'Rejected by Finance']
    
    # Base query for finance-related statuses (exclude archived)
    query = PaymentRequest.query.filter(
        PaymentRequest.status.in_(finance_statuses),
        PaymentRequest.is_archived == False
    )
    # Visibility hardening: Finance Staff must not see CEO-submitted requests
    if current_user.role == 'Finance Staff':
        query = query.filter(~PaymentRequest.user.has(User.role == 'CEO'))
    
    # Add Finance Staff's own requests with Pending Manager Approval and Rejected by Manager (exclude archived)
    if current_user.role == 'Finance Staff':
        own_pending_requests = PaymentRequest.query.filter(
            db.and_(
                PaymentRequest.user_id == current_user.user_id,
                PaymentRequest.status.in_(['Pending Manager Approval', 'Rejected by Manager']),
                PaymentRequest.is_archived == False
            )
        )
        query = query.union(own_pending_requests)
    
    # Add Abdalaziz's special permissions: Finance department requests OR his own requests (exclude archived)
    elif current_user.name == 'Abdalaziz Al-Brashdi':
        abdalaziz_special_requests = PaymentRequest.query.filter(
            db.and_(
                PaymentRequest.status.in_(['Pending Manager Approval', 'Rejected by Manager']),
                db.or_(
                    PaymentRequest.department == 'Finance',
                    PaymentRequest.user_id == current_user.user_id
                ),
                PaymentRequest.is_archived == False
            )
        )
        query = query.union(abdalaziz_special_requests)
    
    # Apply department filter (before tab filtering)
    if department_filter:
        query = query.filter(PaymentRequest.department == department_filter)
    
    # Apply search filter (before tab filtering)
    if search_query:
        # Search by request ID or requestor name
        try:
            # Try to convert to integer for exact match
            search_id = int(search_query)
            query = query.filter(PaymentRequest.request_id == search_id)
        except ValueError:
            # If not a number, search ONLY by requestor name
            search_term = f'%{search_query}%'
            query = query.filter(
                PaymentRequest.requestor_name.ilike(search_term)
            )
    
    # Apply urgent filter (before tab filtering)
    if urgent_filter:
        if urgent_filter == 'urgent':
            query = query.filter(PaymentRequest.is_urgent == True)
        elif urgent_filter == 'not_urgent':
            query = query.filter(PaymentRequest.is_urgent == False)
    
    # Apply tab-based filtering
    if tab == 'completed':
        query = query.filter(PaymentRequest.status == 'Completed')
    elif tab == 'rejected':
        query = query.filter(db.or_(
            PaymentRequest.status == 'Rejected by Manager',
            PaymentRequest.status == 'Rejected by Finance'
        ))
    elif tab == 'recurring':
        query = query.filter(PaymentRequest.status == 'Recurring')
    elif tab == 'my_requests':
        # For 'my_requests' tab, show only the current user's requests
        query = query.filter(PaymentRequest.user_id == current_user.user_id)
    elif tab == 'all':
        # 'all' tab - apply status filter if provided (only on all tab)
        if status_filter:
            query = query.filter(PaymentRequest.status == status_filter)
        # Otherwise show all requests that the user can see
    
    # Get paginated requests
    # For 'all' tab, sort by status priority then by date (Completed by completion_date, others by created_at)
    if tab == 'all':
        requests_pagination = query.order_by(
            get_status_priority_order(),
            get_all_tab_datetime_order()
        ).paginate(
            page=page, per_page=per_page, error_out=False
        )
    elif tab == 'completed':
        requests_pagination = query.order_by(get_completed_datetime_order()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    else:
        requests_pagination = query.order_by(PaymentRequest.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    
    notifications = get_notifications_for_user(current_user)
    unread_count = get_unread_count_for_user(current_user)
    
    # Get overdue requests count
    overdue_count = get_overdue_requests_count()
    
    # Get user's own requests for the My Requests tab (exclude archived)
    my_requests_query = PaymentRequest.query.filter(
        PaymentRequest.user_id == current_user.user_id,
        PaymentRequest.is_archived == False
    )
    my_requests_pagination = my_requests_query.order_by(PaymentRequest.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return render_template('finance_dashboard.html', 
                         requests=requests_pagination.items, 
                         my_requests=my_requests_pagination.items,
                         pagination=requests_pagination,
                         user=current_user, 
                         notifications=notifications, 
                         unread_count=unread_count,
                         overdue_count=overdue_count,
                         department_filter=department_filter,
                         search_query=search_query,
                         urgent_filter=urgent_filter,
                         active_tab=tab)


@app.route('/gm/dashboard')
@login_required
@role_required('GM')
def gm_dashboard():
    """Dashboard for General Manager - view all reports (Approved/Pending only)"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    department_filter = request.args.get('department', None)
    status_filter = request.args.get('status', None)
    search_query = request.args.get('search', None)
    urgent_filter = request.args.get('urgent', None)
    tab = request.args.get('tab', 'all')  # 'all' tab shows all requests
    
    # Validate per_page to prevent abuse
    if per_page not in [10, 20, 50, 100]:
        per_page = 10
    
    # Build query with optional department and search filters
    # GM can see ALL requests from ALL departments including rejected by manager
    # Exclude archived requests
    query = PaymentRequest.query.filter(PaymentRequest.is_archived == False)
    if department_filter:
        query = query.filter(PaymentRequest.department == department_filter)
    if search_query:
        # Search by request ID or requestor name
        try:
            # Try to convert to integer for exact match
            search_id = int(search_query)
            query = query.filter(PaymentRequest.request_id == search_id)
        except ValueError:
            # If not a number, search ONLY by requestor name
            search_term = f'%{search_query}%'
            query = query.filter(
                PaymentRequest.requestor_name.ilike(search_term)
            )
    if urgent_filter:
        if urgent_filter == 'urgent':
            query = query.filter(PaymentRequest.is_urgent == True)
        elif urgent_filter == 'not_urgent':
            query = query.filter(PaymentRequest.is_urgent == False)
    
    # Apply tab-based filtering
    if tab == 'completed':
        query = query.filter(PaymentRequest.status.in_(['Completed', 'Paid', 'Approved']))
    elif tab == 'rejected':
        query = query.filter(PaymentRequest.status.in_(['Rejected by Manager', 'Rejected by Finance', 'Proof Rejected']))
    elif tab == 'recurring':
        query = query.filter(PaymentRequest.status == 'Recurring')
    elif tab == 'my_requests':
        # For 'my_requests' tab, show only the current user's requests
        query = query.filter(PaymentRequest.user_id == current_user.user_id)
    elif tab == 'all':
        # 'all' tab (All Requests) shows all requests
        # Apply status filter if provided (excludes Completed, Rejected, Recurring from dropdown)
        if status_filter:
            query = query.filter(PaymentRequest.status == status_filter)
    # Default case also shows all requests
    
    # Get paginated requests
    # For 'all' tab, sort by status priority then by date (Completed by completion_date, others by created_at)
    if tab == 'all':
        requests_pagination = query.order_by(
            get_status_priority_order(),
            get_all_tab_datetime_order()
        ).paginate(
            page=page, per_page=per_page, error_out=False
        )
    elif tab == 'completed':
        requests_pagination = query.order_by(get_completed_datetime_order()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    else:
        requests_pagination = query.order_by(PaymentRequest.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    
    # Calculate statistics (all requests from all departments - exclude archived)
    all_requests = PaymentRequest.query.filter(PaymentRequest.is_archived == False).all()
    total_requests = len(all_requests)
    approved = len([r for r in all_requests if r.status == 'Approved'])
    pending = len([r for r in all_requests if r.status == 'Pending'])
    total_amount = sum([float(r.amount) for r in all_requests if r.status == 'Approved'])
    
    stats = {
        'total_requests': total_requests,
        'approved': approved,
        'pending': pending,
        'total_amount': total_amount
    }
    
    # Get notifications for GM (all notifications)
    notifications = get_notifications_for_user(current_user)
    unread_count = get_unread_count_for_user(current_user)
    
    # Get separate queries for completed, rejected, and recurring requests for tab content (exclude archived)
    completed_query = PaymentRequest.query.filter(PaymentRequest.is_archived == False)
    rejected_query = PaymentRequest.query.filter(PaymentRequest.is_archived == False)
    recurring_query = PaymentRequest.query.filter(PaymentRequest.is_archived == False)
    
    if department_filter:
        completed_query = completed_query.filter(PaymentRequest.department == department_filter)
        rejected_query = rejected_query.filter(PaymentRequest.department == department_filter)
        recurring_query = recurring_query.filter(PaymentRequest.department == department_filter)
    
    completed_query = completed_query.filter(PaymentRequest.status == 'Completed')
    rejected_query = rejected_query.filter(PaymentRequest.status.in_(['Rejected by Manager', 'Rejected by Finance', 'Proof Rejected']))
    recurring_query = recurring_query.filter(PaymentRequest.status == 'Recurring')
    
    completed_requests = completed_query.order_by(get_completed_datetime_order()).all()
    rejected_requests = rejected_query.order_by(get_rejected_datetime_order()).all()
    recurring_requests = recurring_query.order_by(PaymentRequest.created_at.desc()).all()
    
    # Get user's own requests for the My Requests tab (exclude archived)
    my_requests_query = PaymentRequest.query.filter(
        PaymentRequest.user_id == current_user.user_id,
        PaymentRequest.is_archived == False
    )
    my_requests_pagination = my_requests_query.order_by(PaymentRequest.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return render_template('gm_dashboard.html', 
                         requests=requests_pagination.items, 
                         pagination=requests_pagination,
                         stats=stats, 
                         user=current_user,
                         notifications=notifications,
                         unread_count=unread_count,
                         department_filter=department_filter,
                         status_filter=status_filter,
                         search_query=search_query,
                         urgent_filter=urgent_filter,
                         completed_requests=completed_requests,
                         rejected_requests=rejected_requests,
                         recurring_requests=recurring_requests,
                         my_requests=my_requests_pagination.items,
                         active_tab=tab)


@app.route('/ceo/dashboard')
@login_required
@role_required('CEO')
def ceo_dashboard():
    """Dashboard for CEO - identical to GM but view-only (no approval actions)"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    department_filter = request.args.get('department', None)
    status_filter = request.args.get('status', None)
    search_query = request.args.get('search', None)
    urgent_filter = request.args.get('urgent', None)
    tab = request.args.get('tab', 'all')

    if per_page not in [10, 20, 50, 100]:
        per_page = 10

    # Exclude archived requests
    query = PaymentRequest.query.filter(PaymentRequest.is_archived == False)
    if department_filter:
        query = query.filter(PaymentRequest.department == department_filter)
    if search_query:
        # Search by request ID or requestor name
        try:
            search_id = int(search_query)
            query = query.filter(PaymentRequest.request_id == search_id)
        except ValueError:
            # If not a number, search ONLY by requestor name
            search_term = f'%{search_query}%'
            query = query.filter(
                PaymentRequest.requestor_name.ilike(search_term)
            )
    if urgent_filter:
        if urgent_filter == 'urgent':
            query = query.filter(PaymentRequest.is_urgent == True)
        elif urgent_filter == 'not_urgent':
            query = query.filter(PaymentRequest.is_urgent == False)

    if tab == 'completed':
        query = query.filter(PaymentRequest.status.in_(['Completed', 'Paid', 'Approved']))
    elif tab == 'rejected':
        query = query.filter(PaymentRequest.status.in_(['Rejected by Manager', 'Rejected by Finance', 'Proof Rejected']))
    elif tab == 'recurring':
        query = query.filter(PaymentRequest.status == 'Recurring')
    elif tab == 'my_requests':
        query = query.filter(PaymentRequest.user_id == current_user.user_id)
    elif tab == 'all':
        if status_filter:
            query = query.filter(PaymentRequest.status == status_filter)

    # Get paginated requests
    # For 'all' tab, sort by status priority then by date (Completed by completion_date, others by created_at)
    if tab == 'all':
        requests_pagination = query.order_by(
            get_status_priority_order(),
            get_all_tab_datetime_order()
        ).paginate(
            page=page, per_page=per_page, error_out=False
        )
    elif tab == 'completed':
        requests_pagination = query.order_by(get_completed_datetime_order()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    else:
        requests_pagination = query.order_by(PaymentRequest.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )

    # Exclude archived requests from stats
    all_requests = PaymentRequest.query.filter(PaymentRequest.is_archived == False).all()
    stats = {
        'total_requests': len(all_requests),
        'approved': len([r for r in all_requests if r.status == 'Approved']),
        'pending': len([r for r in all_requests if r.status == 'Pending']),
        'total_amount': sum([float(r.amount) for r in all_requests if r.status == 'Approved'])
    }

    notifications = get_notifications_for_user(current_user)
    unread_count = get_unread_count_for_user(current_user)

    # Exclude archived requests from tab queries
    completed_query = PaymentRequest.query.filter(PaymentRequest.is_archived == False)
    rejected_query = PaymentRequest.query.filter(PaymentRequest.is_archived == False)
    recurring_query = PaymentRequest.query.filter(PaymentRequest.is_archived == False)
    if department_filter:
        completed_query = completed_query.filter(PaymentRequest.department == department_filter)
        rejected_query = rejected_query.filter(PaymentRequest.department == department_filter)
        recurring_query = recurring_query.filter(PaymentRequest.department == department_filter)

    completed_query = completed_query.filter(PaymentRequest.status == 'Completed')
    rejected_query = rejected_query.filter(PaymentRequest.status.in_(['Rejected by Manager', 'Rejected by Finance', 'Proof Rejected']))
    recurring_query = recurring_query.filter(PaymentRequest.status == 'Recurring')

    completed_requests = completed_query.order_by(get_completed_datetime_order()).all()
    rejected_requests = rejected_query.order_by(PaymentRequest.created_at.desc()).all()
    recurring_requests = recurring_query.order_by(PaymentRequest.created_at.desc()).all()

    my_requests_query = PaymentRequest.query.filter(
        PaymentRequest.user_id == current_user.user_id,
        PaymentRequest.is_archived == False
    )
    my_requests_pagination = my_requests_query.order_by(PaymentRequest.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    return render_template('ceo_dashboard.html',
                          requests=requests_pagination.items,
                          pagination=requests_pagination,
                          stats=stats,
                          user=current_user,
                          notifications=notifications,
                          unread_count=unread_count,
                          department_filter=department_filter,
                          status_filter=status_filter,
                          search_query=search_query,
                          urgent_filter=urgent_filter,
                          completed_requests=completed_requests,
                          rejected_requests=rejected_requests,
                          recurring_requests=recurring_requests,
                          my_requests=my_requests_pagination.items,
                          active_tab=tab)

@app.route('/it/dashboard')
@login_required
@role_required('IT Staff', 'Department Manager')
def it_dashboard():
    """Dashboard for IT - full CRUD access"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    department_filter = request.args.get('department', None)
    status_filter = request.args.get('status', None)
    search_query = request.args.get('search', None)
    urgent_filter = request.args.get('urgent', None)
    location_filter = request.args.get('location', None)
    tab = request.args.get('tab', 'all')  # 'all' tab shows all requests
    
    # If department filter is provided, default to request-types tab
    if department_filter and not request.args.get('tab'):
        tab = 'request-types'
    
    # Validate per_page to prevent abuse
    if per_page not in [10, 20, 50, 100]:
        per_page = 10
    
    # Restrict Department Managers to IT department only
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Build query with optional department and search filters
    if current_user.role == 'IT Staff' or (current_user.role == 'Department Manager' and current_user.department == 'IT'):
        # IT users and IT Department Managers see all requests (sensitive fields masked in UI)
        # Exclude archived requests
        query = PaymentRequest.query.filter(PaymentRequest.is_archived == False)
    else:
        # Other users should not see requests that are still pending manager approval
        # Exclude archived requests
        query = PaymentRequest.query.filter(
            PaymentRequest.status != 'Pending Manager Approval',
            PaymentRequest.is_archived == False
        )
    
    if department_filter:
        query = query.filter(PaymentRequest.department == department_filter)
    if search_query:
        # Search by request ID or requestor name
        try:
            # Try to convert to integer for exact match
            search_id = int(search_query)
            query = query.filter(PaymentRequest.request_id == search_id)
        except ValueError:
            # If not a number, search ONLY by requestor name
            search_term = f'%{search_query}%'
            query = query.filter(
                PaymentRequest.requestor_name.ilike(search_term)
            )
    if urgent_filter:
        if urgent_filter == 'urgent':
            query = query.filter(PaymentRequest.is_urgent == True)
        elif urgent_filter == 'not_urgent':
            query = query.filter(PaymentRequest.is_urgent == False)
    
    # Apply tab-based filtering
    if tab == 'completed':
        query = query.filter(PaymentRequest.status.in_(['Completed', 'Paid', 'Approved']))
    elif tab == 'rejected':
        query = query.filter(PaymentRequest.status.in_(['Rejected by Manager', 'Rejected by Finance', 'Proof Rejected']))
    elif tab == 'recurring':
        query = query.filter(PaymentRequest.status == 'Recurring')
    elif tab == 'all':
        # 'all' tab (All Requests) shows all requests
        # Apply status filter if provided (excludes Completed, Rejected, Recurring from dropdown)
        if status_filter:
            query = query.filter(PaymentRequest.status == status_filter)
    # Default case also shows all requests
    
    # Get paginated requests
    # For 'all' tab, sort by status priority then by date (Completed by completion_date, others by created_at)
    if tab == 'all':
        requests_pagination = query.order_by(
            get_status_priority_order(),
            get_all_tab_datetime_order()
        ).paginate(
            page=page, per_page=per_page, error_out=False
        )
    elif tab == 'completed':
        requests_pagination = query.order_by(get_completed_datetime_order()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    else:
        requests_pagination = query.order_by(PaymentRequest.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    
    # Get notifications for IT users and IT Department Managers
    notifications = get_notifications_for_user(current_user)
    unread_count = get_unread_count_for_user(current_user)
    
    # Get separate queries for completed, rejected, and recurring requests for tab content
    # Exclude archived requests from all queries
    completed_query = PaymentRequest.query.filter(PaymentRequest.is_archived == False)
    rejected_query = PaymentRequest.query.filter(PaymentRequest.is_archived == False)
    recurring_query = PaymentRequest.query.filter(PaymentRequest.is_archived == False)
    
    if current_user.role == 'IT Staff' or (current_user.role == 'Department Manager' and current_user.department == 'IT'):
        # IT users and IT Department Managers see all requests
        pass
    else:
        # Other users should not see requests that are still pending manager approval
        completed_query = completed_query.filter(PaymentRequest.status != 'Pending Manager Approval')
        rejected_query = rejected_query.filter(PaymentRequest.status != 'Pending Manager Approval')
        recurring_query = recurring_query.filter(PaymentRequest.status != 'Pending Manager Approval')
    
    if department_filter:
        completed_query = completed_query.filter(PaymentRequest.department == department_filter)
        rejected_query = rejected_query.filter(PaymentRequest.department == department_filter)
        recurring_query = recurring_query.filter(PaymentRequest.department == department_filter)
    
    completed_query = completed_query.filter(PaymentRequest.status == 'Completed')
    rejected_query = rejected_query.filter(PaymentRequest.status.in_(['Rejected by Manager', 'Rejected by Finance', 'Proof Rejected']))
    recurring_query = recurring_query.filter(PaymentRequest.status == 'Recurring')
    
    completed_requests = completed_query.order_by(get_completed_datetime_order()).all()
    rejected_requests = rejected_query.order_by(PaymentRequest.created_at.desc()).all()
    recurring_requests = recurring_query.order_by(PaymentRequest.created_at.desc()).all()
    
    # Get user's own requests for the My Requests tab (exclude archived)
    my_requests_query = PaymentRequest.query.filter(
        PaymentRequest.user_id == current_user.user_id,
        PaymentRequest.is_archived == False
    )
    my_requests_pagination = my_requests_query.order_by(PaymentRequest.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    users = User.query.all()
    logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(50).all()
    
    # Get request types for the Request Types Management section with department filter
    request_types_query = RequestType.query
    
    # Apply department filter if specified
    if department_filter and department_filter != 'all':
        if department_filter == 'Management':
            request_types_query = request_types_query.filter(RequestType.department.in_(['Management', 'General Manager']))
        else:
            request_types_query = request_types_query.filter(RequestType.department == department_filter)
    
    request_types = request_types_query.order_by(RequestType.id).all()
    
    # Get branches for the Branches Management section
    branches_query = Branch.query.filter_by(is_active=True)
    
    # Apply location filter if specified
    if location_filter:
        branches_query = branches_query.filter(Branch.restaurant == location_filter)
    
    branches = branches_query.order_by(Branch.id).all()
    
    # Get all departments for the filter dropdown and normalize legacy label
    all_departments = db.session.query(RequestType.department).distinct().order_by(RequestType.department).all()
    departments = []
    for dept in all_departments:
        d = dept[0]
        departments.append('Management' if d == 'General Manager' else d)
    
    return render_template('it_dashboard.html', 
                         requests=requests_pagination.items, 
                         my_requests=my_requests_pagination.items,
                         pagination=requests_pagination,
                         users=users, 
                         logs=logs, 
                         request_types=request_types,
                         branches=branches,
                         departments=departments,
                         user=current_user,
                         notifications=notifications,
                         unread_count=unread_count,
                         department_filter=department_filter,
                         status_filter=status_filter,
                         search_query=search_query,
                         location_filter=location_filter,
                         urgent_filter=urgent_filter,
                         completed_requests=completed_requests,
                         rejected_requests=rejected_requests,
                         recurring_requests=recurring_requests,
                         active_tab=tab)


@app.route('/it/archives')
@login_required
@role_required('IT Staff', 'Department Manager')
def archives():
    """Archives page for IT department - shows archived requests"""
    # Restrict Department Managers to IT department only
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    department_filter = request.args.get('department', None)
    status_filter = request.args.get('status', None)
    search_query = request.args.get('search', None)
    urgent_filter = request.args.get('urgent', None)
    
    # Validate per_page to prevent abuse
    if per_page not in [10, 20, 50, 100]:
        per_page = 10
    
    # Build query for archived requests only
    query = PaymentRequest.query.filter(PaymentRequest.is_archived == True)
    
    # Apply filters
    if department_filter:
        query = query.filter(PaymentRequest.department == department_filter)
    
    if search_query:
        # Search by request ID or requestor name
        try:
            search_id = int(search_query)
            query = query.filter(PaymentRequest.request_id == search_id)
        except ValueError:
            search_term = f'%{search_query}%'
            query = query.filter(
                PaymentRequest.requestor_name.ilike(search_term)
            )
    
    if urgent_filter:
        if urgent_filter == 'urgent':
            query = query.filter(PaymentRequest.is_urgent == True)
        elif urgent_filter == 'not_urgent':
            query = query.filter(PaymentRequest.is_urgent == False)
    
    if status_filter:
        query = query.filter(PaymentRequest.status == status_filter)
    
    # Get paginated archived requests
    requests_pagination = query.order_by(PaymentRequest.archived_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Get all departments for the filter dropdown
    all_departments = db.session.query(RequestType.department).distinct().order_by(RequestType.department).all()
    departments = []
    for dept in all_departments:
        d = dept[0]
        departments.append('Management' if d == 'General Manager' else d)
    
    return render_template('archives.html',
                         requests=requests_pagination.items,
                         pagination=requests_pagination,
                         departments=departments,
                         user=current_user,
                         department_filter=department_filter,
                         status_filter=status_filter,
                         search_query=search_query,
                         urgent_filter=urgent_filter)


@app.route('/request/<int:request_id>/restore', methods=['POST'])
@login_required
@role_required('IT Staff', 'Department Manager')
def restore_request(request_id):
    """Restore an archived payment request (IT only)"""
    # Restrict Department Managers to IT department only
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    
    req = PaymentRequest.query.get_or_404(request_id)
    
    # Check if request is archived
    if not req.is_archived:
        flash(f'Payment request #{request_id} is not archived.', 'warning')
        return redirect(url_for('archives'))
    
    # Resume any paused approval timing when restoring
    # If timing was paused (end_time was set during archiving), resume it from where it left off
    # We adjust the start_time to account for the pause duration, so elapsed time continues correctly
    
    # Store archived_at before clearing it (needed for timing check)
    archived_at_time = req.archived_at
    restore_time = datetime.utcnow()
    
    # Resume manager approval timing if it was paused
    if req.manager_approval_start_time and req.manager_approval_end_time:
        # Check if end_time was set during archiving (close to archived_at time)
        if archived_at_time and abs((req.manager_approval_end_time - archived_at_time).total_seconds()) < 60:
            # Timing was paused during archiving - calculate elapsed time and adjust start_time
            # Elapsed time = time from start to when it was paused
            elapsed_time = req.manager_approval_end_time - req.manager_approval_start_time
            # Adjust start_time so elapsed time calculation continues from where it left off
            # New start_time = now - elapsed_time (so elapsed = now - new_start = elapsed_time + time_since_restore)
            req.manager_approval_start_time = restore_time - elapsed_time
            req.manager_approval_end_time = None  # Clear end_time to resume timing
            print(f"DEBUG: Resumed manager approval timing for request #{request_id} (elapsed: {elapsed_time.total_seconds()}s)")
    
    # Resume finance approval timing if it was paused
    if req.finance_approval_start_time and req.finance_approval_end_time:
        # Check if end_time was set during archiving (close to archived_at time)
        if archived_at_time and abs((req.finance_approval_end_time - archived_at_time).total_seconds()) < 60:
            # Timing was paused during archiving - calculate elapsed time and adjust start_time
            # Elapsed time = time from start to when it was paused
            elapsed_time = req.finance_approval_end_time - req.finance_approval_start_time
            # Adjust start_time so elapsed time calculation continues from where it left off
            # New start_time = now - elapsed_time (so elapsed = now - new_start = elapsed_time + time_since_restore)
            req.finance_approval_start_time = restore_time - elapsed_time
            req.finance_approval_end_time = None  # Clear end_time to resume timing
            print(f"DEBUG: Resumed finance approval timing for request #{request_id} (elapsed: {elapsed_time.total_seconds()}s)")
    
    # Restore the request
    req.is_archived = False
    req.archived_at = None
    req.archived_by = None
    req.archived_by_user_id = None
    
    db.session.commit()
    
    log_action(f"Restored payment request #{request_id}")
    
    # Notify all IT department users about the restore
    it_users = User.query.filter(
        db.or_(
            User.role == 'IT Staff',
            db.and_(User.role == 'Department Manager', User.department == 'IT')
        )
    ).all()
    
    print(f"DEBUG: Found {len(it_users)} IT department users for restore notification")
    for u in it_users:
        print(f"DEBUG: IT user: {u.username} (ID: {u.user_id}, Role: {u.role}, Dept: {u.department})")
    
    # Exclude the user who restored (they already know)
    it_users_to_notify = [user for user in it_users if user.user_id != current_user.user_id]
    
    print(f"DEBUG: Will notify {len(it_users_to_notify)} IT users about restore (excluding restorer: {current_user.username})")
    
    notification_title = "Payment Request Restored"
    notification_message = f"Payment request #{request_id} submitted by {req.requestor_name} has been restored by {current_user.name}."
    
    for it_user in it_users_to_notify:
        print(f"DEBUG: Creating restore notification for IT user: {it_user.username} (ID: {it_user.user_id})")
        create_notification(
            user_id=it_user.user_id,
            title=notification_title,
            message=notification_message,
            notification_type="request_restored",
            request_id=request_id
        )
    
    # Emit real-time notification to all users after creating database notifications
    if it_users_to_notify:
        try:
            socketio.emit('new_notification', {
                'title': notification_title,
                'message': notification_message,
                'type': 'request_restored',
                'request_id': request_id
            }, room='all_users')
            
            # Also emit a general update event to trigger notification count updates
            socketio.emit('notification_update', {
                'action': 'new_notification',
                'type': 'request_restored'
            }, room='all_users')
            
            print(f"DEBUG: WebSocket events emitted for request_restored")
        except Exception as e:
            print(f"Error emitting WebSocket notification: {e}")
    
    flash(f'Payment request #{request_id} has been restored.', 'success')
    return redirect(url_for('archives'))


@app.route('/request/<int:request_id>/delete_permanently', methods=['POST'])
@login_required
@role_required('IT Staff', 'Department Manager')
def delete_request_permanently(request_id):
    """Permanently delete a payment request from the database (IT only)"""
    # Restrict Department Managers to IT department only
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to perform this action.', 'danger')
        return redirect(url_for('dashboard'))
    
    req = PaymentRequest.query.get_or_404(request_id)
    
    # Check if request is archived
    if not req.is_archived:
        flash(f'Payment request #{request_id} must be archived before it can be permanently deleted.', 'warning')
        return redirect(url_for('archives'))
    
    try:
        # Store request info for notifications before deletion
        requestor_name = req.requestor_name
        request_id_val = request_id
        
        # Notify all IT department users about the permanent deletion (same as archive notifications)
        it_users = User.query.filter(
            db.or_(
                User.role == 'IT Staff',
                db.and_(User.role == 'Department Manager', User.department == 'IT')
            )
        ).all()
        
        # Exclude the user who deleted (they already know)
        it_users_to_notify = [user for user in it_users if user.user_id != current_user.user_id]
        
        notification_title = "Payment Request Permanently Deleted"
        notification_message = f"Payment request #{request_id_val} submitted by {requestor_name} has been permanently deleted from the database by {current_user.name}. This action cannot be undone."
        
        # Delete old notifications for this request first (before creating new ones)
        Notification.query.filter_by(request_id=request_id).delete()
        
        # Create new "permanently deleted" notifications BEFORE deleting the request (so request_id is still valid)
        for it_user in it_users_to_notify:
            create_notification(
                user_id=it_user.user_id,
                title=notification_title,
                message=notification_message,
                notification_type="request_permanently_deleted",
                request_id=request_id_val
            )
        
        # Commit notifications before deleting the request
        db.session.commit()
        
        # Delete all related records (cascade deletes)
        
        # 1. Get schedule IDs before deleting schedules (for edit history cleanup)
        schedule_ids = db.session.query(RecurringPaymentSchedule.schedule_id).filter_by(request_id=request_id).all()
        schedule_ids_list = [sid[0] for sid in schedule_ids] if schedule_ids else []
        
        # 2. Delete InstallmentEditHistory entries (via schedule relationship)
        if schedule_ids_list:
            InstallmentEditHistory.query.filter(InstallmentEditHistory.schedule_id.in_(schedule_ids_list)).delete()
        
        # 3. Delete RecurringPaymentSchedule entries
        RecurringPaymentSchedule.query.filter_by(request_id=request_id).delete()
        
        # 4. Delete PaidNotification entries
        PaidNotification.query.filter_by(request_id=request_id).delete()
        
        # 5. Delete LateInstallment entries
        LateInstallment.query.filter_by(request_id=request_id).delete()
        
        # 6. Delete FinanceAdminNote entries
        FinanceAdminNote.query.filter_by(request_id=request_id).delete()
        
        # 7. Delete the PaymentRequest itself (notifications were already handled above)
        db.session.delete(req)
        
        # Commit all deletions
        db.session.commit()
        
        log_action(f"Permanently deleted payment request #{request_id_val} from database")
        
        # Emit real-time notification to IT users after creating database notifications
        if it_users_to_notify:
            try:
                socketio.emit('new_notification', {
                    'title': notification_title,
                    'message': notification_message,
                    'type': 'request_permanently_deleted',
                    'request_id': request_id_val
                }, room='all_users')
                
                socketio.emit('notification_update', {
                    'action': 'new_notification',
                    'type': 'request_permanently_deleted'
                }, room='all_users')
            except Exception as e:
                print(f"DEBUG: Error emitting WebSocket notification: {e}")
        
        # Emit real-time update to remove from all dashboards
        try:
            socketio.emit('request_deleted', {
                'request_id': request_id_val
            }, room='all_users')
        except Exception as e:
            print(f"DEBUG: Error emitting request_deleted event: {e}")
        
        flash(f'Payment request #{request_id_val} has been permanently deleted from the database.', 'success')
        return redirect(url_for('archives'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting request: {str(e)}', 'error')
        app.logger.error(f"Error permanently deleting request #{request_id}: {str(e)}")
        return redirect(url_for('archives'))


# ==================== REQUEST TYPES MANAGEMENT ROUTES ====================

@app.route('/it/request-types')
@login_required
@role_required('IT Staff', 'Department Manager')
def manage_request_types():
    """Manage request types for all departments - IT only"""
    # Restrict Department Managers to IT department only
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    department_filter = request.args.get('department', '')
    search_query = request.args.get('search', '')
    
    # Validate per_page to prevent abuse
    if per_page not in [10, 20, 50, 100]:
        per_page = 20
    
    # Build query
    query = RequestType.query
    
    if department_filter:
        if department_filter == 'Management':
            query = query.filter(RequestType.department.in_(['Management', 'General Manager']))
        else:
            query = query.filter(RequestType.department == department_filter)
    
    if search_query:
        query = query.filter(RequestType.name.contains(search_query))
    
    # Get paginated results
    request_types_pagination = query.order_by(RequestType.id).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Get all departments for filter dropdown (normalize legacy values)
    departments = db.session.query(RequestType.department).distinct().all()
    departments = [('Management' if dept[0] == 'General Manager' else dept[0]) for dept in departments]
    
    return render_template('manage_request_types.html',
                         request_types=request_types_pagination.items,
                         pagination=request_types_pagination,
                         departments=departments,
                         department_filter=department_filter,
                         search_query=search_query,
                         user=current_user)


@app.route('/it/request-types/add', methods=['GET', 'POST'])
@login_required
@role_required('IT Staff', 'Department Manager')
def add_request_type():
    """Add new request type - IT only"""
    # Restrict Department Managers to IT department only
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        department = request.form.get('department', '').strip()
        is_active = request.form.get('is_active') == 'on'
        
        if not name or not department:
            flash('Name and department are required.', 'danger')
            return redirect(url_for('add_request_type'))
        
        # Check if request type already exists for this department
        existing = RequestType.query.filter_by(name=name, department=department).first()
        if existing:
            flash(f'Request type "{name}" already exists for {department} department.', 'danger')
            return redirect(url_for('add_request_type'))
        
        try:
            request_type = RequestType(
                name=name,
                department=department,
                is_active=is_active,
                created_by_user_id=current_user.user_id
            )
            
            db.session.add(request_type)
            db.session.commit()
            
            # Log the action
            log_action(f'Added request type: {name} for {department} department')
            
            flash(f'Request type "{name}" added successfully for {department} department.', 'success')
            return redirect(url_for('manage_request_types'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding request type: {str(e)}', 'danger')
            return redirect(url_for('add_request_type'))
    
    # Get all departments for dropdown
    departments = ['Management', 'Finance', 'Operation', 'PR', 'Maintenance', 'Marketing', 
                   'Logistic', 'HR', 'Quality Control', 'Procurement', 'IT', 'Customer Service', 
                   'Project']
    
    return render_template('add_request_type.html', departments=departments, user=current_user)


@app.route('/it/request-types/edit/<int:request_type_id>', methods=['GET', 'POST'])
@login_required
@role_required('IT Staff', 'Department Manager')
def edit_request_type(request_type_id):
    """Edit request type - IT only"""
    # Restrict Department Managers to IT department only
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    
    request_type = RequestType.query.get_or_404(request_type_id)
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        department = request.form.get('department', '').strip()
        is_active = request.form.get('is_active') == 'on'
        
        if not name or not department:
            flash('Name and department are required.', 'danger')
            return redirect(url_for('edit_request_type', request_type_id=request_type_id))
        
        # Check if request type already exists for this department (excluding current one)
        existing = RequestType.query.filter(
            RequestType.name == name,
            RequestType.department == department,
            RequestType.id != request_type_id
        ).first()
        
        if existing:
            flash(f'Request type "{name}" already exists for {department} department.', 'danger')
            return redirect(url_for('edit_request_type', request_type_id=request_type_id))
        
        try:
            old_name = request_type.name
            old_department = request_type.department
            
            request_type.name = name
            request_type.department = department
            request_type.is_active = is_active
            request_type.updated_at = datetime.utcnow()
            
            db.session.commit()
            
            # Log the action
            log_action(f'Updated request type: {old_name} ({old_department}) to {name} ({department})')
            
            flash(f'Request type updated successfully.', 'success')
            return redirect(url_for('manage_request_types'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating request type: {str(e)}', 'danger')
            return redirect(url_for('edit_request_type', request_type_id=request_type_id))
    
    # Get all departments for dropdown
    departments = ['Management', 'Finance', 'Operation', 'PR', 'Maintenance', 'Marketing', 
                   'Logistic', 'HR', 'Quality Control', 'Procurement', 'IT', 'Customer Service', 
                   'Project']
    
    return render_template('edit_request_type.html', 
                         request_type=request_type, 
                         departments=departments, 
                         user=current_user)


@app.route('/it/request-types/delete/<int:request_type_id>', methods=['POST'])
@login_required
@role_required('IT Staff', 'Department Manager')
def delete_request_type(request_type_id):
    """Delete request type - IT only"""
    # Restrict Department Managers to IT department only
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    
    request_type = RequestType.query.get_or_404(request_type_id)
    
    try:
        # Check if any payment requests are using this request type
        existing_requests = PaymentRequest.query.filter_by(request_type=request_type.name).first()
        if existing_requests:
            flash(f'Cannot delete request type "{request_type.name}" because it is being used by existing payment requests.', 'danger')
            return redirect(url_for('it_dashboard', tab='request-types'))
        
        # Log the action before deletion
        log_action(f'Deleted request type: {request_type.name} ({request_type.department})')
        
        db.session.delete(request_type)
        db.session.commit()
        
        flash(f'Request type "{request_type.name}" deleted successfully.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting request type: {str(e)}', 'danger')
    
    # Redirect back to IT dashboard with request-types tab active
    return redirect(url_for('it_dashboard', tab='request-types'))


@app.route('/it/request-types/bulk-delete', methods=['POST'])
@login_required
@role_required('IT Staff', 'Department Manager')
def bulk_delete_request_types():
    """Bulk delete request types - IT only"""
    # Restrict Department Managers to IT department only
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    
    request_type_ids = request.form.getlist('request_type_ids')
    
    if not request_type_ids:
        flash('No request types selected for deletion.', 'warning')
        return redirect(url_for('it_dashboard', tab='request-types'))
    
    deleted_count = 0
    failed_deletions = []
    
    try:
        for request_type_id in request_type_ids:
            request_type = RequestType.query.get(request_type_id)
            if not request_type:
                failed_deletions.append(f"Request type with ID {request_type_id} not found")
                continue
            
            # Check if any payment requests are using this request type
            existing_requests = PaymentRequest.query.filter_by(request_type=request_type.name).first()
            if existing_requests:
                failed_deletions.append(f'"{request_type.name}" (used by existing payment requests)')
                continue
            
            # Log the action before deletion
            log_action(f'Bulk deleted request type: {request_type.name} ({request_type.department})')
            
            db.session.delete(request_type)
            deleted_count += 1
        
        if deleted_count > 0:
            db.session.commit()
            flash(f'Successfully deleted {deleted_count} request type(s).', 'success')
        
        if failed_deletions:
            flash(f'Could not delete: {", ".join(failed_deletions)}', 'warning')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error during bulk deletion: {str(e)}', 'danger')
    
    return redirect(url_for('it_dashboard', tab='request-types'))


@app.route('/it/request-types/toggle/<int:request_type_id>', methods=['POST'])
@login_required
@role_required('IT Staff', 'Department Manager')
def toggle_request_type(request_type_id):
    """Toggle request type active status - IT only"""
    # Restrict Department Managers to IT department only
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    
    request_type = RequestType.query.get_or_404(request_type_id)
    
    try:
        request_type.is_active = not request_type.is_active
        request_type.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        status = 'activated' if request_type.is_active else 'deactivated'
        log_action(f'{status.title()} request type: {request_type.name} ({request_type.department})')
        
        flash(f'Request type "{request_type.name}" {status} successfully.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating request type: {str(e)}', 'danger')
    
    return redirect(url_for('manage_request_types'))


# Branch Management Routes
@app.route('/it/branches')
@login_required
@role_required('IT Staff', 'Department Manager')
def manage_branches():
    """Manage branches - IT only"""
    # Restrict Department Managers to IT department only
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search_query = request.args.get('search', '')
    restaurant_filter = request.args.get('restaurant', '')
    
    # Validate per_page to prevent abuse
    if per_page not in [10, 20, 50, 100]:
        per_page = 20
    
    # Build query
    query = Branch.query
    
    if search_query:
        query = query.filter(Branch.name.contains(search_query))
    
    if restaurant_filter:
        query = query.filter(Branch.restaurant == restaurant_filter)
    
    # Get paginated results
    branches_pagination = query.order_by(Branch.id).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return render_template('manage_branches.html',
                         branches=branches_pagination.items,
                         pagination=branches_pagination,
                         search_query=search_query,
                         restaurant_filter=restaurant_filter,
                         user=current_user)


@app.route('/it/branches/add', methods=['GET', 'POST'])
@login_required
@role_required('IT Staff', 'Department Manager')
def add_branch():
    """Add new branch - IT only"""
    # Restrict Department Managers to IT department only
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        restaurant = request.form.get('restaurant', '').strip()
        is_active = request.form.get('is_active') == 'on'
        
        if not restaurant:
            flash('Location is required.', 'danger')
            return redirect(url_for('add_branch'))
        
        if not name:
            flash('Name is required.', 'danger')
            return redirect(url_for('add_branch'))
        
        # Check if branch already exists
        existing = Branch.query.filter_by(name=name).first()
        if existing:
            flash(f'Branch "{name}" already exists.', 'danger')
            return redirect(url_for('add_branch'))
        
        try:
            branch = Branch(
                name=name,
                restaurant=restaurant,
                is_active=is_active,
                created_by_user_id=current_user.user_id
            )
            
            db.session.add(branch)
            db.session.commit()
            
            # Log the action
            log_action(f'Added branch: {name}')
            
            flash(f'Branch "{name}" added successfully.', 'success')
            return redirect(url_for('manage_branches'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding branch: {str(e)}', 'danger')
            return redirect(url_for('add_branch'))
    
    return render_template('add_branch.html', user=current_user)


@app.route('/it/branches/edit/<int:branch_id>', methods=['GET', 'POST'])
@login_required
@role_required('IT Staff', 'Department Manager')
def edit_branch(branch_id):
    """Edit branch - IT only"""
    # Restrict Department Managers to IT department only
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    
    branch = Branch.query.get_or_404(branch_id)
    
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        restaurant = request.form.get('restaurant', '').strip()
        is_active = request.form.get('is_active') == 'on'
        
        if not restaurant:
            flash('Location is required.', 'danger')
            return redirect(url_for('edit_branch', branch_id=branch_id))
        
        if not name:
            flash('Name is required.', 'danger')
            return redirect(url_for('edit_branch', branch_id=branch_id))
        
        # Check if another branch with same name exists
        existing = Branch.query.filter(Branch.name == name, Branch.id != branch_id).first()
        if existing:
            flash(f'Branch "{name}" already exists.', 'danger')
            return redirect(url_for('edit_branch', branch_id=branch_id))
        
        try:
            old_name = branch.name
            old_restaurant = branch.restaurant
            
            branch.name = name
            branch.restaurant = restaurant
            branch.is_active = is_active
            branch.updated_at = datetime.utcnow()
            
            db.session.commit()
            
            # Log the action
            log_action(f'Updated branch: {old_name} ({old_restaurant}) -> {name} ({restaurant})')
            
            flash(f'Branch "{name}" updated successfully.', 'success')
            return redirect(url_for('manage_branches'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating branch: {str(e)}', 'danger')
            return redirect(url_for('edit_branch', branch_id=branch_id))
    
    return render_template('edit_branch.html', branch=branch, user=current_user)


@app.route('/it/branches/<int:branch_id>/aliases/add', methods=['POST'])
@login_required
@role_required('IT Staff', 'Department Manager')
def add_branch_alias(branch_id):
    """Add an alias name for a branch"""
    # Restrict Department Managers to IT department only
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to perform this action.', 'danger')
        return redirect(url_for('dashboard'))

    branch = Branch.query.get_or_404(branch_id)
    alias_name = (request.form.get('alias_name') or '').strip()
    if not alias_name:
        flash('Alias name is required.', 'danger')
        return redirect(url_for('edit_branch', branch_id=branch_id))

    # Prevent duplicate alias for the same branch or alias equal to current name
    existing = BranchAlias.query.filter(
        db.func.lower(BranchAlias.alias_name) == alias_name.lower(),
        BranchAlias.branch_id == branch_id
    ).first()
    if existing or alias_name.lower() == (branch.name or '').lower():
        flash('This alias already exists or matches the current branch name.', 'warning')
        return redirect(url_for('edit_branch', branch_id=branch_id))

    try:
        new_alias = BranchAlias(branch_id=branch_id, alias_name=alias_name)
        db.session.add(new_alias)
        db.session.commit()
        log_action(f'Added alias "{alias_name}" for branch {branch.name} (ID {branch.id})')
        flash('Alias added successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding alias: {str(e)}', 'danger')

    return redirect(url_for('edit_branch', branch_id=branch_id))


@app.route('/it/branches/<int:branch_id>/aliases/<int:alias_id>/delete', methods=['POST'])
@login_required
@role_required('IT Staff', 'Department Manager')
def delete_branch_alias(branch_id, alias_id):
    """Delete an alias from a branch"""
    # Restrict Department Managers to IT department only
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to perform this action.', 'danger')
        return redirect(url_for('dashboard'))

    alias = BranchAlias.query.get_or_404(alias_id)
    if alias.branch_id != branch_id:
        flash('Alias does not belong to the specified branch.', 'danger')
        return redirect(url_for('edit_branch', branch_id=branch_id))

    try:
        db.session.delete(alias)
        db.session.commit()
        log_action(f'Deleted alias "{alias.alias_name}" from branch ID {branch_id}')
        flash('Alias deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting alias: {str(e)}', 'danger')

    return redirect(url_for('edit_branch', branch_id=branch_id))

@app.route('/it/branches/delete/<int:branch_id>', methods=['POST'])
@login_required
@role_required('IT Staff', 'Department Manager')
def delete_branch(branch_id):
    """Delete branch - IT only"""
    # Restrict Department Managers to IT department only
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    
    branch = Branch.query.get_or_404(branch_id)
    
    try:
        branch_name = branch.name
        db.session.delete(branch)
        db.session.commit()
        
        # Log the action
        log_action(f'Deleted branch: {branch_name}')
        
        flash(f'Branch "{branch_name}" deleted successfully.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting branch: {str(e)}', 'danger')
    
    return redirect(url_for('manage_branches'))


@app.route('/it/branches/bulk-delete', methods=['POST'])
@login_required
@role_required('IT Staff', 'Department Manager')
def bulk_delete_branches():
    """Bulk delete branches - IT only"""
    # Restrict Department Managers to IT department only
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    
    branch_ids = request.form.getlist('branch_ids')
    
    if not branch_ids:
        flash('No branches selected for deletion.', 'warning')
        return redirect(url_for('manage_branches'))
    
    try:
        deleted_count = 0
        for branch_id in branch_ids:
            branch = Branch.query.get(branch_id)
            if branch:
                db.session.delete(branch)
                deleted_count += 1
        
        db.session.commit()
        
        # Log the action
        log_action(f'Bulk deleted {deleted_count} branches')
        
        flash(f'{deleted_count} branch(es) deleted successfully.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting branches: {str(e)}', 'danger')
    
    return redirect(url_for('manage_branches'))


@app.route('/it/branches/toggle/<int:branch_id>', methods=['POST'])
@login_required
@role_required('IT Staff', 'Department Manager')
def toggle_branch(branch_id):
    """Toggle branch active status - IT only"""
    # Restrict Department Managers to IT department only
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    
    branch = Branch.query.get_or_404(branch_id)
    
    try:
        branch.is_active = not branch.is_active
        branch.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        status = 'activated' if branch.is_active else 'deactivated'
        log_action(f'{status.title()} branch: {branch.name}')
        
        flash(f'Branch "{branch.name}" {status} successfully.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating branch: {str(e)}', 'danger')
    
    return redirect(url_for('manage_branches'))


@app.route('/it/backup-database')
@login_required
@role_required('IT Staff', 'Department Manager')
def backup_database():
    """Create a backup of the database and download it"""
    try:
        # Restrict Department Managers to IT department only
        if current_user.role == 'Department Manager' and current_user.department != 'IT':
            flash('You do not have permission to perform this action.', 'danger')
            return redirect(url_for('dashboard'))
        
        # Get the database path directly from config
        from config import Config
        db_path = Config._DB_PATH
        
        # Check if database file exists
        if not os.path.exists(db_path):
            flash(f'Database file not found at: {db_path}', 'danger')
            return redirect(url_for('it_dashboard'))
        
        # Create backups directory if it doesn't exist (same directory as database)
        backups_dir = os.path.join(os.path.dirname(db_path), 'backups')
        os.makedirs(backups_dir, exist_ok=True)
        
        # Generate backup filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f'payment_system_backup_{timestamp}.db'
        backup_path = os.path.join(backups_dir, backup_filename)
        
        # Copy the database file using shutil
        import shutil
        shutil.copy2(db_path, backup_path)
        
        # Verify backup was created
        if not os.path.exists(backup_path):
            flash('Failed to create backup file.', 'danger')
            return redirect(url_for('it_dashboard'))
        
        # Log the backup action
        log_action(f"Database backup created: {backup_filename}")
        
        # Send the backup file for download
        return send_file(
            backup_path,
            as_attachment=True,
            download_name=backup_filename,
            mimetype='application/x-sqlite3'
        )
        
    except Exception as e:
        import traceback
        error_msg = f'Error creating backup: {str(e)}'
        print(f"Backup error: {error_msg}")
        print(traceback.format_exc())
        flash(error_msg, 'danger')
        try:
            log_action(f"Database backup failed: {str(e)}")
        except:
            pass
        return redirect(url_for('it_dashboard'))


@app.route('/project/dashboard')
@login_required
@role_required('Project Staff', 'Department Manager')
def project_dashboard():
    """Dashboard for project users - can request payments and view due dates"""
    # Ensure only Project department users can access this dashboard
    if current_user.role == 'Department Manager' and current_user.department != 'Project':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('department_dashboard'))
    
    # Check for recurring payments due today
    check_recurring_payments_due()
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    status_filter = request.args.get('status', None)
    search_query = request.args.get('search', None)
    urgent_filter = request.args.get('urgent', None)
    tab = request.args.get('tab', 'all')  # 'all' tab shows all requests
    
    # Validate per_page to prevent abuse
    if per_page not in [10, 20, 50, 100]:
        per_page = 10
    
    # Build query with optional status and search filters
    # Exclude archived requests from all queries
    if current_user.role == 'Project Staff':
        # Project Staff see only their own requests
        query = PaymentRequest.query.filter(
            PaymentRequest.user_id == current_user.user_id,
            PaymentRequest.is_archived == False
        )
    elif current_user.role == 'Department Manager' and current_user.department == 'Project':
        # Project Manager sees all requests from Project department
        query = PaymentRequest.query.filter(
            PaymentRequest.department == 'Project',
            PaymentRequest.is_archived == False
        )
    else:
        # Fallback - should not happen due to role_required decorator
        query = PaymentRequest.query.filter(
            PaymentRequest.user_id == current_user.user_id,
            PaymentRequest.is_archived == False
        )
    # Only apply status_filter for non-all tabs (All Requests tab shows all statuses)
    if status_filter and tab != 'all':
        query = query.filter(PaymentRequest.status == status_filter)
    if search_query:
        # Search by request ID or requestor name
        try:
            # Try to convert to integer for exact match
            search_id = int(search_query)
            query = query.filter(PaymentRequest.request_id == search_id)
        except ValueError:
            # If not a number, search by requestor name or other text fields
            search_term = f'%{search_query}%'
            query = query.filter(
                db.or_(
                    PaymentRequest.requestor_name.ilike(search_term),
                    PaymentRequest.purpose.ilike(search_term),
                    PaymentRequest.account_name.ilike(search_term)
                )
            )
    if urgent_filter:
        if urgent_filter == 'urgent':
            query = query.filter(PaymentRequest.is_urgent == True)
        elif urgent_filter == 'not_urgent':
            query = query.filter(PaymentRequest.is_urgent == False)
    
    # Apply tab-based filtering
    if tab == 'completed':
        query = query.filter(PaymentRequest.status.in_(['Completed', 'Paid', 'Approved']))
    elif tab == 'rejected':
        query = query.filter(PaymentRequest.status.in_(['Rejected by Manager', 'Rejected by Finance', 'Proof Rejected']))
    elif tab == 'recurring':
        query = query.filter(PaymentRequest.recurring == 'Recurring')
    elif tab == 'all':
        # 'all' tab (All Requests) shows all requests (no additional filtering)
        pass
    # Default case also shows all requests
    
    # Get paginated requests
    # For 'all' tab, sort by status priority then by date (Completed by completion_date, others by created_at)
    if tab == 'all':
        requests_pagination = query.order_by(
            get_status_priority_order(),
            get_all_tab_datetime_order()
        ).paginate(
            page=page, per_page=per_page, error_out=False
        )
    elif tab == 'completed':
        requests_pagination = query.order_by(get_completed_datetime_order()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    else:
        requests_pagination = query.order_by(PaymentRequest.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    
    # Get notifications for project users (only due date notifications)
    notifications = get_notifications_for_user(current_user)
    unread_count = get_unread_count_for_user(current_user)
    
    # Get separate queries for completed, rejected, and recurring requests for tab content
    # Exclude archived requests from all queries
    if current_user.role == 'Project Staff':
        # Project Staff see only their own requests
        completed_query = PaymentRequest.query.filter(
            PaymentRequest.user_id == current_user.user_id,
            PaymentRequest.is_archived == False
        )
        rejected_query = PaymentRequest.query.filter(
            PaymentRequest.user_id == current_user.user_id,
            PaymentRequest.is_archived == False
        )
        recurring_query = PaymentRequest.query.filter(
            PaymentRequest.user_id == current_user.user_id,
            PaymentRequest.is_archived == False
        )
    elif current_user.role == 'Department Manager' and current_user.department == 'Project':
        # Project Manager sees all requests from Project department
        completed_query = PaymentRequest.query.filter(
            PaymentRequest.department == 'Project',
            PaymentRequest.is_archived == False
        )
        rejected_query = PaymentRequest.query.filter(
            PaymentRequest.department == 'Project',
            PaymentRequest.is_archived == False
        )
        recurring_query = PaymentRequest.query.filter(
            PaymentRequest.department == 'Project',
            PaymentRequest.is_archived == False
        )
    else:
        # Fallback - should not happen due to role_required decorator
        completed_query = PaymentRequest.query.filter(
            PaymentRequest.user_id == current_user.user_id,
            PaymentRequest.is_archived == False
        )
        rejected_query = PaymentRequest.query.filter(
            PaymentRequest.user_id == current_user.user_id,
            PaymentRequest.is_archived == False
        )
        recurring_query = PaymentRequest.query.filter(
            PaymentRequest.user_id == current_user.user_id,
            PaymentRequest.is_archived == False
        )
    
    completed_query = completed_query.filter(PaymentRequest.status == 'Completed')
    rejected_query = rejected_query.filter(PaymentRequest.status.in_(['Rejected by Manager', 'Rejected by Finance', 'Proof Rejected']))
    recurring_query = recurring_query.filter(PaymentRequest.recurring == 'Recurring')
    
    completed_requests = completed_query.order_by(get_completed_datetime_order()).all()
    rejected_requests = rejected_query.order_by(PaymentRequest.created_at.desc()).all()
    recurring_requests = recurring_query.order_by(PaymentRequest.created_at.desc()).all()
    
    return render_template('project_dashboard.html', 
                         requests=requests_pagination.items, 
                         pagination=requests_pagination,
                         user=current_user,
                         notifications=notifications,
                         unread_count=unread_count,
                         status_filter=status_filter,
                         search_query=search_query,
                         urgent_filter=urgent_filter,
                         completed_requests=completed_requests,
                         rejected_requests=rejected_requests,
                         recurring_requests=recurring_requests,
                         active_tab=tab)


@app.route('/operation/dashboard')
@login_required
@role_required('Operation Manager')
def operation_dashboard():
    """Dashboard for operation manager - can view all requests but only in dashboard"""
    # Check for recurring payments due today
    check_recurring_payments_due()
    
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    status_filter = request.args.get('status', None)
    department_filter = request.args.get('department', None)
    search_query = request.args.get('search', None)
    urgent_filter = request.args.get('urgent', None)
    tab = request.args.get('tab', 'all')  # 'all' tab shows all requests
    
    # Validate per_page to prevent abuse
    if per_page not in [10, 20, 50, 100]:
        per_page = 10
    
    # Build query with optional status, department, and search filters - Operation Manager sees ALL departments
    # Exclude archived requests
    query = PaymentRequest.query.filter(PaymentRequest.is_archived == False)
    
    if department_filter:
        query = query.filter(PaymentRequest.department == department_filter)
    if search_query:
        # Search by request ID or requestor name
        try:
            # Try to convert to integer for exact match
            search_id = int(search_query)
            query = query.filter(PaymentRequest.request_id == search_id)
        except ValueError:
            # If not a number, search by requestor name or other text fields
            search_term = f'%{search_query}%'
            query = query.filter(
                db.or_(
                    PaymentRequest.requestor_name.ilike(search_term),
                    PaymentRequest.purpose.ilike(search_term),
                    PaymentRequest.account_name.ilike(search_term)
                )
            )
    if urgent_filter:
        if urgent_filter == 'urgent':
            query = query.filter(PaymentRequest.is_urgent == True)
        elif urgent_filter == 'not_urgent':
            query = query.filter(PaymentRequest.is_urgent == False)
    
    # Apply tab-based filtering
    if tab == 'completed':
        query = query.filter(PaymentRequest.status.in_(['Completed', 'Paid', 'Approved']))
    elif tab == 'rejected':
        query = query.filter(PaymentRequest.status.in_(['Rejected by Manager', 'Rejected by Finance', 'Proof Rejected']))
    elif tab == 'recurring':
        query = query.filter(PaymentRequest.status == 'Recurring')
    elif tab == 'my_requests':
        # For 'my_requests' tab, show only the current user's requests
        query = query.filter(PaymentRequest.user_id == current_user.user_id)
    elif tab == 'all':
        # 'all' tab - apply status filter if provided (only on all tab)
        if status_filter:
            query = query.filter(PaymentRequest.status == status_filter)
    
    # Get paginated requests
    # For 'all' tab, sort by status priority then by per-status recency
    if tab == 'all':
        requests_pagination = query.order_by(
            get_status_priority_order(),
            get_all_tab_datetime_order()
        ).paginate(
            page=page, per_page=per_page, error_out=False
        )
    elif tab == 'completed':
        requests_pagination = query.order_by(get_completed_datetime_order()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    elif tab == 'rejected':
        requests_pagination = query.order_by(get_rejected_datetime_order()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    elif tab == 'recurring':
        requests_pagination = query.order_by(get_recurring_datetime_order()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    else:
        requests_pagination = query.order_by(PaymentRequest.created_at.desc()).paginate(
            page=page, per_page=per_page, error_out=False
        )
    
    # Get notifications for operation manager (all notifications, same as admin)
    notifications = get_notifications_for_user(current_user)
    unread_count = get_unread_count_for_user(current_user)
    
    # Get separate queries for completed and rejected requests for tab content (exclude archived)
    completed_query = PaymentRequest.query.filter(PaymentRequest.is_archived == False)
    rejected_query = PaymentRequest.query.filter(PaymentRequest.is_archived == False)
    
    if department_filter:
        completed_query = completed_query.filter(PaymentRequest.department == department_filter)
        rejected_query = rejected_query.filter(PaymentRequest.department == department_filter)
    
    completed_query = completed_query.filter(PaymentRequest.status == 'Completed')
    rejected_query = rejected_query.filter(PaymentRequest.status.in_(['Rejected by Manager', 'Rejected by Finance', 'Proof Rejected']))
    
    completed_requests = completed_query.order_by(get_completed_datetime_order()).all()
    rejected_requests = rejected_query.order_by(get_rejected_datetime_order()).all()
    
    # Get user's own requests for the My Requests tab (exclude archived)
    my_requests_query = PaymentRequest.query.filter(
        PaymentRequest.user_id == current_user.user_id,
        PaymentRequest.is_archived == False
    )
    my_requests_pagination = my_requests_query.order_by(PaymentRequest.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return render_template('operation_dashboard.html', 
                         requests=requests_pagination.items, 
                         pagination=requests_pagination,
                         user=current_user,
                         notifications=notifications,
                         unread_count=unread_count,
                         status_filter=status_filter,
                         department_filter=department_filter,
                         search_query=search_query,
                         urgent_filter=urgent_filter,
                         completed_requests=completed_requests,
                         rejected_requests=rejected_requests,
                         my_requests=my_requests_pagination.items,
                         active_tab=tab)


# ==================== PAYMENT REQUEST ROUTES ====================

def get_available_request_types():
    """Helper function to get available request types for the current user"""
    # Get user's department and role
    user_department = current_user.department
    user_role = current_user.role
    
    # Query request types based on user's department and role
    if user_role in ['GM', 'CEO']:
        # Management roles see request types under the Management department
        return RequestType.query.filter(
            RequestType.department == 'Management',
            RequestType.is_active == True
        ).order_by(RequestType.id).all()
    elif user_role in ['Finance Admin', 'Finance Staff']:
        # Finance users can see Finance department request types
        return RequestType.query.filter(
            RequestType.department == 'Finance',
            RequestType.is_active == True
        ).order_by(RequestType.id).all()
    elif user_role == 'Operation Manager':
        # Operation Manager can see Operation and Project request types
        return RequestType.query.filter(
            RequestType.department.in_(['Operation', 'Project']),
            RequestType.is_active == True
        ).order_by(RequestType.id).all()
    elif user_role == 'Project Staff':
        # Project Staff can see Project request types
        return RequestType.query.filter(
            RequestType.department == 'Project',
            RequestType.is_active == True
        ).order_by(RequestType.id).all()
    elif user_role == 'Department Manager':
        # Department Managers can see their department's request types
        return RequestType.query.filter(
            RequestType.department == user_department,
            RequestType.is_active == True
        ).order_by(RequestType.id).all()
    else:
        # Other staff roles can see their department's request types
        return RequestType.query.filter(
            RequestType.department == user_department,
            RequestType.is_active == True
        ).order_by(RequestType.id).all()

@app.route('/request/new', methods=['GET', 'POST'])
@login_required
def new_request():
    """Create a new payment request"""
    if request.method == 'POST':
        request_type = request.form.get('request_type')
        requestor_name = request.form.get('requestor_name')
        branch_name = request.form.get('branch_name')
        date = datetime.utcnow().date()  # Automatically use today's date
        purpose = request.form.get('purpose')
        
        # Validate required fields
        if not branch_name:
            flash('Branch name is required.', 'error')
            available_request_types = get_available_request_types()
            # Custom order: Office, Kucu, Boom, Thoum, Kitchen
            from sqlalchemy import case
            location_order = case(
                (Branch.restaurant == 'Office', 1),
                (Branch.restaurant == 'Kucu', 2),
                (Branch.restaurant == 'Boom', 3),
                (Branch.restaurant == 'Thoum', 4),
                (Branch.restaurant == 'Kitchen', 5),
                else_=6
            )
            available_branches = Branch.query.filter_by(is_active=True).order_by(location_order, Branch.name).all()
            return render_template('new_request.html', user=current_user, today=datetime.utcnow().date().strftime('%Y-%m-%d'), available_request_types=available_request_types, available_branches=available_branches)
        account_name = request.form.get('account_name')
        account_number = request.form.get('account_number')
        bank_name = request.form.get('bank_name')
        amount = request.form.get('amount')
        recurring = request.form.get('recurring', 'One-Time')
        recurring_interval = request.form.get('recurring_interval')
        
        # Validate account number length (maximum 16 digits)
        if account_number and len(account_number) > 16:
            flash('Account number cannot exceed 16 digits.', 'error')
            available_request_types = get_available_request_types()
            # Custom order: Office, Kucu, Boom, Thoum, Kitchen
            from sqlalchemy import case
            location_order = case(
                (Branch.restaurant == 'Office', 1),
                (Branch.restaurant == 'Kucu', 2),
                (Branch.restaurant == 'Boom', 3),
                (Branch.restaurant == 'Thoum', 4),
                (Branch.restaurant == 'Kitchen', 5),
                else_=6
            )
            available_branches = Branch.query.filter_by(is_active=True).order_by(location_order, Branch.name).all()
            return render_template('new_request.html', user=current_user, today=datetime.utcnow().date().strftime('%Y-%m-%d'), available_request_types=available_request_types, available_branches=available_branches)
        
        # Validate account number contains only digits
        if account_number and not account_number.isdigit():
            flash('Account number must contain only numbers.', 'error')
            available_request_types = get_available_request_types()
            # Custom order: Office, Kucu, Boom, Thoum, Kitchen
            from sqlalchemy import case
            location_order = case(
                (Branch.restaurant == 'Office', 1),
                (Branch.restaurant == 'Kucu', 2),
                (Branch.restaurant == 'Boom', 3),
                (Branch.restaurant == 'Thoum', 4),
                (Branch.restaurant == 'Kitchen', 5),
                else_=6
            )
            available_branches = Branch.query.filter_by(is_active=True).order_by(location_order, Branch.name).all()
            return render_template('new_request.html', user=current_user, today=datetime.utcnow().date().strftime('%Y-%m-%d'), available_request_types=available_request_types, available_branches=available_branches)
        
        # Validate bank name is selected
        if not bank_name:
            flash('Please select a bank name.', 'error')
            available_request_types = get_available_request_types()
            # Custom order: Office, Kucu, Boom, Thoum, Kitchen
            from sqlalchemy import case
            location_order = case(
                (Branch.restaurant == 'Office', 1),
                (Branch.restaurant == 'Kucu', 2),
                (Branch.restaurant == 'Boom', 3),
                (Branch.restaurant == 'Thoum', 4),
                (Branch.restaurant == 'Kitchen', 5),
                else_=6
            )
            available_branches = Branch.query.filter_by(is_active=True).order_by(location_order, Branch.name).all()
            return render_template('new_request.html', user=current_user, today=datetime.utcnow().date().strftime('%Y-%m-%d'), available_request_types=available_request_types, available_branches=available_branches)
        
        # Validate "Others" description if "Others" is selected
        if request_type == 'Others':
            others_description = request.form.get('others_description')
            if not others_description or not others_description.strip():
                flash('Please specify the type of request when selecting "Others".', 'error')
                available_request_types = get_available_request_types()
                return render_template('new_request.html', user=current_user, today=datetime.utcnow().date().strftime('%Y-%m-%d'), available_request_types=available_request_types)
        
        # Handle comma-formatted amount
        if amount:
            # Remove commas from amount for processing
            amount_clean = amount.replace(',', '')
            try:
                amount_float = float(amount_clean)
                if amount_float <= 0:
                    flash('Amount must be greater than 0.', 'error')
                    available_request_types = get_available_request_types()
                    return render_template('new_request.html', user=current_user, today=datetime.utcnow().date().strftime('%Y-%m-%d'), available_request_types=available_request_types)
            except ValueError:
                flash('Invalid amount format.', 'error')
                available_request_types = get_available_request_types()
                return render_template('new_request.html', user=current_user, today=datetime.utcnow().date().strftime('%Y-%m-%d'), available_request_types=available_request_types)
        
        # Handle multiple file uploads for receipts
        receipt_path = None  # Initialize receipt_path
        receipt_paths = []
        if 'receipt_files' in request.files:
            receipt_files = request.files.getlist('receipt_files')
            if receipt_files and any(f.filename for f in receipt_files):
                import uuid
                import json
                import os
                
                # Create uploads directory if it doesn't exist
                upload_folder = os.path.join(app.root_path, 'uploads', 'receipts')
                os.makedirs(upload_folder, exist_ok=True)
                
                allowed_extensions = {'pdf', 'jpg', 'jpeg', 'png', 'doc', 'docx', 'xls', 'xlsx'}
                
                for receipt_file in receipt_files:
                    if receipt_file and receipt_file.filename:
                        # Validate file size (50MB max)
                        max_file_size = app.config.get('MAX_FILE_SIZE', 50 * 1024 * 1024)
                        file_size = len(receipt_file.read())
                        if file_size > max_file_size:
                            file_size_mb = file_size / (1024 * 1024)
                            error_msg = f'File "{receipt_file.filename}" is too large. Maximum size is {max_file_size // (1024 * 1024)}MB. Your file size is {file_size_mb:.2f}MB.'
                            # Check if this is a fetch request (FormData submission)
                            # Fetch requests with FormData typically don't have Accept: text/html
                            if 'text/html' not in request.headers.get('Accept', ''):
                                return jsonify({'error': 'File too large', 'message': error_msg}), 400
                            flash(error_msg, 'error')
                            available_request_types = get_available_request_types()
                            return render_template('new_request.html', user=current_user, today=datetime.utcnow().date().strftime('%Y-%m-%d'), available_request_types=available_request_types)
                        
                        # Reset file pointer
                        receipt_file.seek(0)
                        
                        # Validate file extension
                        file_extension = receipt_file.filename.rsplit('.', 1)[1].lower() if '.' in receipt_file.filename else ''
                        if file_extension not in allowed_extensions:
                            error_msg = f'Invalid file type for "{receipt_file.filename}". Allowed types: PDF, JPG, PNG, DOC, DOCX, XLS, XLSX'
                            # Check if this is a fetch request (FormData submission)
                            if 'text/html' not in request.headers.get('Accept', ''):
                                return jsonify({'error': 'Invalid file type', 'message': error_msg}), 400
                            flash(error_msg, 'error')
                            available_request_types = get_available_request_types()
                            return render_template('new_request.html', user=current_user, today=datetime.utcnow().date().strftime('%Y-%m-%d'), available_request_types=available_request_types)
                        
                        # Generate unique filename
                        filename = f"{uuid.uuid4()}_{receipt_file.filename}"
                        
                        # Save file
                        full_path = os.path.join(upload_folder, filename)
                        receipt_file.save(full_path)
                        receipt_paths.append(filename)  # Store only the filename, not the full path
                
                # Convert list to JSON string for storage
                receipt_path = json.dumps(receipt_paths) if receipt_paths else None
        
        # Get dynamic fields based on request type
        item_name = request.form.get('item_name')
        person_company = request.form.get('person_company')
        company_name = request.form.get('company_name')
        others_description = request.form.get('others_description')
        
        # All departments go to their manager first for approval
        initial_status = 'Pending Manager Approval'
        
        # Create new request
        current_time = datetime.utcnow()
        
        # Handle "Others" request type with description
        final_request_type = request_type
        if request_type == 'Others' and others_description:
            final_request_type = f"Others: {others_description}"
        
        new_req = PaymentRequest(
            request_type=final_request_type,
            requestor_name=requestor_name,
            branch_name=branch_name,
            item_name=item_name if request_type == 'Item' else None,
            person_company=person_company if person_company else None,
            company_name=company_name if request_type == 'Supplier/Rental' else None,
            department=current_user.department,
            date=date,
            purpose=purpose,
            account_name=account_name,
            account_number=account_number,
            bank_name=bank_name,
            amount=amount_clean if amount else amount,  # Use cleaned amount without commas
            recurring=recurring,
            recurring_interval=recurring_interval if recurring == 'Recurring' else None,
            status=initial_status,
            requestor_receipt_path=receipt_path,  # Store requestor receipts in separate column
            user_id=current_user.user_id,
            # Start timing immediately when request is submitted
            manager_approval_start_time=current_time
        )
        
        
        db.session.add(new_req)
        db.session.commit()
        
        # Handle recurring payment schedules (both variable amounts and custom)
        if recurring == 'Recurring':
            recurring_interval = request.form.get('recurring_interval', '')
            print(f"🔧 DEBUG: Processing recurring payment - interval: {recurring_interval}")
            
            # Check if it's a custom payment schedule
            if recurring_interval.startswith('custom:'):
                print(f"🔧 DEBUG: Detected custom payment schedule")
                try:
                    # Parse custom payment schedule
                    custom_data = recurring_interval[7:]  # Remove 'custom:' prefix
                    print(f"🔧 DEBUG: Custom data: {custom_data}")
                    payment_schedule_data = []
                    
                    if custom_data:
                        # Split by comma to get individual date:amount pairs
                        date_amount_pairs = custom_data.split(',')
                        print(f"🔧 DEBUG: Date amount pairs: {date_amount_pairs}")
                        
                        for i, pair in enumerate(date_amount_pairs, 1):
                            if ':' in pair:
                                date_str, amount_str = pair.split(':', 1)
                                payment_schedule_data.append({
                                    'date': date_str,
                                    'amount': float(amount_str)
                                })
                                print(f"🔧 DEBUG: Added payment {i}: {date_str} - {amount_str}")
                    
                    print(f"🔧 DEBUG: Final payment schedule data: {payment_schedule_data}")
                    
                    # Create the payment schedule
                    if payment_schedule_data:
                        success = create_recurring_payment_schedule(new_req.request_id, amount_clean if amount else amount, payment_schedule_data)
                        if success:
                            print(f"🔧 DEBUG: Successfully created custom payment schedule for request #{new_req.request_id}")
                            log_action(f"Created custom payment schedule for request #{new_req.request_id}")
                        else:
                            print(f"🔧 DEBUG: Failed to create custom payment schedule for request #{new_req.request_id}")
                    else:
                        print(f"🔧 DEBUG: No payment schedule data to create")
                        
                except Exception as e:
                    print(f"Error creating custom payment schedule: {e}")
                    import traceback
                    traceback.print_exc()
                    flash('Payment request created but custom schedule configuration failed. Please contact admin.', 'warning')
            
            # Handle variable amount recurring payments (monthly)
            elif request.form.get('variable_amounts') == 'true':
                try:
                    # Get the payment schedule data from the form
                    payment_schedule_data = []
                    schedule_data = request.form.get('payment_schedule', '[]')
                    import json
                    schedule_list = json.loads(schedule_data)
                    
                    for payment in schedule_list:
                        payment_schedule_data.append({
                            'date': payment['date'],
                            'amount': payment['amount']
                        })
                    
                    # Create the payment schedule
                    if payment_schedule_data:
                        create_recurring_payment_schedule(new_req.request_id, amount_clean if amount else amount, payment_schedule_data)
                        log_action(f"Created variable amount payment schedule for request #{new_req.request_id}")
                except Exception as e:
                    print(f"Error creating payment schedule: {e}")
                    flash('Payment request created but schedule configuration failed. Please contact admin.', 'warning')
        
        log_action(f"Created payment request #{new_req.request_id} - {request_type}")
        
        # Create notifications based on request status and RBAC rules
        try:
            if new_req.status == 'Completed':
                # Finance department requests are auto-approved - notify Finance Admin
                notify_users_by_role(
                    request=new_req,
                    notification_type="ready_for_finance_review",
                    title="New Payment Request Submitted",
                    message=f"New {request_type} request submitted by {requestor_name} from {current_user.department} department for OMR {amount}",
                    request_id=new_req.request_id
                )
            else:
                # Other departments - notify based on RBAC rules
                notify_users_by_role(
                    request=new_req,
                    notification_type="new_submission",
                    title="New Payment Request for Approval",
                    message=f"New {request_type} request submitted by {requestor_name} from {current_user.department} department for OMR {amount} - requires your approval",
                    request_id=new_req.request_id
                )
                
                
        except Exception as e:
            print(f"Error creating notifications: {e}")
            # Don't fail the request creation if notification fails
        
        # Emit real-time event to all users
        try:
            socketio.emit('new_request', {
                'request_id': new_req.request_id,
                'request_type': new_req.request_type,
                'requestor_name': new_req.requestor_name,
                'department': new_req.department,
                'amount': float(new_req.amount),
                'status': new_req.status,
                'date': new_req.date.strftime('%Y-%m-%d')
            }, room='all_users')
            
            # Also emit specifically to finance admin room for immediate updates
            socketio.emit('new_request', {
                'request_id': new_req.request_id,
                'request_type': new_req.request_type,
                'requestor_name': new_req.requestor_name,
                'department': new_req.department,
                'amount': float(new_req.amount),
                'status': new_req.status,
                'date': new_req.date.strftime('%Y-%m-%d')
            }, room='finance_admin')
            
            # Also emit to department staff room for department dashboard users
            socketio.emit('new_request', {
                'request_id': new_req.request_id,
                'request_type': new_req.request_type,
                'requestor_name': new_req.requestor_name,
                'department': new_req.department,
                'amount': float(new_req.amount),
                'status': new_req.status,
                'date': new_req.date.strftime('%Y-%m-%d')
            }, room='department_staff')
            
            # Also emit to department managers room
            socketio.emit('new_request', {
                'request_id': new_req.request_id,
                'request_type': new_req.request_type,
                'requestor_name': new_req.requestor_name,
                'department': new_req.department,
                'amount': float(new_req.amount),
                'status': new_req.status,
                'date': new_req.date.strftime('%Y-%m-%d')
            }, room='department_managers')
            
            print(f"DEBUG: Emitted new_request event to all_users, finance_admin, department_staff, and department_managers rooms")
        except Exception as e:
            print(f"Error emitting real-time notification: {e}")
            # Don't fail the request creation if real-time notification fails
        
        flash('Payment request submitted successfully!', 'success')
        return redirect(url_for('dashboard'))
    
    # Get available request types and branches for the New Request page
    available_request_types = get_available_request_types()
    # Custom order: Office, Kucu, Boom, Thoum, Kitchen
    from sqlalchemy import case
    location_order = case(
        (Branch.restaurant == 'Office', 1),
        (Branch.restaurant == 'Kucu', 2),
        (Branch.restaurant == 'Boom', 3),
        (Branch.restaurant == 'Thoum', 4),
        (Branch.restaurant == 'Kitchen', 5),
        else_=6
    )
    available_branches = Branch.query.filter_by(is_active=True).order_by(location_order, Branch.name).all()
    today = datetime.utcnow().date().strftime('%Y-%m-%d')
    return render_template('new_request.html', user=current_user, today=today, available_request_types=available_request_types, available_branches=available_branches)


@app.route('/populate-request-types')
@login_required
@role_required('IT Staff', 'Department Manager')
def populate_request_types_route():
    """Populate request types in the database"""
    try:
        from populate_request_types import populate_request_types, verify_population
        
        # Populate request types
        populate_request_types()
        
        # Verify population
        verify_population()
        
        flash('Request types populated successfully!', 'success')
        return redirect(url_for('it_dashboard'))
        
    except Exception as e:
        flash(f'Error populating request types: {str(e)}', 'error')
        return redirect(url_for('it_dashboard'))

@app.route('/test-timezone')
def test_timezone():
    """Test timezone conversion"""
    from datetime import datetime
    test_time = datetime(2025, 10, 20, 6, 36, 21)
    local_time = utc_to_local(test_time)
    return f"UTC: {test_time}, Local: {local_time}"

@app.route('/fix-durations')
@login_required
@role_required('Admin')
def fix_durations():
    """Fix existing incorrect duration calculations in the database"""
    try:
        # Get all requests with timing data
        requests = PaymentRequest.query.filter(
            PaymentRequest.manager_approval_start_time.isnot(None),
            PaymentRequest.manager_approval_end_time.isnot(None)
        ).all()
        
        fixed_count = 0
        
        for req in requests:
            # Recalculate manager approval duration
            if req.manager_approval_start_time and req.manager_approval_end_time:
                duration = req.manager_approval_end_time - req.manager_approval_start_time
                correct_duration_seconds = int(duration.total_seconds())
                
                # If the stored duration is much larger than the actual duration,
                # it was likely calculated in minutes instead of seconds
                if req.manager_approval_duration_minutes and req.manager_approval_duration_minutes > correct_duration_seconds:
                    req.manager_approval_duration_minutes = correct_duration_seconds
                    fixed_count += 1
            
            # Recalculate finance approval duration
            if req.finance_approval_start_time and req.finance_approval_end_time:
                duration = req.finance_approval_end_time - req.finance_approval_start_time
                correct_duration_seconds = int(duration.total_seconds())
                
                # If the stored duration is much larger than the actual duration,
                # it was likely calculated in minutes instead of seconds
                if req.finance_approval_duration_minutes and req.finance_approval_duration_minutes > correct_duration_seconds:
                    req.finance_approval_duration_minutes = correct_duration_seconds
                    fixed_count += 1
        
        db.session.commit()
        return f"Fixed {fixed_count} duration calculations"
        
    except Exception as e:
        return f"Error fixing durations: {str(e)}"

@app.route('/api/timing/<int:request_id>')
@login_required
def get_timing_api(request_id):
    """API endpoint to get real-time timing data for a request"""
    try:
        req = PaymentRequest.query.get_or_404(request_id)
        current_time = datetime.utcnow()
        
        timing_data = {}
        
        # Manager approval timing
        if req.manager_approval_start_time:
            if req.manager_approval_end_time:
                # Completed - show final duration
                duration = req.manager_approval_end_time - req.manager_approval_start_time
                timing_data['manager'] = {
                    'status': 'completed',
                    'start_time': req.manager_approval_start_time.isoformat(),
                    'end_time': req.manager_approval_end_time.isoformat(),
                    'duration_seconds': int(duration.total_seconds()),
                    'duration_minutes': req.manager_approval_duration_minutes
                }
            else:
                # In progress - calculate elapsed time
                elapsed = current_time - req.manager_approval_start_time
                timing_data['manager'] = {
                    'status': 'in_progress',
                    'start_time': req.manager_approval_start_time.isoformat(),
                    'elapsed_seconds': int(elapsed.total_seconds())
                }
        
        # Finance approval timing
        if req.finance_approval_start_time:
            if req.finance_approval_end_time:
                # Completed - show final duration
                duration = req.finance_approval_end_time - req.finance_approval_start_time
                timing_data['finance'] = {
                    'status': 'completed',
                    'start_time': req.finance_approval_start_time.isoformat(),
                    'end_time': req.finance_approval_end_time.isoformat(),
                    'duration_seconds': int(duration.total_seconds()),
                    'duration_minutes': req.finance_approval_duration_minutes
                }
            else:
                # In progress - calculate elapsed time
                elapsed = current_time - req.finance_approval_start_time
                timing_data['finance'] = {
                    'status': 'in_progress',
                    'start_time': req.finance_approval_start_time.isoformat(),
                    'elapsed_seconds': int(elapsed.total_seconds())
                }
        
        return jsonify(timing_data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/payment-schedule/<int:request_id>')
@login_required
def get_payment_schedule_api(request_id):
    """API endpoint to get payment schedule for a request"""
    try:
        schedule = get_payment_schedule(request_id)
        return jsonify({'success': True, 'schedule': schedule})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/admin/check-timing-alerts')
@login_required
@role_required('Finance Admin')
def check_timing_alerts():
    """Manual endpoint to check and send timing alerts (for testing and manual triggers)"""
    try:
        check_finance_approval_timing_alerts()
        flash('Timing alerts check completed successfully.', 'success')
        return redirect(url_for('admin_dashboard'))
    except Exception as e:
        flash(f'Error checking timing alerts: {str(e)}', 'error')
        return redirect(url_for('admin_dashboard')), 500


@app.route('/admin/overdue-requests')
@login_required
@role_required('Finance Admin', 'Finance Staff')
def overdue_requests():
    """Display all overdue finance approval requests"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)  # Default to 50 to show more requests
        
        # Validate per_page to prevent abuse - allow higher limits to show all overdue requests
        if per_page not in [10, 20, 50, 100, 200, 500]:
            per_page = 50
        
        # Get ALL overdue requests (no limit in the query)
        overdue_requests_list = get_overdue_requests()
        overdue_count = len(overdue_requests_list)
        
        # Calculate pagination
        total_pages = (overdue_count + per_page - 1) // per_page if overdue_count > 0 else 1  # Ceiling division
        start_index = (page - 1) * per_page
        end_index = start_index + per_page
        
        # Get paginated results
        paginated_overdue_requests = overdue_requests_list[start_index:end_index]
        
        return render_template('overdue_requests.html', 
                             overdue_requests=paginated_overdue_requests,
                             overdue_count=overdue_count,
                             page=page,
                             per_page=per_page,
                             total_pages=total_pages)
    except Exception as e:
        flash(f'Error loading overdue requests: {str(e)}', 'error')
        import traceback
        traceback.print_exc()
        return redirect(url_for('admin_dashboard'))


@app.route('/favicon.ico')
def favicon():
    """Serve favicon to prevent 404 errors"""
    return '', 204  # No content response


@app.route('/write-cheque', methods=['GET', 'POST'])
@login_required
@role_required('GM', 'CEO', 'Operation Manager')
def write_cheque():
    """Write a cheque for approved payment requests"""
    if request.method == 'POST':
        # Handle cheque writing logic here
        flash('Cheque written successfully!', 'success')
        return redirect(url_for('dashboard'))
    
    # Get approved payment requests that need cheques
    approved_requests = PaymentRequest.query.filter_by(status='Approved').all()
    
    return render_template('write_cheque.html', 
                         user=current_user, 
                         approved_requests=approved_requests)


@app.route('/generate-cheque-pdf', methods=['POST'])
@login_required
@role_required('GM', 'CEO', 'Operation Manager')
def generate_cheque_pdf():
    """Generate PDF from cheque data"""
    try:
        data = request.get_json()
        
        # Extract cheque data
        cheque_date = data.get('chequeDate', '')
        payee_name = data.get('payeeName', '')
        amount = data.get('amount', '')
        currency = data.get('currency', 'OMR')
        crossing = data.get('crossing', '')
        bank = data.get('bank', 'dhofar_islamic')
        
        # Format date
        if cheque_date:
            try:
                date_obj = datetime.strptime(cheque_date, '%Y-%m-%d')
                formatted_date = date_obj.strftime('%d/%m/%Y')
            except:
                formatted_date = cheque_date
        else:
            formatted_date = ''
        
        # Format amount
        formatted_amount = ''
        if amount:
            try:
                numeric = float(amount)
                formatted_amount = f"{numeric:,.3f}".replace(',', '')
                if currency and currency != 'OMR':
                    formatted_amount = f"{formatted_amount} {currency}"
            except:
                formatted_amount = amount
        
        # Convert amount to words (simplified - you may want to use the same function from JS)
        amount_words = ''
        if amount:
            try:
                numeric = float(amount)
                # Simple conversion (you can enhance this)
                amount_words = convert_amount_to_words(numeric)
            except:
                pass
        
        # Get bank-specific positions
        bank_positions = {
            'dhofar_islamic': {
                'date': {'top': '90px', 'left': '220px'},  # Moved slightly higher (92->90)
                'payee': {'top': '132px', 'left': '40px'},  # Moved 5px lower (127->132)
                'amount': {'top': '165px', 'left': '367px'},  # Moved 3px left (370->367)
                'amountWords': {'top': '155px', 'left': '41px'},  # Moved 3px right (38->41)
                'crossing': {'top': '20px', 'left': '60px'}
            },
            'oman_arab': {
                'date': {'top': '62px', 'left': '365px', 'extra': 'font-size: 11px;'},  # +20px right (345->365)
                'payee': {'top': '130px', 'left': '-5px', 'extra': 'font-size: 13px;'},   # Nudge ~1px down (129->130)
                'amount': {'top': '135px', 'left': '383px', 'extra': 'font-size: 12px;'}, # 8px down (127->135), 3px right (380->383), smaller font
                'amountWords': {'top': '157px', 'left': '58px', 'extra': 'max-width: 230px; width: 230px; line-height: 2.5;'},  # Bigger line-height
                'crossing': {'top': '20px', 'left': '60px'}  # Same as Dhofar Islamic Bank for testing
            },
            'sohar': {
                'date': {'top': '52px', 'left': '555px'},
                'payee': {'top': '116px', 'left': '115px'},
                'amount': {'top': '163px', 'left': '568px'},
                'amountWords': {'top': '153px', 'left': '80px'},
                'crossing': {'top': '25px', 'left': '70px'}
            }
        }
        
        positions = bank_positions.get(bank, bank_positions['dhofar_islamic'])
        
        # Get bank image
        bank_images = {
            'dhofar_islamic': 'DHOFAR-ISLAMIC-BANK.png',
            'oman_arab': 'OMAN-ARAB-BANK.png',
            'sohar': 'SOHAR-BANK.png'
        }
        bank_image = bank_images.get(bank, 'DHOFAR-ISLAMIC-BANK.png')
        
        # Get full path to bank image
        bank_image_path = os.path.join(app.static_folder, 'cheque_templates', bank_image)
        bank_image_url = url_for('static', filename=f'cheque_templates/{bank_image}', _external=True)
        
        # Convert image to base64 for xhtml2pdf compatibility
        bank_image_base64 = ''
        if os.path.exists(bank_image_path):
            with open(bank_image_path, 'rb') as img_file:
                img_data = img_file.read()
                bank_image_base64 = base64.b64encode(img_data).decode('utf-8')
                bank_image_base64 = f"data:image/png;base64,{bank_image_base64}"
        
        # Render PDF template
        html_content = render_template('cheque_pdf.html',
                                     cheque_date=formatted_date,
                                     payee_name=payee_name,
                                     amount=formatted_amount,
                                     amount_words=amount_words,
                                     crossing=crossing,
                                     bank_image_base64=bank_image_base64,
                                     positions=positions)
        
        # Generate PDF using Playwright
        pdf_buffer = BytesIO()
        
        with sync_playwright() as p:
            # Launch browser
            browser = p.chromium.launch(headless=True)
            page = browser.new_page()
            
            # Set page size to cheque dimensions (7.5in x 3.5in)
            page.set_viewport_size({"width": 720, "height": 336})  # 7.5in * 96dpi = 720px, 3.5in * 96dpi = 336px
            
            # Load HTML content
            page.set_content(html_content, wait_until='networkidle')
            
            # Generate PDF with exact page size
            # Using scale: 1.0 to ensure actual size (100% scale)
            pdf_bytes = page.pdf(
                format=None,
                width='7.5in',
                height='3.5in',
                margin={'top': '0', 'right': '0', 'bottom': '0', 'left': '0'},
                print_background=True,
                scale=1.0  # Force 100% scale (actual size)
            )
            
            browser.close()
            
            # Write PDF to buffer
            pdf_buffer.write(pdf_bytes)
            pdf_buffer.seek(0)
        
        # Return PDF
        return Response(
            pdf_buffer.getvalue(),
            mimetype='application/pdf',
            headers={
                'Content-Disposition': 'inline; filename=cheque.pdf'
            }
        )
        
    except Exception as e:
        print(f"Error generating PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def convert_amount_to_words(amount):
    """Convert numeric amount to words"""
    ones = ['zero', 'one', 'two', 'three', 'four', 'five', 'six', 'seven', 'eight', 'nine', 
            'ten', 'eleven', 'twelve', 'thirteen', 'fourteen', 'fifteen', 'sixteen', 
            'seventeen', 'eighteen', 'nineteen']
    tens = ['', '', 'twenty', 'thirty', 'forty', 'fifty', 'sixty', 'seventy', 'eighty', 'ninety']
    
    def number_to_words(n):
        if n < 20:
            return ones[int(n)]
        if n < 100:
            t = int(n // 10)
            r = int(n % 10)
            return f"{tens[t]}-{ones[r]}" if r else tens[t]
        if n < 1000:
            h = int(n // 100)
            r = int(n % 100)
            return f"{ones[h]} hundred {number_to_words(r)}" if r else f"{ones[h]} hundred"
        
        units = [
            (1_000_000_000_000, 'trillion'),
            (1_000_000_000, 'billion'),
            (1_000_000, 'million'),
            (1_000, 'thousand')
        ]
        
        for unit_value, unit_name in units:
            if n >= unit_value:
                q = int(n // unit_value)
                r = int(n % unit_value)
                return f"{number_to_words(q)} {unit_name} {number_to_words(r)}" if r else f"{number_to_words(q)} {unit_name}"
        return ''
    
    try:
        fixed = f"{amount:.3f}"
        int_str, dec_str = fixed.split('.')
        int_num = int(int_str)
        
        words = number_to_words(abs(int_num))
        if int_num == 0:
            words = 'zero'
        
        if dec_str and int(dec_str) > 0:
            dec_words = ' '.join([number_to_words(int(d)) for d in dec_str])
            words = f"{words} point {dec_words}"
        
        if amount < 0:
            words = f"minus {words}"
        
        # Capitalize and add "Omani Rials Only"
        capitalized = ' '.join([word.capitalize() for word in words.split()])
        return f"{capitalized} Omani Rials Only"
    except:
        return ''


@app.route('/request/<int:request_id>')
@login_required
def view_request(request_id):
    """View a specific payment request"""
    # Use eager loading to reduce database queries
    req = PaymentRequest.query.options(
        db.joinedload(PaymentRequest.user)
    ).get_or_404(request_id)
    
    # If request is archived, only IT department can view it (they can access via archives page)
    if req.is_archived:
        it_allowed = (current_user.role == 'IT Staff') or (current_user.role == 'Department Manager' and current_user.department == 'IT')
        if not it_allowed:
            flash('This request has been archived and is only accessible from the Archives page (IT Department only).', 'warning')
            return redirect(url_for('dashboard'))
    
    # Hard restriction: CEO requests are viewable by Finance Admins, GM, Operation Manager, IT roles, or the CEO themself
    if getattr(req.user, 'role', None) == 'CEO':
        it_allowed = (current_user.role == 'IT Staff') or (current_user.role == 'Department Manager' and current_user.department == 'IT')
        if current_user.user_id != req.user_id and current_user.role not in ['Finance Admin', 'GM', 'Operation Manager'] and not it_allowed:
            return render_template('403.html'), 403

    # Check permissions
    # Allow Operation Manager, IT users, and IT Department Managers to view all requests (same as GM visibility)
    if current_user.role not in ['Finance Admin', 'Finance Staff', 'GM', 'CEO', 'IT Staff', 'Project Staff', 'Operation Manager']:
        # Auditing Department users (Staff and Department Manager) can view their own requests OR Completed/Recurring requests from other departments
        if current_user.department == 'Auditing' and (current_user.role == 'Auditing Staff' or current_user.role == 'Department Manager'):
            # Allow if:
            # 1. It's their own request, OR
            # 2. (For Department Manager) It's from Auditing department, OR
            # 3. It's Completed/Recurring from another department, OR
            # 4. They are temporary manager
            if req.user_id != current_user.user_id:
                # Not their own request
                if current_user.role == 'Department Manager':
                    # Department Manager can also view their department's requests
                    if req.department == 'Auditing' or getattr(req, 'temporary_manager_id', None) == current_user.user_id:
                        pass  # Allow access
                    elif req.status not in ['Completed', 'Recurring']:
                        flash('You do not have permission to view this request.', 'danger')
                        return redirect(url_for('dashboard'))
                else:
                    # Auditing Staff - check if it's Completed/Recurring from another department
                    if req.department == 'Auditing' or req.status not in ['Completed', 'Recurring']:
                        flash('You do not have permission to view this request.', 'danger')
                        return redirect(url_for('dashboard'))
        # Other Department Managers can view requests from their department
        elif current_user.role == 'Department Manager':
            # IT Department Managers can view all requests
            if current_user.department == 'IT':
                pass  # Allow access to all requests
            # Temporary manager may view regardless of department
            elif getattr(req, 'temporary_manager_id', None) == current_user.user_id:
                pass
            elif req.department != current_user.department:
                flash('You do not have permission to view this request.', 'danger')
                return redirect(url_for('dashboard'))
        # Regular users can only view their own requests
        elif req.user_id != current_user.user_id:
            flash('You do not have permission to view this request.', 'danger')
            return redirect(url_for('dashboard'))
    elif current_user.role in ['Finance Admin', 'Finance Staff']:
        # If assigned as temporary manager, allow viewing regardless of finance status rules
        if getattr(req, 'temporary_manager_id', None) == current_user.user_id:
            pass
        else:
            # Always allow finance users to view their own requests regardless of status
            if req.user_id == current_user.user_id:
                pass
            else:
                # Finance users can only view requests in finance-related statuses
                finance_statuses = ['Pending Finance Approval', 'Proof Pending', 'Proof Sent', 'Proof Rejected', 'Recurring', 'Completed', 'Rejected by Finance']
                
                # For Abdalaziz, also allow viewing PMA/Rejected-by-Manager for Finance dept, his own, and GM/CEO/Operation Manager submissions
                if current_user.name == 'Abdalaziz Al-Brashdi' and req.status in ['Pending Manager Approval', 'Rejected by Manager']:
                    if (req.department == 'Finance' or req.user_id == current_user.user_id or getattr(req.user, 'role', None) in ['GM', 'CEO', 'Operation Manager']):
                        pass  # Allow access
                    else:
                        flash('You do not have permission to view this request.', 'danger')
                        return redirect(url_for('dashboard'))
                elif req.status not in finance_statuses:
                    flash('You do not have permission to view this request.', 'danger')
                    return redirect(url_for('dashboard'))
    
    # Mark notifications related to this request as read for Finance and Admin users
    if current_user.role in ['Finance Staff', 'Finance Admin']:
        # Use a more efficient update query
        db.session.query(Notification).filter_by(
            user_id=current_user.user_id,
            request_id=request_id,
            is_read=False
        ).update({'is_read': True})
        db.session.commit()
    
    # Determine if current user can schedule a one-time payment date
    can_schedule_one_time = False
    one_time_scheduled_by = None

    try:
        # Authorization for scheduling one-time payment date
        # 1) If a temporary manager is assigned, only they can schedule
        if getattr(req, 'temporary_manager_id', None):
            can_schedule_one_time = (req.temporary_manager_id == current_user.user_id)
        else:
            # 2) Global roles
            if current_user.role in ['GM', 'Operation Manager']:
                can_schedule_one_time = True
            # 3) IT department (IT Staff and IT Department Manager)
            elif current_user.department == 'IT' and current_user.role in ['IT Staff', 'Department Manager']:
                can_schedule_one_time = True
            # 4) Assigned manager of the requestor
            elif getattr(req.user, 'manager_id', None) == current_user.user_id:
                can_schedule_one_time = True
            # 5) Department Manager of the same department as the requestor
            elif current_user.role == 'Department Manager' and current_user.department == req.department:
                can_schedule_one_time = True

        # Resolve the name of the scheduler from audit logs (latest)
        if req.payment_date:
            from models import AuditLog
            keyword = f"request #{request_id}"
            log = (
                AuditLog.query
                .filter(AuditLog.action.like('%Scheduled one-time payment date%'))
                .filter(AuditLog.action.like(f"%{keyword}%"))
                .order_by(AuditLog.timestamp.desc())
                .first()
            )
            if log:
                one_time_scheduled_by = getattr(log.user, 'name', None) or log.username_snapshot or 'Unknown'
    except Exception as e:
        # Fail-safe: do not break view if any of the above fails
        print(f"DEBUG: Error computing can_schedule_one_time or scheduled_by: {e}")

    # Get schedule rows for variable payments - show for Admin review, but only allow payments when approved
    schedule_rows = []
    total_paid_amount = 0
    
    # Process schedule if it's a recurring payment (monthly or custom)
    print(f"🔧 DEBUG: view_request - recurring_interval: {req.recurring_interval}")
    if req.recurring_interval and ('monthly' in req.recurring_interval or req.recurring_interval.startswith('custom:')):
        print(f"🔧 DEBUG: Processing recurring payment schedule")
        # Get variable payment schedule if exists - use single query with ordering
        schedule = RecurringPaymentSchedule.query.filter_by(
            request_id=request_id
        ).order_by(RecurringPaymentSchedule.payment_order).all()
        print(f"🔧 DEBUG: Found {len(schedule)} schedule entries")
        
        if schedule:
            # Optimize: Get all paid notifications and late installments in single queries
            # Use list comprehension for better performance
            paid_notifications = PaidNotification.query.filter_by(request_id=request_id).all()
            late_installments = LateInstallment.query.filter_by(request_id=request_id).all()
            
            # Create sets for O(1) lookup instead of O(n) list search
            paid_dates = {paid.paid_date for paid in paid_notifications}
            late_dates = {late.payment_date for late in late_installments}
            
            # Calculate total paid and remaining amounts
            total_paid_amount = 0
            
            # Use list comprehension for better performance
            for entry in schedule:
                # Check if this installment is already paid (use the is_paid field from the schedule)
                is_paid = entry.is_paid
                # Check if this installment is marked late (optimized lookup)
                is_late = entry.payment_date in late_dates
                
                # If this installment is paid, add its amount to total paid
                if is_paid:
                    total_paid_amount += entry.amount
                
                schedule_rows.append({
                    'schedule_id': entry.schedule_id,
                    'date': entry.payment_date,
                    'payment_date': entry.payment_date,  # Add both for compatibility
                    'amount': entry.amount,
                    'is_paid': is_paid,
                    'is_late': is_late,
                    'receipt_path': entry.receipt_path,
                    'invoice_path': entry.invoice_path,
                    'has_been_edited': entry.has_been_edited
                })
    
    # Determine the manager's name for display
    manager_name = None
    temporary_manager_name = None
    
    # Check if there's a temporary manager assigned (IT Department feature)
    if req.temporary_manager_id:
        temp_manager = User.query.get(req.temporary_manager_id)
        if temp_manager:
            temporary_manager_name = temp_manager.name
    
    # Determine manager name for all statuses (pending and completed)
    if req.user.manager_id:
        # Get the manager's name from the manager_id
        manager = User.query.get(req.user.manager_id)
        if manager:
            manager_name = manager.name
        else:
            # If manager_id exists but user not found, try to find Department Manager
            dept_manager = User.query.filter_by(role='Department Manager', department=req.department).first()
            if dept_manager:
                manager_name = dept_manager.name
    else:
        # If no manager_id is set, find the Department Manager for the requestor's department
        dept_manager = User.query.filter_by(role='Department Manager', department=req.department).first()
        if dept_manager:
            manager_name = dept_manager.name
        elif req.department in ['Operation', 'Project']:
            # For Operation and Project, try Operation Manager as fallback
            operation_manager = User.query.filter_by(role='Operation Manager').first()
            if operation_manager:
                manager_name = operation_manager.name
        elif req.department == 'Office':
            # For Office, fallback to the General Manager
            gm_user_fallback = User.query.filter_by(role='GM').first()
            if gm_user_fallback:
                manager_name = gm_user_fallback.name
        
    # Also resolve GM and Operation Manager names (used for Department Manager submissions)
    gm_user = User.query.filter_by(role='GM').first()
    gm_name = gm_user.name if gm_user else 'General Manager'
    op_manager_user = User.query.filter_by(role='Operation Manager').first()
    op_manager_name = op_manager_user.name if op_manager_user else 'Operation Manager'
    
    # Get all proof files for this request grouped by batch
    proof_files = []
    proof_batches = []
    if req.status in ['Proof Sent', 'Proof Rejected', 'Completed', 'Recurring']:
        import os
        import glob
        upload_folder = app.config['UPLOAD_FOLDER']
        # Look for all proof files for this request (files starting with proof_{request_id}_)
        proof_pattern = os.path.join(upload_folder, f"proof_{request_id}_*")
        all_files = [os.path.basename(f) for f in glob.glob(proof_pattern)]
        # Parse batch numbers from filenames
        from collections import defaultdict
        batches = defaultdict(list)
        for fname in all_files:
            batch_num = 1
            try:
                prefix = f"proof_{request_id}_"
                if fname.startswith(prefix + "b"):
                    after = fname[len(prefix)+1:]  # skip 'b'
                    part = after.split('_', 1)[0]
                    batch_num = int(part)
            except Exception:
                batch_num = 1
            batches[batch_num].append(fname)

        # Sort files within each batch by filename (timestamp included)
        for bn in batches:
            batches[bn].sort(reverse=True)

        # Build ordered list of batches, latest first
        ordered_batch_nums = sorted(batches.keys(), reverse=True)
        for bn in ordered_batch_nums:
            proof_batches.append({
                'batch_num': bn,
                'files': batches[bn]
            })

        # Also keep a flat list for legacy sections if any
        proof_files = [f for bn in ordered_batch_nums for f in batches[bn]]
    
    # Get current server time for timer calculations
    current_server_time = datetime.utcnow()
    
    # Get finance admin notes for this request (ordered by creation date, newest first)
    finance_notes = FinanceAdminNote.query.filter_by(request_id=request_id).order_by(FinanceAdminNote.created_at.desc()).all()
    
    # Ensure finance approval duration is calculated if needed
    calculate_finance_approval_duration(req)
    if req.finance_approval_duration_minutes is not None:
        db.session.commit()
    
    # Get list of managers for IT dropdown (only if user is IT Staff or IT Department Manager)
    available_managers = []
    if current_user.department == 'IT' and current_user.role in ['IT Staff', 'Department Manager']:
        available_managers = User.query.filter(
            User.role.in_(['Department Manager', 'GM', 'Operation Manager', 'Finance Admin'])
        ).order_by(User.department, User.name).all()
    
    # Prepare receipt files for template (both requestor and finance admin receipts)
    requestor_receipts = []
    finance_admin_receipts = []
    import json
    
    # Get requestor receipts
    if req.requestor_receipt_path:
        try:
            requestor_receipts = json.loads(req.requestor_receipt_path)
        except (json.JSONDecodeError, TypeError, ValueError):
            # Fallback: try Python literal (in case it's a repr list)
            try:
                import ast
                parsed = ast.literal_eval(req.requestor_receipt_path)
                if isinstance(parsed, list):
                    requestor_receipts = parsed
                elif isinstance(parsed, str) and parsed:
                    requestor_receipts = [parsed]
            except Exception:
                # Final fallback: split on commas
                if isinstance(req.requestor_receipt_path, str) and req.requestor_receipt_path:
                    if ',' in req.requestor_receipt_path:
                        requestor_receipts = [p.strip() for p in req.requestor_receipt_path.split(',') if p.strip()]
                    else:
                        requestor_receipts = [req.requestor_receipt_path]
    
    # Get finance admin receipts
    if req.finance_admin_receipt_path:
        try:
            finance_admin_receipts = json.loads(req.finance_admin_receipt_path)
        except (json.JSONDecodeError, TypeError, ValueError):
            # Fallback: try Python literal (in case it's a repr list)
            try:
                import ast
                parsed = ast.literal_eval(req.finance_admin_receipt_path)
                if isinstance(parsed, list):
                    finance_admin_receipts = parsed
                elif isinstance(parsed, str) and parsed:
                    finance_admin_receipts = [parsed]
            except Exception:
                # Final fallback: split on commas
                if isinstance(req.finance_admin_receipt_path, str) and req.finance_admin_receipt_path:
                    if ',' in req.finance_admin_receipt_path:
                        finance_admin_receipts = [p.strip() for p in req.finance_admin_receipt_path.split(',') if p.strip()]
                    else:
                        finance_admin_receipts = [req.finance_admin_receipt_path]
    
    # Backward compatibility: If new columns are empty but legacy receipt_path exists,
    # try to determine which column it should be in based on request status
    if not requestor_receipts and not finance_admin_receipts and req.receipt_path:
        try:
            legacy_receipts = json.loads(req.receipt_path)
            if isinstance(legacy_receipts, list):
                # Check status to determine if it's from requestor or finance admin
                finance_statuses = ['Proof Pending', 'Proof Sent', 'Proof Rejected', 'Recurring', 'Completed']
                if req.approver and req.status in finance_statuses:
                    finance_admin_receipts = legacy_receipts
                else:
                    requestor_receipts = legacy_receipts
        except (json.JSONDecodeError, TypeError):
            # Handle legacy single file format
            if isinstance(req.receipt_path, str):
                finance_statuses = ['Proof Pending', 'Proof Sent', 'Proof Rejected', 'Recurring', 'Completed']
                if req.approver and req.status in finance_statuses:
                    finance_admin_receipts = [req.receipt_path]
                else:
                    requestor_receipts = [req.receipt_path]
    
    # Provide the same option lists used by the New Request form so edit UI can mirror creation behavior
    from models import RequestType, Branch
    # Map Office → Management for request type catalog
    effective_department = 'Management' if req.department == 'Office' else req.department
    available_request_types = RequestType.query.filter_by(department=effective_department, is_active=True).order_by(RequestType.name).all()
    # Custom order: Office, Kucu, Boom, Thoum, Kitchen
    from sqlalchemy import case
    location_order = case(
        (Branch.restaurant == 'Office', 1),
        (Branch.restaurant == 'Kucu', 2),
        (Branch.restaurant == 'Boom', 3),
        (Branch.restaurant == 'Thoum', 4),
        (Branch.restaurant == 'Kitchen', 5),
        else_=6
    )
    available_branches = Branch.query.filter_by(is_active=True).order_by(location_order, Branch.name).all()

    was_just_edited = request.args.get('edited') == '1'
    edited_fields_param = request.args.get('edited_fields', '')
    edited_fields = [f for f in edited_fields_param.split(',') if f]
    # Fetch all-time edited fields from audit table (created on demand)
    try:
        result = db.session.execute(db.text('''
            CREATE TABLE IF NOT EXISTS request_field_edits (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_id INTEGER NOT NULL,
                field_name TEXT NOT NULL,
                first_edited_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                last_edited_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(request_id, field_name)
            );
        '''))
        result.close()
        rows = db.session.execute(db.text('SELECT field_name FROM request_field_edits WHERE request_id = :rid'), { 'rid': req.request_id }).fetchall()
        cumulative_fields = [r[0] for r in rows]
    except Exception:
        cumulative_fields = []
    # Union of current-save markers and cumulative markers
    edited_fields_all = sorted(set(edited_fields) | set(cumulative_fields))
    return render_template('view_request.html', request=req, user=current_user, schedule_rows=schedule_rows, total_paid_amount=float(total_paid_amount), manager_name=manager_name, temporary_manager_name=temporary_manager_name, available_managers=available_managers, proof_files=proof_files, proof_batches=proof_batches, current_server_time=current_server_time, finance_notes=finance_notes, gm_name=gm_name, op_manager_name=op_manager_name, requestor_receipts=requestor_receipts, finance_admin_receipts=finance_admin_receipts, available_request_types=available_request_types, available_branches=available_branches, was_just_edited=was_just_edited, edited_fields=edited_fields_all, can_schedule_one_time=can_schedule_one_time, one_time_scheduled_by=one_time_scheduled_by)


@app.route('/request/<int:request_id>/schedule_one_time_payment', methods=['POST'])
@login_required
def schedule_one_time_payment(request_id):
    """Allow authorized users to set a payment date for ONE-TIME requests (optional)."""
    req = PaymentRequest.query.get_or_404(request_id)

    # Only for One-Time requests
    if req.recurring == 'Recurring':
        flash('You can only schedule a payment date for one-time requests.', 'error')
        return redirect(url_for('view_request', request_id=request_id))

    # Only allowed when request is in these statuses
    if req.status not in ['Pending Manager Approval', 'Pending Finance Approval']:
        flash('Payment date can only be scheduled when status is Pending Manager Approval or Pending Finance Approval.', 'error')
        return redirect(url_for('view_request', request_id=request_id))

    # Authorization logic mirrors the view logic
    authorized = False
    if getattr(req, 'temporary_manager_id', None):
        authorized = (req.temporary_manager_id == current_user.user_id)
    else:
        if current_user.role in ['GM', 'Operation Manager']:
            authorized = True
        elif current_user.department == 'IT' and current_user.role in ['IT Staff', 'Department Manager']:
            authorized = True
        elif getattr(req.user, 'manager_id', None) == current_user.user_id:
            authorized = True
        elif current_user.role == 'Department Manager' and current_user.department == req.department:
            authorized = True

    if not authorized:
        flash('You are not authorized to schedule a payment date for this request.', 'error')
        return redirect(url_for('view_request', request_id=request_id))

    payment_date_str = request.form.get('payment_date', '').strip()
    if not payment_date_str:
        flash('Please select a payment date.', 'error')
        return redirect(url_for('view_request', request_id=request_id))

    try:
        from datetime import datetime as _dt
        payment_date_val = _dt.strptime(payment_date_str, '%Y-%m-%d').date()
    except Exception:
        flash('Invalid payment date format.', 'error')
        return redirect(url_for('view_request', request_id=request_id))

    # Persist - automatically change payment type to "Scheduled One-Time" when payment date is set
    req.payment_date = payment_date_val
    # Automatically update payment type to "Scheduled One-Time" if it's currently "One-Time" or None
    if req.recurring != 'Recurring':
        req.recurring = 'Scheduled One-Time'
    req.updated_at = datetime.utcnow()
    db.session.commit()

    # Audit and realtime
    log_action(
        f"Request #{request_id} from {getattr(req.user, 'name', 'Unknown')} in the {req.department} department has been scheduled for {payment_date_val} by {current_user.name}"
    )
    
    # Notify all authorized users
    try:
        title = "One-time Payment Date Scheduled"
        msg = (
            f"Request #{request_id} from {getattr(req.user, 'name', 'Unknown')} in the {req.department} department "
            f"has been scheduled for {payment_date_val.strftime('%B %d, %Y')} by {current_user.name}."
        )
        notified_user_ids = set()

        def add_user(u):
            if u and getattr(u, 'user_id', None):
                notified_user_ids.add(u.user_id)

        # Requestor
        add_user(req.user)

        # Department Manager of the requestor's department (all in that role/department)
        try:
            dept_managers = User.query.filter_by(role='Department Manager', department=req.user.department).all()
        except Exception:
            dept_managers = []
        for u in dept_managers:
            add_user(u)

        # Assigned Manager (manager_id)
        if getattr(req.user, 'manager_id', None):
            assigned_manager = User.query.get(req.user.manager_id)
            add_user(assigned_manager)

        # General Managers (all)
        try:
            gms = User.query.filter_by(role='GM').all()
        except Exception:
            gms = []
        for u in gms:
            add_user(u)

        # Operation Managers (all)
        try:
            opms = User.query.filter_by(role='Operation Manager').all()
        except Exception:
            opms = []
        for u in opms:
            add_user(u)

        # Temporary Manager (if any)
        if getattr(req, 'temporary_manager_id', None):
            temp_manager = User.query.get(req.temporary_manager_id)
            add_user(temp_manager)

        # Finance Admin notifications (conditional rules)
        try:
            finance_admins = User.query.filter_by(role='Finance Admin').all()
        except Exception:
            finance_admins = []
        # Find Abdalaziz specifically
        abdalaziz = next((u for u in finance_admins if u and u.name == 'Abdalaziz Al-Brashdi'), None)
        others_finance_admins = [u for u in finance_admins if u and (not abdalaziz or u.user_id != abdalaziz.user_id)]

        # Notify Abdalaziz only if (is requestor OR assigned manager) AND status is PFA
        if abdalaziz and req.status == 'Pending Finance Approval':
            if (req.user_id == abdalaziz.user_id) or (getattr(req.user, 'manager_id', None) == abdalaziz.user_id):
                add_user(abdalaziz)

        # Notify other Finance Admins only if they are the requestor AND status is PFA
        if req.status == 'Pending Finance Approval':
            for fa in others_finance_admins:
                if req.user_id == fa.user_id:
                    add_user(fa)

        # IT Department: IT Staff + IT Department Managers
        try:
            it_staff = User.query.filter_by(department='IT', role='IT Staff').all()
        except Exception:
            it_staff = []
        try:
            it_mgrs = User.query.filter_by(department='IT', role='Department Manager').all()
        except Exception:
            it_mgrs = []
        for u in it_staff + it_mgrs:
            add_user(u)

        # Create notifications (deduped)
        for uid in notified_user_ids:
            create_notification(
                user_id=uid,
                title=title,
                message=msg,
                notification_type='one_time_payment_scheduled',
                request_id=request_id
            )

        # Emit broadcast signal for clients to refresh notifications
        try:
            socketio.emit('new_notification', {
                'title': title,
                'message': msg,
                'type': 'one_time_payment_scheduled',
                'request_id': request_id
            }, room='all_users')
            socketio.emit('notification_update', {
                'action': 'new_notification',
                'type': 'one_time_payment_scheduled'
            }, room='all_users')
        except Exception as e:
            print(f"DEBUG: WebSocket emit failed for notifications: {e}")
    except Exception as e:
        print(f"DEBUG: Failed to create notifications for schedule_one_time_payment: {e}")
    try:
        socketio.emit('request_updated', {
            'request_id': request_id,
            'one_time_payment_scheduled': True,
            'payment_date': payment_date_val.isoformat()
        })
    except Exception as e:
        print(f"DEBUG: socket emit failed for schedule_one_time_payment: {e}")

    flash('Payment date scheduled successfully.', 'success')
    return redirect(url_for('view_request', request_id=request_id))


@app.route('/request/<int:request_id>/edit', methods=['POST'])
@login_required
def edit_request(request_id):
    """Save inline edits by IT users when status is Pending Manager Approval."""
    req = PaymentRequest.query.get_or_404(request_id)

    # Authorization: Only IT Staff and IT Department Manager
    if not (current_user.department == 'IT' and current_user.role in ['IT Staff', 'Department Manager']):
        flash('You are not authorized to edit this request.', 'error')
        return redirect(url_for('view_request', request_id=request_id))

    # Status gate
    if req.status != 'Pending Manager Approval':
        flash('This request cannot be edited in its current status.', 'error')
        return redirect(url_for('view_request', request_id=request_id))

    # Collect form fields (only those we allow)
    # Track original values to compute which fields changed
    original = {
        'request_type': req.request_type or '',
        'branch_name': req.branch_name or '',
        'person_company': req.person_company or '',
        'company_name': req.company_name or '',
        'purpose': req.purpose or '',
        'bank_name': req.bank_name or '',
        'account_name': req.account_name or '',
        'account_number': req.account_number or '',
        'item_name': req.item_name or ''
    }
    new_request_type = request.form.get('request_type') or req.request_type
    others_description = (request.form.get('others_description') or '').strip()
    if new_request_type == 'Others':
        if not others_description:
            flash('Please specify the type of request for "Others".', 'error')
            return redirect(url_for('view_request', request_id=request_id))
        new_request_type = f"Others: {others_description}"

    req.request_type = new_request_type
    req.branch_name = request.form.get('branch_name') or req.branch_name
    req.person_company = request.form.get('person_company') or req.person_company
    req.company_name = request.form.get('company_name') or req.company_name
    req.purpose = request.form.get('purpose') or req.purpose
    req.bank_name = request.form.get('bank_name') or req.bank_name
    req.account_name = request.form.get('account_name') or req.account_name
    req.account_number = request.form.get('account_number') or req.account_number
    req.item_name = request.form.get('item_name') or req.item_name

    db.session.commit()

    # Build edited_fields list for UI badges: only fields submitted AND changed this save
    def norm(v):
        return (v or '').strip()
    updated = {
        'request_type': req.request_type or '',
        'branch_name': req.branch_name or '',
        'person_company': req.person_company or '',
        'company_name': req.company_name or '',
        'purpose': req.purpose or '',
        'bank_name': req.bank_name or '',
        'account_name': req.account_name or '',
        'account_number': req.account_number or '',
        'item_name': req.item_name or ''
    }
    submitted_keys = set(request.form.keys())
    # Map form field names to our keys
    form_to_key = {
        'request_type': 'request_type',
        'others_description': 'request_type',  # affects request_type
        'branch_name': 'branch_name',
        'person_company': 'person_company',
        'company_name': 'company_name',
        'purpose': 'purpose',
        'bank_name': 'bank_name',
        'account_name': 'account_name',
        'account_number': 'account_number',
        'item_name': 'item_name'
    }
    candidate_keys = set(form_to_key[k] for k in submitted_keys if k in form_to_key)
    edited_fields = [key for key in candidate_keys if norm(original.get(key, '')) != norm(updated.get(key, ''))]

    # Persist cumulative edited fields (create table if missing, upsert per field)
    try:
        db.session.execute(db.text('''
            CREATE TABLE IF NOT EXISTS request_field_edits (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_id INTEGER NOT NULL,
                field_name TEXT NOT NULL,
                first_edited_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                last_edited_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(request_id, field_name)
            );
        '''))
        for key in edited_fields:
            db.session.execute(db.text('''
                INSERT INTO request_field_edits (request_id, field_name) VALUES (:rid, :fname)
                ON CONFLICT(request_id, field_name) DO UPDATE SET last_edited_at = CURRENT_TIMESTAMP
            '''), { 'rid': request_id, 'fname': key })
        # Also persist detailed edit history per field (old/new values)
        db.session.execute(db.text('''
            CREATE TABLE IF NOT EXISTS request_field_edit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_id INTEGER NOT NULL,
                field_name TEXT NOT NULL,
                old_value TEXT,
                new_value TEXT,
                edited_by_user_id INTEGER,
                edited_by_name TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            );
        '''))
        for key in edited_fields:
            db.session.execute(db.text('''
                INSERT INTO request_field_edit_logs (request_id, field_name, old_value, new_value, edited_by_user_id, edited_by_name)
                VALUES (:rid, :fname, :oldv, :newv, :uid, :uname)
            '''), {
                'rid': request_id,
                'fname': key,
                'oldv': str(original.get(key, '')),
                'newv': str(updated.get(key, '')),
                'uid': current_user.user_id,
                'uname': current_user.name
            })
        db.session.commit()
    except Exception as e:
        app.logger.warning(f"Failed to persist edited fields: {e}")

    # Create notifications for all authorized persons
    try:
        recipients = set()
        # Requestor
        recipients.add(req.user_id)
        # All IT department users (Staff + Department Manager)
        it_users = User.query.filter(User.department == 'IT', User.role.in_(['IT Staff', 'Department Manager'])).all()
        for u in it_users:
            recipients.add(u.user_id)
        # Assigned manager (direct manager)
        if getattr(req.user, 'manager_id', None):
            recipients.add(req.user.manager_id)
        # Department Manager(s) of the requestor's department (in addition to direct manager)
        dept_manager_users = User.query.filter_by(role='Department Manager', department=req.user.department).all()
        for u in dept_manager_users:
            if u.user_id != req.user_id:
                recipients.add(u.user_id)
        # Temporary manager
        if getattr(req, 'temporary_manager_id', None):
            recipients.add(req.temporary_manager_id)
        # GM(s)
        for u in User.query.filter_by(role='GM').all():
            recipients.add(u.user_id)
        # Operation Manager(s)
        for u in User.query.filter_by(role='Operation Manager').all():
            recipients.add(u.user_id)
        # Finance Admins are NOT notified here (unless they are the requestor, handled above)

        title = 'Request Updated'
        requestor_name = getattr(req.user, 'name', 'Unknown')
        for user_id in recipients:
            if not user_id or user_id == current_user.user_id:
                continue
            if user_id == req.user_id:
                # Personalized message for the requestor
                message = f"Your request #{req.request_id} has been edited by {current_user.name}."
            else:
                # Include requestor name for all others
                message = f"{requestor_name}'s request #{req.request_id} has been edited by {current_user.name}."
            create_notification(user_id, title, message, 'status_changed', request_id=req.request_id)
    except Exception as e:
        app.logger.warning(f"Failed to create edit notifications: {e}")

    # Real-time emit to all relevant rooms
    try:
        emit_request_update_to_all_rooms('request_updated', {
            'request_id': req.request_id,
            'status': req.status,
            'department': req.department,
            'request_type': req.request_type,
            'branch_name': req.branch_name,
            'action': 'edited'
        })
    except Exception:
        pass

    # Audit log entry
    try:
        requestor_name = getattr(req.user, 'name', 'Unknown')
        log_action(f"{requestor_name}'s request #{request_id} edited by {current_user.name} - type: {req.request_type}, branch: {req.branch_name}")
    except Exception:
        pass

    flash('Edits saved successfully.', 'success')
    return redirect(url_for('view_request', request_id=request_id, tab='submit', edited='1', edited_fields=','.join(edited_fields)))


@app.route('/request/<int:request_id>/field_history')
@login_required
def request_field_history(request_id: int):
    """Return JSON history entries for a specific field for this request."""
    field = request.args.get('field', '').strip()
    if not field:
        return jsonify({'success': False, 'error': 'Missing field parameter'}), 400

    # Basic permission: reuse view permission; if user can't view the request, block
    req = PaymentRequest.query.get_or_404(request_id)
    # Minimal reuse: allow anyone who can open the view_request page (same route did checks)
    # Here, we do a simplified check: if current user is the requestor or IT/GM/Operation Manager/Finance Admin/Department Manager
    if not (
        current_user.user_id == req.user_id or
        current_user.department == 'IT' or
        current_user.role in ['GM', 'Operation Manager', 'Finance Admin', 'Department Manager']
    ):
        return jsonify({'success': False, 'error': 'Not authorized'}), 403

    try:
        db.session.execute(db.text('''
            CREATE TABLE IF NOT EXISTS request_field_edit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                request_id INTEGER NOT NULL,
                field_name TEXT NOT NULL,
                old_value TEXT,
                new_value TEXT,
                edited_by_user_id INTEGER,
                edited_by_name TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            );
        '''))
        rows = db.session.execute(
            db.text('''SELECT old_value, new_value, edited_by_name, created_at
                       FROM request_field_edit_logs
                       WHERE request_id = :rid AND field_name = :fname
                       ORDER BY created_at DESC'''),
            {'rid': request_id, 'fname': field}
        ).fetchall()
        history = [
            {
                'old_value': r[0],
                'new_value': r[1],
                'edited_by': r[2],
                'edited_at': r[3]
            }
            for r in rows
        ]
        return jsonify({'success': True, 'field': field, 'history': history})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/request/<int:request_id>/approve', methods=['POST'])
@login_required
@role_required('Finance Admin')
def approve_request(request_id):
    """Approve a payment request (Finance approval)"""
    req = PaymentRequest.query.get_or_404(request_id)
    
    # Check if request is in correct status for Finance approval
    if req.status not in ['Pending Manager Approval', 'Pending Finance Approval', 'Proof Sent']:
        flash('This request is not ready for Finance approval.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    # Get form data
    approval_status = request.form.get('approval_status')
    
    if approval_status == 'approve':
        # Automatically assign the logged-in finance admin user as the approver
        approver = current_user.name
        proof_required = request.form.get('proof_required') == 'on'
        today = datetime.utcnow().date()
        
        # Require reference number for Finance Admin approval
        reference_number = request.form.get('reference_number', '').strip()
        if not reference_number:
            flash('Reference number is required for Finance Admin approval.', 'error')
            return redirect(url_for('view_request', request_id=request_id))
        
        # Validate reference number is alphanumeric
        if not re.match(r'^[A-Za-z0-9]+$', reference_number):
            flash('Reference number must contain only letters and numbers.', 'error')
            return redirect(url_for('view_request', request_id=request_id))
        
        # Require receipt upload for Finance Admin approval
        if 'receipt_files' not in request.files:
            flash('Receipt upload is required for Finance Admin approval.', 'error')
            return redirect(url_for('view_request', request_id=request_id))
        
        receipt_files = request.files.getlist('receipt_files')
        if not receipt_files or not any(f.filename for f in receipt_files):
            flash('Receipt upload is required for Finance Admin approval.', 'error')
            return redirect(url_for('view_request', request_id=request_id))
        
        # Handle multiple receipt uploads
        uploaded_files = []
        allowed_extensions = {'pdf', 'jpg', 'jpeg', 'png', 'doc', 'docx', 'xls', 'xlsx'}
        
        for receipt_file in receipt_files:
            if receipt_file and receipt_file.filename:
                # Validate file size (50MB max)
                max_file_size = app.config.get('MAX_FILE_SIZE', 50 * 1024 * 1024)
                if len(receipt_file.read()) > max_file_size:
                    flash(f'File "{receipt_file.filename}" is too large. Maximum size is {max_file_size // (1024 * 1024)}MB.', 'error')
                    return redirect(url_for('view_request', request_id=request_id))
                
                # Reset file pointer
                receipt_file.seek(0)
                
                # Validate file extension
                file_extension = receipt_file.filename.rsplit('.', 1)[1].lower() if '.' in receipt_file.filename else ''
                if file_extension not in allowed_extensions:
                    flash(f'Invalid file type for "{receipt_file.filename}". Allowed types: PDF, JPG, PNG, DOC, DOCX, XLS, XLSX', 'error')
                    return redirect(url_for('view_request', request_id=request_id))
                
                # Generate unique filename
                filename = secure_filename(receipt_file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"receipt_{timestamp}_{filename}"
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                receipt_file.save(filepath)
                uploaded_files.append(filename)
        
        if not uploaded_files:
            flash('No valid receipt files were uploaded.', 'error')
            return redirect(url_for('view_request', request_id=request_id))
        
        # Store all finance admin receipts as JSON string in finance_admin_receipt_path
        # Never overwrite requestor receipts - they stay in requestor_receipt_path
        import json
        finance_receipt_path = json.dumps(uploaded_files)
        req.finance_admin_receipt_path = finance_receipt_path
        
        # Save reference number
        req.reference_number = reference_number
        
        req.approver = approver
        req.proof_required = proof_required
        req.updated_at = datetime.utcnow()
        
        if proof_required:
            # Proof is required - set status to Proof Pending
            req.status = 'Proof Pending'
            flash(f'Payment request #{request_id} approved. Waiting for proof of payment from department.', 'info')
            log_action(f"Approved payment request #{request_id} - Proof required")
            
            # Notify the requestor
            create_notification(
                user_id=req.user_id,
                title="Proof of Payment Required",
                message=f"Your payment request #{request_id} has been approved. Please upload proof of payment.",
                notification_type="proof_required",
                request_id=request_id
            )
            
            # Emit real-time update
            socketio.emit('request_updated', {
                'request_id': request_id,
                'status': 'Proof Pending',
                'approver': approver
            })
        else:
            # No proof required - check if it's a recurring payment
            if req.recurring == 'Recurring':
                # Recurring payment - set status to Recurring
                req.status = 'Recurring'
                req.approval_date = today  # Set approval_date when status becomes Recurring
                
                # End finance approval timing when recurring payment is approved
                if req.finance_approval_start_time and not req.finance_approval_end_time:
                    current_time = datetime.utcnow()
                    req.finance_approval_end_time = current_time
                    duration = current_time - req.finance_approval_start_time
                    req.finance_approval_duration_minutes = int(duration.total_seconds())
                
                # Automatically mark the first installment as paid
                first_installment = RecurringPaymentSchedule.query.filter_by(
                    request_id=request_id
                ).order_by(RecurringPaymentSchedule.payment_order).first()
                
                if first_installment:
                    first_installment.is_paid = True
                    first_installment.paid_date = today
                    # Copy the finance admin receipt to the first installment (only if it doesn't already have one)
                    if not first_installment.receipt_path and req.finance_admin_receipt_path:
                        import json
                        finance_receipts = json.loads(req.finance_admin_receipt_path)
                        if finance_receipts:
                            first_installment.receipt_path = finance_receipts[0]
                    
                    # Create a paid notification for the first installment
                    create_notification(
                        user_id=req.user_id,
                        title="First Installment Paid",
                        message=f'First installment for {first_installment.payment_date} has been automatically marked as paid (Amount: {first_installment.amount} OMR)',
                        notification_type="installment_paid",
                        request_id=request_id
                    )
                    
                    # Also notify Finance Admin and Finance Staff
                    finance_users = User.query.filter(User.role.in_(['Finance Staff', 'Finance Admin'])).all()
                    for user in finance_users:
                        create_notification(
                            user_id=user.user_id,
                            title="First Installment Paid",
                            message=f'First installment for {first_installment.payment_date} has been automatically marked as paid (Amount: {first_installment.amount} OMR)',
                            notification_type="installment_paid",
                            request_id=request_id
                        )
                
                flash(f'Recurring payment request #{request_id} approved. First installment automatically marked as paid. Payment schedule will be managed.', 'success')
                log_action(f"Approved recurring payment request #{request_id} - No proof required - First installment marked as paid")
                
                # Notify the requestor
                create_notification(
                    user_id=req.user_id,
                    title="Recurring Payment Approved",
                    message=f"Your recurring payment request #{request_id} has been approved. Payment schedule will be managed.",
                    notification_type="recurring_approved",
                    request_id=request_id
                )
                
                # Notify Auditing Staff and Auditing Department Manager
                auditing_users = User.query.filter(
                    db.and_(
                        User.department == 'Auditing',
                        User.role.in_(['Auditing Staff', 'Department Manager'])
                    )
                ).all()
                for auditing_user in auditing_users:
                    create_notification(
                        user_id=auditing_user.user_id,
                        title="Recurring Payment Approved",
                        message=f"Recurring payment request #{request_id} from {req.department} department has been approved by Finance. Payment schedule will be managed.",
                        notification_type="recurring_approved",
                        request_id=request_id
                    )
                
                # Check if all installments are now paid and mark as completed if so
                check_recurring_payment_completion(request_id)
                
                # Emit real-time update
                socketio.emit('request_updated', {
                    'request_id': request_id,
                    'status': 'Recurring',
                    'approver': approver,
                    'recurring': True
                })
            else:
                # One-time payment - set status to Completed
                req.status = 'Completed'
                req.completion_date = today
                req.approval_date = today  # Set approval_date when status becomes Completed
                
                # End finance approval timing when completed
                if req.finance_approval_start_time and not req.finance_approval_end_time:
                    current_time = datetime.utcnow()
                    req.finance_approval_end_time = current_time
                    duration = current_time - req.finance_approval_start_time
                    req.finance_approval_duration_minutes = int(duration.total_seconds())
                flash(f'Payment request #{request_id} approved and completed. No proof of payment required.', 'success')
                log_action(f"Approved and completed payment request #{request_id} - No proof required")
                
                # Notify the requestor
                create_notification(
                    user_id=req.user_id,
                    title="Request Completed",
                    message=f"Your payment request #{request_id} has been approved and completed. No proof of payment was required.",
                    notification_type="request_completed",
                    request_id=request_id
                )
                
                # Notify Auditing Staff and Auditing Department Manager
                auditing_users = User.query.filter(
                    db.and_(
                        User.department == 'Auditing',
                        User.role.in_(['Auditing Staff', 'Department Manager'])
                    )
                ).all()
                for auditing_user in auditing_users:
                    create_notification(
                        user_id=auditing_user.user_id,
                        title="Request Completed",
                        message=f"Payment request #{request_id} from {req.department} department has been completed.",
                        notification_type="request_completed",
                        request_id=request_id
                    )
                
                # Emit real-time update
                socketio.emit('request_updated', {
                    'request_id': request_id,
                    'status': 'Completed',
                    'approver': approver,
                    'completed': True
                })
    
    
    elif approval_status == 'paid':
        # Mark as paid - requires receipt upload
        if 'receipt_files' not in request.files:
            flash('Receipt upload is required to mark as paid.', 'error')
            return redirect(url_for('view_request', request_id=request_id))
        
        receipt_files = request.files.getlist('receipt_files')
        if not receipt_files or not any(f.filename for f in receipt_files):
            flash('Receipt upload is required to mark as paid.', 'error')
            return redirect(url_for('view_request', request_id=request_id))
        
        # Require reference number when marking as paid
        reference_number = request.form.get('reference_number', '').strip()
        if not reference_number:
            flash('Reference number is required when marking as paid.', 'error')
            return redirect(url_for('view_request', request_id=request_id))
        
        # Validate reference number is alphanumeric
        if not re.match(r'^[A-Za-z0-9]+$', reference_number):
            flash('Reference number must contain only letters and numbers.', 'error')
            return redirect(url_for('view_request', request_id=request_id))
        
        # Handle multiple receipt uploads
        uploaded_files = []
        allowed_extensions = {'pdf', 'jpg', 'jpeg', 'png', 'doc', 'docx', 'xls', 'xlsx'}
        
        for receipt_file in receipt_files:
            if receipt_file and receipt_file.filename:
                # Validate file size (50MB max)
                max_file_size = app.config.get('MAX_FILE_SIZE', 50 * 1024 * 1024)
                if len(receipt_file.read()) > max_file_size:
                    flash(f'File "{receipt_file.filename}" is too large. Maximum size is {max_file_size // (1024 * 1024)}MB.', 'error')
                    return redirect(url_for('view_request', request_id=request_id))
                
                # Reset file pointer
                receipt_file.seek(0)
                
                # Validate file extension
                file_extension = receipt_file.filename.rsplit('.', 1)[1].lower() if '.' in receipt_file.filename else ''
                if file_extension not in allowed_extensions:
                    flash(f'Invalid file type for "{receipt_file.filename}". Allowed types: PDF, JPG, PNG, DOC, DOCX, XLS, XLSX', 'error')
                    return redirect(url_for('view_request', request_id=request_id))
                
                # Generate unique filename
                filename = secure_filename(receipt_file.filename)
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                filename = f"receipt_{timestamp}_{filename}"
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                receipt_file.save(filepath)
                uploaded_files.append(filename)
        
        if not uploaded_files:
            flash('No valid receipt files were uploaded.', 'error')
            return redirect(url_for('view_request', request_id=request_id))
        
        # Store all finance admin receipts as JSON string in finance_admin_receipt_path
        # Never overwrite requestor receipts - they stay in requestor_receipt_path
        import json
        finance_receipt_path = json.dumps(uploaded_files)
        req.finance_admin_receipt_path = finance_receipt_path
        
        # Save reference number
        req.reference_number = reference_number
        
        req.status = 'Completed'
        req.approval_date = datetime.utcnow().date()  # Set approval_date when status becomes Completed
        req.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        log_action(f"Finance admin marked payment request #{request_id} as completed")
        
        # Emit real-time update to all users
        emit_request_update_to_all_rooms('request_updated', {
            'request_id': request_id,
            'status': 'Completed',
            'paid': True
        })
        
        flash(f'Payment request #{request_id} has been marked as paid.', 'success')
    
    elif approval_status == 'proof_sent_approve':
        # Approve proof sent by requestor
        current_time = datetime.utcnow()
        
        # Check if this is a recurring payment
        if req.recurring == 'Recurring':
            # For recurring payments, set status to Recurring and handle payment schedule
            req.status = 'Recurring'
            req.approval_date = current_time.date()  # Set approval_date when status becomes Recurring
            
            # End finance approval timing when recurring payment is approved
            if req.finance_approval_start_time and not req.finance_approval_end_time:
                req.finance_approval_end_time = current_time
                duration = current_time - req.finance_approval_start_time
                req.finance_approval_duration_minutes = int(duration.total_seconds())
            
            # Automatically mark the first installment as paid
            first_installment = RecurringPaymentSchedule.query.filter_by(
                request_id=request_id
            ).order_by(RecurringPaymentSchedule.payment_order).first()
            
            if first_installment:
                first_installment.is_paid = True
                first_installment.paid_date = current_time.date()
                # Copy the finance admin receipt to the first installment (only if it doesn't already have one)
                if not first_installment.receipt_path and req.finance_admin_receipt_path:
                    import json
                    finance_receipts = json.loads(req.finance_admin_receipt_path)
                    if finance_receipts:
                        first_installment.receipt_path = finance_receipts[0]
                
                # Create a paid notification for the first installment
                create_notification(
                    user_id=req.user_id,
                    title="First Installment Paid",
                    message=f'First installment for {first_installment.payment_date} has been automatically marked as paid (Amount: {first_installment.amount} OMR)',
                    notification_type="installment_paid",
                    request_id=request_id
                )
                
                # Also notify Finance Admin and Finance Staff
                finance_users = User.query.filter(User.role.in_(['Finance Staff', 'Finance Admin'])).all()
                for user in finance_users:
                    create_notification(
                        user_id=user.user_id,
                        title="First Installment Paid",
                        message=f'First installment for {first_installment.payment_date} has been automatically marked as paid (Amount: {first_installment.amount} OMR)',
                        notification_type="installment_paid",
                        request_id=request_id
                    )
            
            req.updated_at = current_time
            db.session.commit()
            
            log_action(f"Finance admin approved proof for recurring payment request #{request_id}")
            
            # Notify the requestor
            create_notification(
                user_id=req.user_id,
                title="Recurring Payment Approved",
                message=f"Your proof for recurring payment request #{request_id} has been approved. Payment schedule is now active.",
                notification_type="proof_approved",
                request_id=request_id
            )
            
            # Emit real-time update
            socketio.emit('request_updated', {
                'request_id': request_id,
                'status': 'Recurring',
                'proof_approved': True,
                'recurring': True
            })
            
            flash(f'Recurring payment request #{request_id} has been approved. Payment schedule is now active.', 'success')
        else:
            # For non-recurring payments, set status to Completed when proof is approved
            # Don't end finance approval timing here - it should continue until completed
            # The timer should continue running until status is Completed
            
            req.status = 'Completed'
            req.updated_at = current_time
            req.completion_date = datetime.utcnow().date()
            
            db.session.commit()
            
            log_action(f"Finance admin approved proof for payment request #{request_id}")
            
            # Notify the requestor
            create_notification(
                user_id=req.user_id,
                title="Proof Approved",
                message=f"Your proof for payment request #{request_id} has been approved. Status updated to Completed.",
                notification_type="proof_approved",
                request_id=request_id
            )
            
            # Notify Auditing Staff and Auditing Department Manager
            auditing_users = User.query.filter(
                db.and_(
                    User.department == 'Auditing',
                    User.role.in_(['Auditing Staff', 'Department Manager'])
                )
            ).all()
            for auditing_user in auditing_users:
                create_notification(
                    user_id=auditing_user.user_id,
                    title="Request Completed",
                    message=f"Payment request #{request_id} from {req.department} department has been completed (proof approved).",
                    notification_type="request_completed",
                    request_id=request_id
                )
            
            # Emit real-time update
            socketio.emit('request_updated', {
                'request_id': request_id,
                'status': 'Completed',
                'proof_approved': True
            })
            
            flash(f'Proof for payment request #{request_id} has been approved.', 'success')
    
    elif approval_status == 'proof_sent_reject':
        # Reject proof sent by requestor
        rejection_reason = request.form.get('rejection_reason', '').strip()
        if not rejection_reason:
            flash('Please provide a reason for rejection.', 'error')
            return redirect(url_for('view_request', request_id=request_id))
        
        req.status = 'Proof Rejected'  # Set to Proof Rejected status
        req.rejection_reason = rejection_reason
        req.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        log_action(f"Finance admin rejected proof for payment request #{request_id} - Reason: {rejection_reason}")
        
        # Notify the requestor
        create_notification(
            user_id=req.user_id,
            title="Proof Rejected",
            message=f"Your proof for payment request #{request_id} has been rejected. Please review the feedback and resubmit.",
            notification_type="proof_rejected",
            request_id=request_id
        )
        
        # Emit real-time update
        socketio.emit('request_updated', {
            'request_id': request_id,
            'status': 'Proof Pending',
            'proof_rejected': True
        })
        
        flash(f'Proof for payment request #{request_id} has been rejected.', 'success')
    
    elif approval_status == 'reject':
        # Finance admin rejects - request is rejected
        rejection_reason = request.form.get('rejection_reason', '').strip()
        if not rejection_reason:
            flash('Please provide a reason for rejection.', 'error')
            return redirect(url_for('view_request', request_id=request_id))
        
        current_time = datetime.utcnow()
        
        # End finance approval timing (rejection stops the timer)
        if req.finance_approval_start_time:
            req.finance_approval_end_time = current_time
            duration = current_time - req.finance_approval_start_time
            req.finance_approval_duration_minutes = int(duration.total_seconds())
        
        req.status = 'Rejected by Finance'
        req.rejection_reason = rejection_reason
        req.finance_rejection_date = current_time.date()
        req.updated_at = current_time
        
        db.session.commit()
        
        log_action(f"Finance admin rejected payment request #{request_id} - Reason: {rejection_reason}")
        
        # Notify the requestor
        create_notification(
            user_id=req.user_id,
            title="Payment Request Rejected",
            message=f"Your payment request #{request_id} has been rejected by Finance. Please review the feedback.",
            notification_type="request_rejected",
            request_id=request_id
        )
        
        # Emit real-time update
        socketio.emit('request_updated', {
            'request_id': request_id,
            'status': 'Rejected by Finance',
            'finance_rejected': True
        })
        
        flash(f'Payment request #{request_id} has been rejected by Finance.', 'success')
    
    
    else:
        flash('Invalid approval status selected.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    db.session.commit()
    return redirect(url_for('view_request', request_id=request_id))


@app.route('/request/<int:request_id>/upload_additional_files', methods=['POST'])
@login_required
@role_required('Finance Admin')
def upload_additional_files(request_id):
    """Upload additional files to an approved request"""
    req = PaymentRequest.query.get_or_404(request_id)
    
    # Check if request is in correct status for additional file upload
    if req.status not in ['Proof Pending', 'Proof Sent', 'Proof Rejected', 'Completed', 'Recurring']:
        flash('This request is not in a state that allows file uploads.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    # Get the note about the files
    file_note = request.form.get('file_note', '').strip()
    
    # Handle file uploads
    files = request.files.getlist('additional_files')
    print(f"DEBUG: Received {len(files)} files from request")
    for i, file in enumerate(files):
        print(f"DEBUG: File {i}: {file.filename if file else 'None'}")
    
    uploaded_files = []
    allowed_extensions = {'pdf', 'jpg', 'jpeg', 'png', 'doc', 'docx', 'xls', 'xlsx'}
    
    # First, validate all files and collect valid ones
    valid_files = []
    validation_errors = []
    
    for file in files:
        if file and file.filename:
            # Validate file size (50MB max)
            max_file_size = app.config.get('MAX_FILE_SIZE', 50 * 1024 * 1024)
            if len(file.read()) > max_file_size:
                validation_errors.append(f'File "{file.filename}" is too large. Maximum size is {max_file_size // (1024 * 1024)}MB.')
                continue
            
            # Reset file pointer
            file.seek(0)
            
            # Validate file extension
            file_extension = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
            if file_extension not in allowed_extensions:
                validation_errors.append(f'Invalid file type for "{file.filename}". Allowed types: PDF, JPG, PNG, DOC, DOCX, XLS, XLSX')
                continue
            
            # File is valid, add to valid files list
            valid_files.append(file)
    
    # Show validation errors if any
    if validation_errors:
        for error in validation_errors:
            flash(error, 'error')
    
    # Process only valid files
    for file in valid_files:
        filename = secure_filename(file.filename)
        # Add timestamp to filename to avoid conflicts
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"additional_{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        uploaded_files.append(filename)
    
    if uploaded_files:
        # Store additional files in the database
        import json
        
        # Get existing additional files
        existing_files = []
        if req.additional_files:
            try:
                existing_files = json.loads(req.additional_files)
            except (json.JSONDecodeError, TypeError):
                existing_files = []
        
        # Create file entries with notes
        new_file_entries = []
        for filename in uploaded_files:
            file_entry = {
                'filename': filename,
                'uploaded_at': datetime.utcnow().isoformat(),
                'uploaded_by': current_user.name,
                'note': file_note if file_note else None
            }
            new_file_entries.append(file_entry)
        
        # Add new files to existing ones
        all_files = existing_files + new_file_entries
        
        # Store as JSON string
        req.additional_files = json.dumps(all_files)
        req.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        note_text = f" with note: '{file_note}'" if file_note else ""
        log_action(f"Finance admin uploaded {len(uploaded_files)} additional files to request #{request_id}{note_text}")
        flash(f'Successfully uploaded {len(uploaded_files)} additional files.', 'success')
    else:
        flash('No valid files were uploaded.', 'warning')
    
    return redirect(url_for('view_request', request_id=request_id))


@app.route('/request/<int:request_id>/mark_as_paid', methods=['POST'])
@login_required
@role_required('Finance Admin')
def mark_as_paid(request_id):
    """Mark a payment pending request as paid"""
    req = PaymentRequest.query.get_or_404(request_id)
    
    # Check if request is in correct status (Proof Sent means proof has been submitted and can be marked as paid/completed)
    if req.status not in ['Proof Sent', 'Proof Pending']:
        flash('This request is not in a valid status for marking as paid.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    # Mark as completed
    req.status = 'Completed'
    req.approval_date = datetime.utcnow().date()  # Set approval_date when status becomes Completed
    req.completion_date = datetime.utcnow().date()
    req.updated_at = datetime.utcnow()
    
    db.session.commit()
    
    log_action(f"Finance admin marked payment request #{request_id} as paid")
    
    # Notify the requestor
    create_notification(
        user_id=req.user_id,
        title="Payment Completed",
        message=f"Your payment request #{request_id} has been paid.",
        notification_type="payment_completed",
        request_id=request_id
    )
    
    # Emit real-time update
    socketio.emit('request_updated', {
        'request_id': request_id,
        'status': 'Approved',
        'paid': True
    })
    
    flash(f'Payment request #{request_id} has been marked as paid.', 'success')
    return redirect(url_for('view_request', request_id=request_id))


@app.route('/request/<int:request_id>/close', methods=['POST'])
@login_required
@role_required('Finance Admin')
def close_request(request_id):
    """Close a request (mark as completed)"""
    req = PaymentRequest.query.get_or_404(request_id)
    
    # Check if request is in correct status
    if req.status not in ['Proof Pending', 'Proof Sent', 'Proof Rejected']:
        flash('This request cannot be closed in its current status.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    # Close the request
    req.status = 'Completed'
    req.completion_date = datetime.utcnow().date()
    req.approval_date = datetime.utcnow().date()  # Set approval_date when status becomes Completed
    req.updated_at = datetime.utcnow()
    
    # End finance approval timing when completed
    if req.finance_approval_start_time and not req.finance_approval_end_time:
        current_time = datetime.utcnow()
        req.finance_approval_end_time = current_time
        duration = current_time - req.finance_approval_start_time
        req.finance_approval_duration_minutes = int(duration.total_seconds())
    
    db.session.commit()
    
    log_action(f"Finance admin closed payment request #{request_id}")
    
    # Notify the requestor
    create_notification(
        user_id=req.user_id,
        title="Request Completed",
        message=f"Your payment request #{request_id} has been completed and closed.",
        notification_type="request_completed",
        request_id=request_id
    )
    
    # Notify Auditing Staff and Auditing Department Manager
    auditing_users = User.query.filter(
        db.and_(
            User.department == 'Auditing',
            User.role.in_(['Auditing Staff', 'Department Manager'])
        )
    ).all()
    for auditing_user in auditing_users:
        create_notification(
            user_id=auditing_user.user_id,
            title="Request Completed",
            message=f"Payment request #{request_id} from {req.department} department has been completed and closed.",
            notification_type="request_completed",
            request_id=request_id
        )
    
    # Emit real-time update
    socketio.emit('request_updated', {
        'request_id': request_id,
        'status': 'Completed',
        'completed': True
    })
    
    flash(f'Payment request #{request_id} has been closed.', 'success')
    return redirect(url_for('view_request', request_id=request_id))


@app.route('/request/<int:request_id>/pending', methods=['POST'])
@login_required
@role_required('Admin')
def mark_pending(request_id):
    """Mark a payment request as pending with reason"""
    req = PaymentRequest.query.get_or_404(request_id)
    
    reason = request.form.get('reason_pending')
    
    if not reason:
        flash('Please provide a reason for marking as pending.', 'warning')
        return redirect(url_for('view_request', request_id=request_id))
    
    req.status = 'Pending'
    req.reason_pending = reason
    req.updated_at = datetime.utcnow()
    
    db.session.commit()
    
    log_action(f"Marked payment request #{request_id} as pending")
    
    # Emit real-time update for pending status
    socketio.emit('request_updated', {
        'request_id': request_id,
        'status': 'Pending',
        'reason': reason
    })
    
    flash(f'Payment request #{request_id} has been marked as pending.', 'info')
    return redirect(url_for('admin_dashboard'))


@app.route('/request/<int:request_id>/upload_proof', methods=['POST'])
@login_required
def upload_proof(request_id):
    """Department user uploads proof of payment"""
    req = PaymentRequest.query.get_or_404(request_id)
    
    # Check if user owns this request
    if req.user_id != current_user.user_id:
        flash('You can only upload proof for your own requests.', 'error')
        return redirect(url_for('dashboard'))
    
    # Check if request is in "Proof Pending" or "Proof Rejected" status
    if req.status not in ['Proof Pending', 'Proof Rejected']:
        flash('This request does not require proof upload.', 'error')
        return redirect(url_for('dashboard'))
    
    # Handle multiple file uploads
    if 'proof_files' not in request.files:
        flash('No files uploaded.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    proof_files = request.files.getlist('proof_files')
    
    if not proof_files or not any(f.filename for f in proof_files):
        flash('No files selected.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    uploaded_files = []
    allowed_extensions = {'pdf', 'jpg', 'jpeg', 'png', 'doc', 'docx', 'xls', 'xlsx'}
    
    # Determine next batch number for this request
    import glob, os
    existing_files_pattern = os.path.join(app.config['UPLOAD_FOLDER'], f"proof_{request_id}_*")
    existing_files = [os.path.basename(f) for f in glob.glob(existing_files_pattern)]
    max_batch = 0
    for ef in existing_files:
        # Expect formats:
        #  - proof_{request_id}_b{batch}_{timestamp}_{original}
        #  - legacy: proof_{request_id}_{timestamp}_{original}
        try:
            if ef.startswith(f"proof_{request_id}_b"):
                after_prefix = ef[len(f"proof_{request_id}_b"):]
                batch_str = after_prefix.split('_', 1)[0]
                batch_num = int(batch_str)
                if batch_num > max_batch:
                    max_batch = batch_num;
            elif ef.startswith(f"proof_{request_id}_"):
                # Treat legacy as batch 1 candidate
                if 1 > max_batch:
                    max_batch = 1
        except Exception:
            pass

    next_batch = max(1, max_batch + 1) if existing_files else 1

    for file in proof_files:
        if file and file.filename:
            # Validate file size (50MB max)
            max_file_size = app.config.get('MAX_FILE_SIZE', 50 * 1024 * 1024)
            if len(file.read()) > max_file_size:
                flash(f'File "{file.filename}" is too large. Maximum size is {max_file_size // (1024 * 1024)}MB.', 'error')
                return redirect(url_for('view_request', request_id=request_id))
            
            # Reset file pointer
            file.seek(0)
            
            # Validate file extension
            file_extension = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
            if file_extension not in allowed_extensions:
                flash(f'Invalid file type for "{file.filename}". Allowed types: PDF, JPG, PNG, DOC, DOCX, XLS, XLSX', 'error')
                return redirect(url_for('view_request', request_id=request_id))
            
            # Generate unique filename
            filename = secure_filename(f"proof_{request_id}_b{next_batch}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}")
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            uploaded_files.append(filename)
    
    if uploaded_files:
        # Update request - store the first file as primary proof, others are additional
        req.proof_of_payment = uploaded_files[0]
        req.status = 'Proof Sent'
        req.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        log_action(f"Uploaded {len(uploaded_files)} proof files for payment request #{request_id}")
        
        # Notify Finance Admin and requestor about proof upload
        notify_users_by_role(
            request=req,
            notification_type="proof_uploaded",
            title="Proof of Payment Uploaded",
            message=f"{len(uploaded_files)} proof file(s) have been uploaded for request #{request_id} by {current_user.name}",
            request_id=request_id
        )
        
        # Emit real-time update to all users
        emit_request_update_to_all_rooms('request_updated', {
            'request_id': request_id,
            'status': 'Proof Sent',
            'requestor': current_user.username
        })
        
        flash(f'Successfully uploaded {len(uploaded_files)} proof file(s)! Finance will review your proof.', 'success')
    else:
        flash('No valid files were uploaded.', 'error')
    
    return redirect(url_for('view_request', request_id=request_id))


@app.route('/request/<int:request_id>/save_finance_note', methods=['POST'])
@login_required
@role_required('Finance Admin')
def save_finance_note(request_id):
    """Save a note from Finance Admin for a request in Pending Finance Approval status"""
    req = PaymentRequest.query.get_or_404(request_id)
    
    # Check if request is in correct status for Finance admin to add notes
    if req.status != 'Pending Finance Approval':
        flash('Notes can only be added to requests in Pending Finance Approval status.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    # Get the note from form data
    note = request.form.get('finance_note', '').strip()
    
    if not note:
        flash('Note cannot be empty.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    # Create a new finance admin note entry
    finance_note = FinanceAdminNote(
        request_id=request_id,
        note_content=note,
        added_by=current_user.name,
        added_by_id=current_user.user_id,
        created_at=datetime.utcnow()
    )
    
    db.session.add(finance_note)
    req.updated_at = datetime.utcnow()
    
    db.session.commit()
    
    log_action(f"Finance admin added note to payment request #{request_id}")
    
    # Notify all authorized persons about the note: requestor, assigned managers (incl. temporary), and other finance admins
    notification_title = "Finance Admin Note Added"
    # Include who added the note for clarity
    notification_message = f"{current_user.name} added a finance admin note to payment request #{request_id}."
    # Notify requestor
    create_notification(
        user_id=req.user_id,
        title=notification_title,
        message=notification_message,
        notification_type="finance_note_added",
        request_id=request_id
    )
    
    # Notify assigned managers and temporary manager (if any)
    try:
        authorized_managers = get_authorized_manager_approvers(req)
    except Exception as e:
        authorized_managers = []
        print(f"DEBUG: get_authorized_manager_approvers failed in save_finance_note: {e}")
    
    notified_user_ids = {req.user_id}
    for manager_user in authorized_managers:
        if manager_user and manager_user.user_id not in notified_user_ids:
            create_notification(
                user_id=manager_user.user_id,
                title=notification_title,
                message=notification_message,
                notification_type="finance_note_added",
                request_id=request_id
            )
            notified_user_ids.add(manager_user.user_id)
    
    # Notify other Finance Admins
    try:
        other_finance_admins = User.query.filter_by(role='Finance Admin').all()
    except Exception as e:
        other_finance_admins = []
        print(f"DEBUG: querying Finance Admins failed in save_finance_note: {e}")
    for fa_user in other_finance_admins:
        if fa_user and fa_user.user_id not in notified_user_ids and fa_user.user_id != current_user.user_id:
            create_notification(
                user_id=fa_user.user_id,
                title=notification_title,
                message=notification_message,
                notification_type="finance_note_added",
                request_id=request_id
            )
            notified_user_ids.add(fa_user.user_id)

    # Notify ALL IT Staff regardless of request owner
    try:
        it_staff_users = User.query.filter_by(role='IT Staff').all()
    except Exception as e:
        it_staff_users = []
        print(f"DEBUG: querying IT Staff failed in save_finance_note: {e}")
    for it_user in it_staff_users:
        if it_user and it_user.user_id not in notified_user_ids:
            create_notification(
                user_id=it_user.user_id,
                title=notification_title,
                message=notification_message,
                notification_type="finance_note_added",
                request_id=request_id
            )
            notified_user_ids.add(it_user.user_id)

    # Notify IT Department Manager(s) regardless of request owner
    try:
        it_managers = User.query.filter_by(role='Department Manager', department='IT').all()
    except Exception as e:
        it_managers = []
        print(f"DEBUG: querying IT Department Managers failed in save_finance_note: {e}")
    for mgr in it_managers:
        if mgr and mgr.user_id not in notified_user_ids:
            create_notification(
                user_id=mgr.user_id,
                title=notification_title,
                message=notification_message,
                notification_type="finance_note_added",
                request_id=request_id
            )
            notified_user_ids.add(mgr.user_id)

    # Notify General Managers (ALL GMs)
    try:
        gm_users = User.query.filter_by(role='GM').all()
    except Exception as e:
        gm_users = []
        print(f"DEBUG: querying GM users failed in save_finance_note: {e}")
    for gm in gm_users:
        if gm and gm.user_id not in notified_user_ids:
            create_notification(
                user_id=gm.user_id,
                title=notification_title,
                message=notification_message,
                notification_type="finance_note_added",
                request_id=request_id
            )
            notified_user_ids.add(gm.user_id)

    # Notify Operation Managers (ALL Operation Managers)
    try:
        op_manager_users = User.query.filter_by(role='Operation Manager').all()
    except Exception as e:
        op_manager_users = []
        print(f"DEBUG: querying Operation Managers failed in save_finance_note: {e}")
    for opm in op_manager_users:
        if opm and opm.user_id not in notified_user_ids:
            create_notification(
                user_id=opm.user_id,
                title=notification_title,
                message=notification_message,
                notification_type="finance_note_added",
                request_id=request_id
            )
            notified_user_ids.add(opm.user_id)
    
    # Emit real-time notification events
    try:
        socketio.emit('new_notification', {
            'title': notification_title,
            'message': notification_message,
            'type': 'finance_note_added',
            'request_id': request_id
        }, room='all_users')
        socketio.emit('notification_update', {
            'action': 'new_notification',
            'type': 'finance_note_added'
        }, room='all_users')
    except Exception as e:
        print(f"Error emitting finance_note_added WebSocket events: {e}")
    
    # Emit real-time update
    socketio.emit('request_updated', {
        'request_id': request_id,
        'finance_note_added': True,
        'note': note
    })
    
    flash('Note saved successfully.', 'success')
    return redirect(url_for('view_request', request_id=request_id))


@app.route('/request/<int:request_id>/final_approve', methods=['POST'])
@login_required
@role_required('Admin')
def final_approve_request(request_id):
    """Admin final approval after receiving proof"""
    req = PaymentRequest.query.get_or_404(request_id)
    
    # Check if request is in "Proof Sent" status
    if req.status != 'Proof Sent':
        flash('This request is not ready for final approval.', 'error')
        return redirect(url_for('admin_dashboard'))
    
    # Final approval - set to Completed
    req.status = 'Completed'
    req.approval_date = datetime.utcnow().date()  # Set approval_date when status becomes Completed
    req.completion_date = datetime.utcnow().date()
    req.updated_at = datetime.utcnow()
    
    db.session.commit()
    
    log_action(f"Final approved payment request #{request_id}")
    
    # Emit real-time update
    socketio.emit('request_updated', {
        'request_id': request_id,
        'status': 'Completed',
        'final_approval': True
    })
    
    flash(f'Payment request #{request_id} has been finally approved.', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/request/<int:request_id>/manager_approve', methods=['POST'])
@login_required
def manager_approve_request(request_id):
    """Manager approves a payment request"""
    req = PaymentRequest.query.get_or_404(request_id)
    
    # Never allow users to approve their own requests at the manager stage
    if req.user_id == current_user.user_id:
        flash('You cannot approve your own request at the manager stage.', 'error')
        return redirect(url_for('view_request', request_id=request_id))

    # Debug information
    print(f"DEBUG: Current user: {current_user.name} (ID: {current_user.user_id}, Role: {current_user.role}, Department: {current_user.department})")
    print(f"DEBUG: Request submitter: {req.user.name} (ID: {req.user.user_id}, Role: {req.user.role}, Department: {req.user.department})")
    print(f"DEBUG: Request submitter's manager_id: {req.user.manager_id}")
    print(f"DEBUG: Request status: {req.status}")
    
    # Check if current user is authorized to approve this request
    is_authorized = False
    
    # First, enforce temporary manager exclusivity (IT Department feature)
    if req.temporary_manager_id:
        if req.temporary_manager_id == current_user.user_id:
            is_authorized = True
            print("DEBUG: Authorized via temporary manager assignment")
        else:
            # When a temporary manager is assigned, only they can approve this request
            print("DEBUG: Approval blocked - temporary manager is assigned and current user is not the assignee")
            is_authorized = False
    else:
        # No temporary manager assigned, use standard authorization checks
        # Hard rule: Requests submitted by GM/CEO/Operation Manager can ONLY be approved by Abdalaziz (Finance Admin)
        if req.user.role in ['GM', 'CEO', 'Operation Manager']:
            if current_user.name == 'Abdalaziz Al-Brashdi':
                is_authorized = True
                print("DEBUG: Authorized via Abdalaziz-only rule for GM/CEO/Operation Manager submitter")
            else:
                is_authorized = False
                print("DEBUG: Blocked - only Abdalaziz can approve GM/CEO/Operation Manager submitter")
        # General rule for other requests
        elif current_user.role in ['GM', 'Operation Manager']:
            is_authorized = True
            print("DEBUG: Authorized via global GM/Operation Manager rule")
        else:
        # Check if current user is the manager of the request submitter
            if req.user.manager_id and req.user.manager_id == current_user.user_id:
                is_authorized = True
                print("DEBUG: Authorized via manager_id relationship")
    
    # If not yet authorized and no temporary manager restriction applied, check special cases
    if not is_authorized and not req.temporary_manager_id:
        # Special case: General Manager can approve Department Manager requests
        if (current_user.role == 'GM' and req.user.role == 'Department Manager'):
            is_authorized = True
            print("DEBUG: Authorized via GM role for Department Manager")
        # Special case: Operation Manager can also approve Department Manager requests (global)
        elif (current_user.role == 'Operation Manager' and req.user.role == 'Department Manager'):
            is_authorized = True
            print("DEBUG: Authorized via Operation Manager role for Department Manager")
        # Special case: Abdalaziz can approve General Manager and CEO requests
        elif (current_user.name == 'Abdalaziz Al-Brashdi' and req.user.role in ['GM','CEO']):
            is_authorized = True
            print("DEBUG: Authorized via Abdalaziz role for GM/CEO")
        # Special case: Abdalaziz can approve Finance Staff requests
        elif (current_user.name == 'Abdalaziz Al-Brashdi' and req.user.role == 'Finance Staff'):
            is_authorized = True
            print("DEBUG: Authorized via Abdalaziz role for Finance Staff")
        # Special case: Abdalaziz can approve Operation Manager requests
        elif (current_user.name == 'Abdalaziz Al-Brashdi' and req.user.role == 'Operation Manager'):
            is_authorized = True
            print("DEBUG: Authorized via Abdalaziz role for Operation Manager")
        # Special case: Operation Manager can approve Operation department and Project requests
        elif (current_user.role == 'Operation Manager' and 
              (req.user.department == 'Operation' or req.user.department == 'Project') and 
              req.user.role != 'Operation Manager'):  # Operation Manager can't approve their own requests
            is_authorized = True
            print("DEBUG: Authorized via Operation Manager role")
        # Special case: Finance Admin can approve Finance department requests
        elif (current_user.role == 'Finance Admin' and 
              req.user.department == 'Finance' and 
              req.user.role != 'Finance Admin'):  # Finance Admin can't approve their own requests
            is_authorized = True
            print("DEBUG: Authorized via Finance Admin role")
        # Special case: Department Manager can approve same department requests
        elif (current_user.role == 'Department Manager' and 
              req.user.department == current_user.department and 
              req.user.role != 'Department Manager'):  # Department Manager can't approve their own requests
            is_authorized = True
            print("DEBUG: Authorized via Department Manager role")
    
    print(f"DEBUG: Authorization result: {is_authorized}")
    
    if not is_authorized:
        flash('You are not authorized to approve this request.', 'error')
        return redirect(url_for('dashboard'))
    
    # Check if request is in correct status
    if req.status != 'Pending Manager Approval':
        flash('This request is not pending manager approval.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get form data
    approval_status = request.form.get('approval_status')
    
    if approval_status == 'approve':
        # Manager approves - move to Finance for final approval
        current_time = datetime.utcnow()
        
        # Start manager approval timing if not already started
        if not req.manager_approval_start_time:
            req.manager_approval_start_time = current_time
        
        # End manager approval timing
        req.manager_approval_end_time = current_time
        
        # Calculate duration in seconds
        if req.manager_approval_start_time:
            duration = current_time - req.manager_approval_start_time
            req.manager_approval_duration_minutes = int(duration.total_seconds())
        
        # Start finance approval timing
        req.finance_approval_start_time = current_time
        
        req.status = 'Pending Finance Approval'
        req.manager_approval_date = current_time.date()
        # Track who actually approved as manager (covers GM, Operation Manager, Department Manager, Finance Admin, and temporary manager)
        req.manager_approver = current_user.name
        req.manager_approver_user_id = current_user.user_id
        req.is_urgent = request.form.get('is_urgent') == 'on'
        req.manager_approval_reason = request.form.get('approval_reason', '').strip()
        req.updated_at = current_time
        
        db.session.commit()
        
        log_action(f"Manager approved payment request #{request_id}")
        
        # Notify the requestor about approval
        create_notification(
            user_id=req.user_id,
            title="Payment Request Approved",
            message=f"Your payment request #{request_id} has been approved by your manager and sent to Finance for final approval.",
            notification_type="request_approved",
            request_id=request_id
        )
        
        # Notify Finance Admin that request is ready for their review
        notify_users_by_role(
            request=req,
            notification_type="ready_for_finance_review",
            title="Payment Request Ready for Review",
            message=f"Payment request #{request_id} from {req.department} department has been approved by manager and is ready for Finance review",
            request_id=request_id
        )
        
        # Emit real-time update to Finance Admin and management dashboards
        socketio.emit('request_updated', {
            'request_id': request_id,
            'status': 'Pending Finance Approval',
            'manager_approved': True
        }, room='finance_admin')
        socketio.emit('request_updated', {
            'request_id': request_id,
            'status': 'Pending Finance Approval',
            'manager_approved': True
        }, room='gm')
        socketio.emit('request_updated', {
            'request_id': request_id,
            'status': 'Pending Finance Approval',
            'manager_approved': True
        }, room='operation_manager')
        socketio.emit('request_updated', {
            'request_id': request_id,
            'status': 'Pending Finance Approval',
            'manager_approved': True
        })
        
        flash(f'Payment request #{request_id} has been approved by manager. Sent to Finance for final approval.', 'success')
        
        
    elif approval_status == 'reject':
        # Manager rejects - request is rejected
        rejection_reason = request.form.get('rejection_reason', '').strip()
        if not rejection_reason:
            flash('Please provide a reason for rejection.', 'error')
            return redirect(url_for('view_request', request_id=request_id))
        
        current_time = datetime.utcnow()
        
        # Start manager approval timing if not already started (for rejection tracking)
        if not req.manager_approval_start_time:
            req.manager_approval_start_time = current_time
        
        # End manager approval timing (rejection stops the timer)
        req.manager_approval_end_time = current_time
        
        # Calculate duration in seconds
        if req.manager_approval_start_time:
            duration = current_time - req.manager_approval_start_time
            req.manager_approval_duration_minutes = int(duration.total_seconds())
        
        req.status = 'Rejected by Manager'
        req.rejection_reason = rejection_reason
        req.manager_rejection_date = current_time.date()
        # Track who actually rejected as manager (covers GM, Operation Manager, Department Manager, Finance Admin, and temporary manager)
        req.manager_rejector = current_user.name
        req.manager_rejector_user_id = current_user.user_id
        req.updated_at = current_time
        
        db.session.commit()
        
        log_action(f"Manager rejected payment request #{request_id} - Reason: {rejection_reason}")
        
        # Notify the requestor
        create_notification(
            user_id=req.user_id,
            title="Payment Request Rejected",
            message=f"Your payment request #{request_id} has been rejected by your manager. Please review the feedback.",
            notification_type="request_rejected",
            request_id=request_id
        )
        
        # Emit real-time update to everyone (including GM and Operation Manager dashboards)
        socketio.emit('request_updated', {
            'request_id': request_id,
            'status': 'Rejected by Manager',
            'manager_rejected': True
        })
        socketio.emit('request_updated', {
            'request_id': request_id,
            'status': 'Rejected by Manager',
            'manager_rejected': True
        }, room='gm')
        socketio.emit('request_updated', {
            'request_id': request_id,
            'status': 'Rejected by Manager',
            'manager_rejected': True
        }, room='operation_manager')
        
        flash(f'Payment request #{request_id} has been rejected by manager.', 'success')
    
    else:
        flash('Invalid approval status selected.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    return redirect(url_for('view_request', request_id=request_id))


@app.route('/request/<int:request_id>/manager_reject', methods=['POST'])
@login_required
def manager_reject_request(request_id):
    """Manager rejects a payment request"""
    req = PaymentRequest.query.get_or_404(request_id)
    
    # Never allow users to reject their own requests at the manager stage
    if req.user_id == current_user.user_id:
        flash('You cannot reject your own request at the manager stage.', 'error')
        return redirect(url_for('view_request', request_id=request_id))

    # Check if current user is authorized to reject this request
    is_authorized = False
    
    # Temporary manager exclusivity: if a temporary manager is assigned, only they can reject
    if req.temporary_manager_id:
        if req.temporary_manager_id == current_user.user_id:
            is_authorized = True
        else:
            is_authorized = False
    else:
        # New global rule: GM and Operation Manager can reject ANY request at manager stage
        if current_user.role in ['GM', 'Operation Manager']:
            is_authorized = True
        # Check if current user is the manager of the request submitter
        elif req.user.manager_id and req.user.manager_id == current_user.user_id:
            is_authorized = True
        # Special case: Operation Manager can reject Operation department requests
        elif (current_user.role == 'Operation Manager' and 
              req.user.department == 'Operation' and 
              req.user.role != 'Operation Manager'):  # Operation Manager can't reject their own requests
            is_authorized = True
        # Special case: Finance Admin can reject Finance department requests
        elif (current_user.role == 'Finance Admin' and 
              req.user.department == 'Finance' and 
              req.user.role != 'Finance Admin'):  # Finance Admin can't reject their own requests
            is_authorized = True
    
    if not is_authorized:
        flash('You are not authorized to reject this request.', 'error')
        return redirect(url_for('dashboard'))
    
    # Check if request is in correct status
    if req.status != 'Pending Manager Approval':
        flash('This request is not pending manager approval.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get rejection reason
    reason = request.form.get('rejection_reason', '').strip()
    if not reason:
        flash('Please provide a reason for rejection.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    # Manager rejects - request is rejected
    req.status = 'Rejected by Manager'
    req.rejection_reason = reason
    req.manager_rejection_date = datetime.utcnow().date()
    # Track who actually rejected as manager
    req.manager_rejector = current_user.name
    req.manager_rejector_user_id = current_user.user_id
    req.updated_at = datetime.utcnow()
    
    db.session.commit()
    
    log_action(f"Manager rejected payment request #{request_id} - Reason: {reason}")
    
    # Emit real-time update
    socketio.emit('request_updated', {
        'request_id': request_id,
        'status': 'Rejected by Manager',
        'manager_rejected': True
    })
    
    flash(f'Payment request #{request_id} has been rejected by manager.', 'success')
    return redirect(url_for('dashboard'))


@app.route('/request/<int:request_id>/reassign_manager', methods=['POST'])
@login_required
@role_required('IT Staff', 'Department Manager')
def reassign_manager(request_id):
    """IT Department can reassign a temporary manager for a specific request"""
    # Only allow IT Department Staff to perform this action
    if current_user.department != 'IT':
        flash('You do not have permission to reassign managers. Only IT Department can perform this action.', 'danger')
        return redirect(url_for('dashboard'))
    
    req = PaymentRequest.query.get_or_404(request_id)
    
    # Only allow reassignment for pending manager approval requests
    if req.status != 'Pending Manager Approval':
        flash('You can only reassign managers for requests with "Pending Manager Approval" status.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    # Get the new manager ID from form
    new_manager_id = request.form.get('temporary_manager_id')
    
    if not new_manager_id:
        flash('Please select a manager.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    # Validate that the manager exists and is a valid manager
    new_manager = User.query.get(new_manager_id)
    if not new_manager:
        flash('Selected manager does not exist.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    # Only allow reassignment to managers and department managers
    if new_manager.role not in ['Department Manager', 'GM', 'Operation Manager', 'Finance Admin']:
        flash('Selected user is not a manager. Please select a valid manager.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    # Don't allow reassigning to the same manager
    if req.temporary_manager_id == int(new_manager_id):
        flash('Selected manager is already assigned as the temporary manager for this request.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    # Get the old temporary manager (if any) for notification
    old_temp_manager = req.temporary_manager
    old_original_manager = None
    if req.user.manager_id:
        old_original_manager = User.query.get(req.user.manager_id)
    
    # Set the temporary manager
    req.temporary_manager_id = int(new_manager_id)
    req.updated_at = datetime.utcnow()
    
    db.session.commit()
    
    log_action(f"Reassigned temporary manager for request #{request_id} to {new_manager.name} (IT Staff only action)")
    
    # Notify the new temporary manager
    create_notification(
        user_id=new_manager_id,
        title="Temporary Manager Assignment",
        message=f"You have been temporarily assigned to review payment request #{request_id} from {req.department} department. The originally assigned manager is not available.",
        notification_type="temporary_manager_assignment",
        request_id=request_id
    )
    
    # Notify the requestor about the manager change
    create_notification(
        user_id=req.user_id,
        title="Manager Reassigned for Your Request",
        message=f"The manager for your payment request #{request_id} has been temporarily reassigned to {new_manager.name}. This only affects this specific request.",
        notification_type="manager_reassigned",
        request_id=request_id
    )
    
    # Notify the old temporary manager (if any) that they're no longer assigned
    if old_temp_manager and old_temp_manager.user_id != int(new_manager_id):
        create_notification(
            user_id=old_temp_manager.user_id,
            title="Temporary Manager Assignment Removed",
            message=f"You are no longer the temporary manager for payment request #{request_id}.",
            notification_type="temporary_manager_unassigned",
            request_id=request_id
        )
    
    # Emit real-time update
    socketio.emit('request_updated', {
        'request_id': request_id,
        'temporary_manager_id': new_manager_id,
        'temporary_manager_name': new_manager.name
    })
    
    flash(f'Temporary manager has been reassigned to {new_manager.name} for request #{request_id}.', 'success')
    return redirect(url_for('view_request', request_id=request_id))


@app.route('/request/<int:request_id>/delete', methods=['POST'])
@login_required
@role_required('IT Staff', 'Department Manager')
def delete_request(request_id):
    """Archive a payment request (IT only) - soft delete"""
    # Restrict Department Managers to IT department only
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    
    req = PaymentRequest.query.get_or_404(request_id)
    
    # Check if already archived
    if req.is_archived:
        flash(f'Payment request #{request_id} is already archived.', 'warning')
        return redirect(url_for('archives'))
    
    # Pause any ongoing approval timing before archiving
    archive_time = datetime.utcnow()
    
    # Pause manager approval timing if in progress
    if req.manager_approval_start_time and not req.manager_approval_end_time:
        # Timing was in progress - pause it by setting end_time
        req.manager_approval_end_time = archive_time
        # Calculate and store duration
        if not req.manager_approval_duration_minutes:
            duration = archive_time - req.manager_approval_start_time
            req.manager_approval_duration_minutes = int(duration.total_seconds())
        print(f"DEBUG: Paused manager approval timing for request #{request_id}")
    
    # Pause finance approval timing if in progress
    if req.finance_approval_start_time and not req.finance_approval_end_time:
        # Timing was in progress - pause it by setting end_time
        req.finance_approval_end_time = archive_time
        # Calculate and store duration
        if not req.finance_approval_duration_minutes:
            duration = archive_time - req.finance_approval_start_time
            req.finance_approval_duration_minutes = int(duration.total_seconds())
        print(f"DEBUG: Paused finance approval timing for request #{request_id}")
    
    # Archive the request (soft delete) - keep all files and related records
    req.is_archived = True
    req.archived_at = archive_time
    req.archived_by = current_user.name
    req.archived_by_user_id = current_user.user_id
    
    db.session.commit()
    
    log_action(f"Archived payment request #{request_id}")
    
    # Notify all IT department users about the archive
    it_users = User.query.filter(
        db.or_(
            User.role == 'IT Staff',
            db.and_(User.role == 'Department Manager', User.department == 'IT')
        )
    ).all()
    
    print(f"DEBUG: Found {len(it_users)} IT department users")
    for u in it_users:
        print(f"DEBUG: IT user: {u.username} (ID: {u.user_id}, Role: {u.role}, Dept: {u.department})")
    
    # Exclude the user who archived (they already know)
    it_users_to_notify = [user for user in it_users if user.user_id != current_user.user_id]
    
    print(f"DEBUG: Will notify {len(it_users_to_notify)} IT users (excluding archiver: {current_user.username})")
    
    notification_title = "Payment Request Archived"
    notification_message = f"Payment request #{request_id} submitted by {req.requestor_name} has been archived by {current_user.name}."
    
    for it_user in it_users_to_notify:
        print(f"DEBUG: Creating notification for IT user: {it_user.username} (ID: {it_user.user_id})")
        create_notification(
            user_id=it_user.user_id,
            title=notification_title,
            message=notification_message,
            notification_type="request_archived",
            request_id=request_id
        )
    
    # Emit real-time notification to all users after creating database notifications
    if it_users_to_notify:
        try:
            socketio.emit('new_notification', {
                'title': notification_title,
                'message': notification_message,
                'type': 'request_archived',
                'request_id': request_id
            }, room='all_users')
            
            # Also emit a general update event to trigger notification count updates
            socketio.emit('notification_update', {
                'action': 'new_notification',
                'type': 'request_archived'
            }, room='all_users')
            
            print(f"DEBUG: WebSocket events emitted for request_archived")
        except Exception as e:
            print(f"Error emitting WebSocket notification: {e}")
    
    flash(f'Payment request #{request_id} has been archived.', 'success')
    return redirect(url_for('it_dashboard'))

@app.route('/bulk-delete-requests', methods=['POST'])
@login_required
@role_required('IT Staff', 'Department Manager')
def bulk_delete_requests():
    """Bulk archive payment requests (IT only) - soft delete"""
    # Restrict Department Managers to IT department only
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    
    request_ids = request.form.getlist('request_ids')
    
    if not request_ids:
        flash('No requests selected for archiving.', 'warning')
        return redirect(url_for('it_dashboard'))
    
    archived_count = 0
    already_archived_count = 0
    archived_requests = []  # Store archived requests for notifications
    
    for request_id in request_ids:
        try:
            req = PaymentRequest.query.get(int(request_id))
            if req:
                if req.is_archived:
                    already_archived_count += 1
                    continue
                
                # Archive the request (soft delete) - keep all files and related records
                req.is_archived = True
                req.archived_at = datetime.utcnow()
                req.archived_by = current_user.name
                req.archived_by_user_id = current_user.user_id
                
                # Store request info for notifications
                archived_requests.append({
                    'request_id': request_id,
                    'requestor_name': req.requestor_name
                })
                
                # Log the archival
                log_action(f"Bulk archived payment request #{request_id} - {req.request_type} - {req.purpose}")
                archived_count += 1
        except (ValueError, TypeError):
            continue
    
    db.session.commit()
    
    # Notify all IT department users about bulk archive
    if archived_count > 0:
        it_users = User.query.filter(
            db.or_(
                User.role == 'IT Staff',
                db.and_(User.role == 'Department Manager', User.department == 'IT')
            )
        ).all()
        
        # Exclude the user who archived (they already know)
        it_users_to_notify = [user for user in it_users if user.user_id != current_user.user_id]
        
        # Create notifications for each archived request
        for archived_req in archived_requests:
            notification_title = "Payment Request Archived"
            notification_message = f"Payment request #{archived_req['request_id']} submitted by {archived_req['requestor_name']} has been archived by {current_user.name}."
            
            for it_user in it_users_to_notify:
                create_notification(
                    user_id=it_user.user_id,
                    title=notification_title,
                    message=notification_message,
                    notification_type="request_archived",
                    request_id=archived_req['request_id']
                )
            
            # Emit real-time notification for each archived request
            if it_users_to_notify:
                try:
                    socketio.emit('new_notification', {
                        'title': notification_title,
                        'message': notification_message,
                        'type': 'request_archived',
                        'request_id': archived_req['request_id']
                    }, room='all_users')
                    
                    # Also emit a general update event to trigger notification count updates
                    socketio.emit('notification_update', {
                        'action': 'new_notification',
                        'type': 'request_archived'
                    }, room='all_users')
                except Exception as e:
                    print(f"Error emitting WebSocket notification: {e}")
    
    if already_archived_count > 0:
        flash(f'{archived_count} payment request(s) have been archived. {already_archived_count} were already archived.', 'success')
    else:
        flash(f'{archived_count} payment request(s) have been archived.', 'success')
    return redirect(url_for('it_dashboard'))

@app.route('/request/<int:request_id>/mark-installment-paid', methods=['POST'])
@login_required
@role_required('Finance Admin')
def mark_installment_paid_finance(request_id):
    """Mark a specific installment as paid (Finance Admin only)"""
    req = PaymentRequest.query.get_or_404(request_id)
    
    # Check if request is recurring
    if req.status != 'Recurring':
        flash('This endpoint is only for recurring payments.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    schedule_id = request.form.get('schedule_id')
    payment_date = request.form.get('payment_date')
    
    if not schedule_id or not payment_date:
        flash('Missing required parameters.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    try:
        # Get the schedule entry
        schedule_entry = RecurringPaymentSchedule.query.get(int(schedule_id))
        if not schedule_entry or schedule_entry.request_id != request_id:
            flash('Installment not found.', 'error')
            return redirect(url_for('view_request', request_id=request_id))
        
        # Mark as paid
        schedule_entry.is_paid = True
        schedule_entry.paid_date = datetime.strptime(payment_date, '%Y-%m-%d').date()
        
        # Create a paid notification
        create_notification(
            user_id=req.user_id,
            title="Installment Paid",
            message=f'Installment for {payment_date} has been marked as paid (Amount: {schedule_entry.amount} OMR)',
            notification_type="installment_paid",
            request_id=request_id
        )
        
        # Also notify Finance Admin and Finance Staff
        finance_users = User.query.filter(User.role.in_(['Finance Staff', 'Finance Admin'])).all()
        for user in finance_users:
            create_notification(
                user_id=user.user_id,
                title="Installment Paid",
                message=f'Installment for {payment_date} has been marked as paid (Amount: {schedule_entry.amount} OMR)',
                notification_type="installment_paid",
                request_id=request_id
            )
        
        db.session.commit()
        
        # Check if all installments are now paid and mark as completed if so
        check_recurring_payment_completion(request_id)
        
        log_action(f"Marked installment {schedule_id} as paid for request #{request_id}")
        flash(f'Installment for {payment_date} has been marked as paid.', 'success')
        
    except (ValueError, TypeError) as e:
        flash('Invalid date format.', 'error')
    except Exception as e:
        db.session.rollback()
        flash('Error marking installment as paid.', 'error')
    
    return redirect(url_for('view_request', request_id=request_id))

@app.route('/request/<int:request_id>/upload-installment-receipt', methods=['POST'])
@login_required
@role_required('Finance Admin')
def upload_installment_receipt(request_id):
    """Upload receipt for a specific installment (Finance Admin only)"""
    req = PaymentRequest.query.get_or_404(request_id)
    
    # Check if request is recurring
    if req.status != 'Recurring':
        flash('This endpoint is only for recurring payments.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    schedule_id = request.form.get('schedule_id')
    payment_date = request.form.get('payment_date')
    
    if not schedule_id or not payment_date:
        flash('Missing required parameters.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    # Handle single receipt upload for installment
    if 'receipt_file' not in request.files:
        flash('No receipt file uploaded.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    receipt_file = request.files['receipt_file']
    if not receipt_file or not receipt_file.filename:
        flash('No receipt file selected.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    allowed_extensions = {'pdf', 'jpg', 'jpeg', 'png', 'doc', 'docx', 'xls', 'xlsx'}
    
    # Validate file size (50MB max)
    max_file_size = app.config.get('MAX_FILE_SIZE', 50 * 1024 * 1024)
    if len(receipt_file.read()) > max_file_size:
        flash(f'File "{receipt_file.filename}" is too large. Maximum size is {max_file_size // (1024 * 1024)}MB.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    # Reset file pointer
    receipt_file.seek(0)
    
    # Validate file extension
    file_extension = receipt_file.filename.rsplit('.', 1)[1].lower() if '.' in receipt_file.filename else ''
    if file_extension not in allowed_extensions:
        flash(f'Invalid file type for "{receipt_file.filename}". Allowed types: PDF, JPG, PNG, DOC, DOCX, XLS, XLSX', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    try:
        # Get the schedule entry
        schedule_entry = RecurringPaymentSchedule.query.get(int(schedule_id))
        if not schedule_entry or schedule_entry.request_id != request_id:
            flash('Installment not found.', 'error')
            return redirect(url_for('view_request', request_id=request_id))
        
        # Handle single receipt upload for installment
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = secure_filename(receipt_file.filename)
        filename = f"installment_{schedule_id}_{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        receipt_file.save(filepath)
        
        # Store the file as receipt for this installment (only if it doesn't already have one)
        if schedule_entry.receipt_path:
            flash(f'This installment already has a receipt. Please delete the existing receipt first if you want to replace it.', 'error')
            return redirect(url_for('view_request', request_id=request_id))
        
        schedule_entry.receipt_path = filename
        db.session.commit()
        
        log_action(f"Uploaded receipt for installment {schedule_id} for request #{request_id}")
        flash(f'Receipt uploaded successfully for installment on {payment_date}.', 'success')
        
    except (ValueError, TypeError) as e:
        flash('Invalid parameters.', 'error')
    except Exception as e:
        flash('Error uploading receipt.', 'error')
    
    return redirect(url_for('view_request', request_id=request_id))


@app.route('/request/<int:request_id>/upload-installment-invoice', methods=['POST'])
@login_required
def upload_installment_invoice(request_id):
    """Upload invoice for a specific installment (Requestor only)"""
    req = PaymentRequest.query.get_or_404(request_id)
    
    # Check if request is recurring
    if req.status != 'Recurring':
        flash('This endpoint is only for recurring payments.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    # Check if user is the requestor
    if req.user_id != current_user.user_id:
        flash('You are not authorized to upload invoices for this request.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    schedule_id = request.form.get('schedule_id')
    payment_date = request.form.get('payment_date')
    
    if not schedule_id or not payment_date:
        flash('Missing required parameters.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    # Handle single invoice upload for installment
    if 'invoice_file' not in request.files:
        flash('No invoice file uploaded.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    invoice_file = request.files['invoice_file']
    if not invoice_file or not invoice_file.filename:
        flash('No invoice file selected.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    allowed_extensions = {'pdf', 'jpg', 'jpeg', 'png', 'doc', 'docx', 'xls', 'xlsx'}
    
    # Validate file size (50MB max)
    max_file_size = app.config.get('MAX_FILE_SIZE', 50 * 1024 * 1024)
    if len(invoice_file.read()) > max_file_size:
        flash(f'File "{invoice_file.filename}" is too large. Maximum size is {max_file_size // (1024 * 1024)}MB.', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    # Reset file pointer
    invoice_file.seek(0)
    
    # Validate file extension
    file_extension = invoice_file.filename.rsplit('.', 1)[1].lower() if '.' in invoice_file.filename else ''
    if file_extension not in allowed_extensions:
        flash(f'Invalid file type for "{invoice_file.filename}". Allowed types: PDF, JPG, PNG, DOC, DOCX, XLS, XLSX', 'error')
        return redirect(url_for('view_request', request_id=request_id))
    
    try:
        # Get the schedule entry
        schedule_entry = RecurringPaymentSchedule.query.get(int(schedule_id))
        if not schedule_entry or schedule_entry.request_id != request_id:
            flash('Installment not found.', 'error')
            return redirect(url_for('view_request', request_id=request_id))
        
        # Handle single invoice upload for installment
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = secure_filename(invoice_file.filename)
        filename = f"invoice_{schedule_id}_{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        invoice_file.save(filepath)
        
        # Store the file as invoice for this installment (only if it doesn't already have one)
        if schedule_entry.invoice_path:
            flash(f'This installment already has an invoice. Please delete the existing invoice first if you want to replace it.', 'error')
            return redirect(url_for('view_request', request_id=request_id))
        
        schedule_entry.invoice_path = filename
        db.session.commit()
        
        log_action(f"Uploaded invoice for installment {schedule_id} for request #{request_id}")
        flash(f'Invoice uploaded successfully for installment on {payment_date}.', 'success')
        
    except (ValueError, TypeError) as e:
        flash('Invalid parameters.', 'error')
    except Exception as e:
        flash('Error uploading invoice.', 'error')
    
    return redirect(url_for('view_request', request_id=request_id))

@app.route('/request/<int:request_id>/edit_installment', methods=['POST'])
@login_required
def edit_installment(request_id):
    """Edit installment details (Requestor only)"""
    try:
        # Get the request
        req = PaymentRequest.query.get_or_404(request_id)
        
        # Check if user is the requestor
        if req.user_id != current_user.user_id:
            return jsonify({'success': False, 'message': 'You can only edit your own request installments.'}), 403
        
        # Check if request is recurring
        if req.status != 'Recurring':
            return jsonify({'success': False, 'message': 'Can only edit installments for recurring payments.'}), 400
        
        # Get form data
        schedule_id = request.form.get('schedule_id')
        new_payment_date = request.form.get('payment_date')
        edit_reason = request.form.get('edit_reason', '').strip()
        
        if not all([schedule_id, new_payment_date, edit_reason]):
            return jsonify({'success': False, 'message': 'Missing required fields. Please provide a reason for the edit.'}), 400
        
        # Get the schedule entry
        schedule_entry = RecurringPaymentSchedule.query.get(int(schedule_id))
        if not schedule_entry or schedule_entry.request_id != request_id:
            return jsonify({'success': False, 'message': 'Installment not found.'}), 404
        
        # Check if installment is already paid
        if schedule_entry.is_paid:
            return jsonify({'success': False, 'message': 'Cannot edit paid installments.'}), 400
        
        # Validate date
        try:
            payment_date = datetime.strptime(new_payment_date, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'message': 'Invalid date format.'}), 400
        
        # Store original values for logging and history
        original_date = schedule_entry.payment_date
        original_amount = schedule_entry.amount
        
        # Update only the payment date (amount cannot be changed)
        schedule_entry.payment_date = payment_date
        schedule_entry.has_been_edited = True
        
        # Create edit history record
        edit_history = InstallmentEditHistory(
            schedule_id=schedule_entry.schedule_id,
            request_id=request_id,
            edited_by_user_id=current_user.user_id,
            old_payment_date=original_date,
            new_payment_date=payment_date,
            old_amount=original_amount,
            new_amount=original_amount,  # Amount doesn't change
            edit_reason=edit_reason
        )
        db.session.add(edit_history)
        
        # Log the changes
        log_action(f"Edited installment {schedule_id} for request #{request_id}: Payment date changed from {original_date} to {payment_date}")
        
        # Clean up any existing notifications for the old date
        # This prevents notifications from being sent on the old date
        old_date_notifications = Notification.query.filter(
            Notification.request_id == request_id,
            Notification.notification_type == 'recurring_due',
            Notification.message.contains(str(original_date))
        ).all()
        
        for notification in old_date_notifications:
            db.session.delete(notification)
        
        # Send notification to Finance Admin about the edit
        finance_admin_users = User.query.filter_by(role='Finance Admin').all()
        for admin in finance_admin_users:
            create_notification(
                user_id=admin.user_id,
                title="Installment Date Edited",
                message=f"Recurring request #{request_id} installment payment date has been edited from {original_date} to {payment_date} by {current_user.name}",
                notification_type='installment_edited',
                request_id=request_id
            )
        
        # If the new date is today, create immediate notifications
        if payment_date == date.today():
            # Create notifications for Finance Admin and Finance Staff about the payment due today
            finance_users = User.query.filter(User.role.in_(['Finance Admin', 'Finance Staff'])).all()
            for user in finance_users:
                create_notification(
                    user_id=user.user_id,
                    title="Recurring Payment Due",
                    message=f'Recurring payment due today for {req.request_type} - {req.purpose} (Amount: {schedule_entry.amount} OMR) - Date was recently edited',
                    notification_type='recurring_due',
                    request_id=request_id
                )
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Installment updated successfully.'})
        
    except Exception as e:
        db.session.rollback()
        print(f"Error editing installment: {str(e)}")
        return jsonify({'success': False, 'message': 'Error updating installment.'}), 500


@app.route('/api/installment/<int:schedule_id>/edit-history', methods=['GET'])
@login_required
def get_installment_edit_history(schedule_id):
    """Get edit history for a specific installment"""
    try:
        # Get the schedule entry
        schedule_entry = RecurringPaymentSchedule.query.get(schedule_id)
        if not schedule_entry:
            return jsonify({'success': False, 'message': 'Installment not found.'}), 404
        
        # Check if user has permission to view this installment
        if schedule_entry.request.user_id != current_user.user_id and current_user.role not in ['Finance Admin', 'Finance Staff', 'GM', 'IT Staff', 'Operation Manager']:
            return jsonify({'success': False, 'message': 'You do not have permission to view this installment.'}), 403
        
        # Get edit history for this installment
        edit_history = InstallmentEditHistory.query.filter_by(
            schedule_id=schedule_id
        ).order_by(InstallmentEditHistory.created_at.desc()).all()
        
        # Convert to dictionary format
        history_data = [edit.to_dict() for edit in edit_history]
        
        return jsonify({
            'success': True, 
            'edit_history': history_data,
            'installment_info': {
                'schedule_id': schedule_entry.schedule_id,
                'payment_date': schedule_entry.payment_date.strftime('%Y-%m-%d'),
                'amount': float(schedule_entry.amount),
                'has_been_edited': schedule_entry.has_been_edited
            }
        })
        
    except Exception as e:
        print(f"Error fetching installment edit history: {str(e)}")
        return jsonify({'success': False, 'message': 'Error fetching edit history.'}), 500


# ==================== REPORTS ROUTES ====================

@app.route('/reports')
@login_required
@role_required('Finance Admin', 'Finance Staff', 'GM', 'CEO', 'IT Staff', 'Department Manager', 'Operation Manager')
def reports():
    """View reports page"""
    # Get filter parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    department_filter = request.args.get('department', '')
    request_type_filter = request.args.get('request_type', '')
    company_filter = request.args.get('company', '')
    branch_filter = request.args.get('branch', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    status_filter = request.args.get('status', '')
    payment_type_filter = request.args.get('payment_type', '')
    payment_type_filter = request.args.get('payment_type', '')
    payment_type_filter = request.args.get('payment_type', '')
    payment_type_filter = request.args.get('payment_type', '')
    payment_type_filter = request.args.get('payment_type', '')
    payment_type_filter = request.args.get('payment_type', '')  # 'One-Time' or 'Recurring'
    
    # Validate per_page to prevent abuse
    if per_page not in [10, 20, 50, 100]:
        per_page = 10
    
    # Build query - show ALL statuses by default, but exclude archived requests
    query = PaymentRequest.query.filter(PaymentRequest.is_archived == False)
    
    # Filter for Department Managers based on their department
    if current_user.role == 'Department Manager':
        if current_user.department == 'IT':
            # IT Department Manager can see ALL requests from ALL departments (all statuses)
            # No filtering needed - they see everything
            pass
        elif current_user.department == 'Auditing':
            # Auditing Department Manager can see:
            # 1. ALL requests from Auditing department (all statuses)
            # 2. Completed and Recurring requests from OTHER departments
            query = query.filter(
                db.or_(
                    PaymentRequest.department == 'Auditing',
                    PaymentRequest.status.in_(['Completed', 'Recurring'])
                )
            )
        else:
            # Other Department Managers (non-IT, non-Auditing) can ONLY see their own department's requests
            query = query.filter(PaymentRequest.department == current_user.department)
    
    if status_filter:
        if status_filter == 'All Pending':
            # Show both pending statuses
            query = query.filter(PaymentRequest.status.in_(['Pending Manager Approval', 'Pending Finance Approval']))
        else:
            query = query.filter_by(status=status_filter)
    if department_filter:
        query = query.filter_by(department=department_filter)
    if request_type_filter:
        # Special handling for "Others" to match both "Others" and "Others:..."
        if request_type_filter == 'Others':
            query = query.filter(PaymentRequest.request_type.like('Others%'))
        else:
            query = query.filter_by(request_type=request_type_filter)
    if company_filter:
        # Filter by person_company field only (company_name is no longer used)
        query = query.filter(PaymentRequest.person_company.ilike(f'%{company_filter}%'))
    if branch_filter:
        # Alias-aware branch filtering: include canonical name and any aliases
        selected_branch = Branch.query.filter_by(name=branch_filter).first()
        if selected_branch:
            alias_names = [a.alias_name for a in getattr(selected_branch, 'aliases', [])]
            names = [selected_branch.name] + alias_names
            query = query.filter(PaymentRequest.branch_name.in_(names))
        else:
            query = query.filter(PaymentRequest.branch_name == branch_filter)
    # Payment type filter
    if payment_type_filter:
        if payment_type_filter == 'Recurring':
            query = query.filter(PaymentRequest.recurring == 'Recurring')
        elif payment_type_filter == 'Scheduled One-Time':
            query = query.filter(
                db.or_(
                    PaymentRequest.recurring == 'Scheduled One-Time',
                    PaymentRequest.payment_date.isnot(None)
                )
            ).filter(PaymentRequest.recurring != 'Recurring')
        elif payment_type_filter == 'One-Time':
            query = query.filter(
                db.or_(
                    PaymentRequest.recurring == None,
                    PaymentRequest.recurring == '',
                    PaymentRequest.recurring == 'One-Time'
                )
            ).filter(PaymentRequest.payment_date.is_(None))
    
    # Date filtering - use submission date (date field) which exists for all requests
    if date_from:
        query = query.filter(PaymentRequest.date >= datetime.strptime(date_from, '%Y-%m-%d').date())
    if date_to:
        query = query.filter(PaymentRequest.date <= datetime.strptime(date_to, '%Y-%m-%d').date())
    # Sort by status priority then by date (Completed by completion_date, others by created_at)
    # Get all filtered requests for stats calculation (before pagination)
    all_filtered_requests = query.order_by(
        get_status_priority_order(),
        get_all_tab_datetime_order()
    ).all()
    
    # Calculate stats from all filtered requests
    total_requests = len(all_filtered_requests)
    completed_count = len([r for r in all_filtered_requests if r.status == 'Completed'])
    pending_count = len([r for r in all_filtered_requests if r.status in ['Pending Manager Approval', 'Pending Finance Approval']])
    
    # Calculate total amount - handle IT Staff/Department Manager special case
    if current_user.role == 'IT Staff' or (current_user.role == 'Department Manager' and current_user.department == 'IT'):
        it_requests = [r for r in all_filtered_requests if r.department == 'IT']
        total_amount = sum(float(r.amount) for r in it_requests) if it_requests else 0
        it_amount = total_amount
    else:
        total_amount = sum(float(r.amount) for r in all_filtered_requests)
        it_amount = None
    
    # Paginate the query for display with the same ordering
    pagination = query.order_by(
        get_status_priority_order(),
        get_all_tab_datetime_order()
    ).paginate(
        page=page, per_page=per_page, error_out=False
    )
    requests = pagination.items
    
    # Get unique departments for filter (exclude archived)
    # For non-IT, non-Auditing Department Managers, only show their own department
    if current_user.role == 'Department Manager' and current_user.department not in ['IT', 'Auditing']:
        departments = [current_user.department] if current_user.department else []
    else:
        # For IT, Auditing Department Managers, and all other users, show all departments (exclude archived)
        departments = db.session.query(PaymentRequest.department).filter(
            PaymentRequest.is_archived == False
        ).distinct().all()
        departments = [d[0] for d in departments]
    
    # Get unique companies for filter (only person_company field since company_name is no longer used)
    # Show companies from all statuses, or filter by status if status_filter is provided (exclude archived)
    companies_query = db.session.query(PaymentRequest.person_company).filter(
        PaymentRequest.is_archived == False,
        PaymentRequest.person_company.isnot(None),
        PaymentRequest.person_company != ''
    )
    if status_filter:
        if status_filter == 'All Pending':
            companies_query = companies_query.filter(PaymentRequest.status.in_(['Pending Manager Approval', 'Pending Finance Approval']))
        else:
            companies_query = companies_query.filter(PaymentRequest.status == status_filter)
    companies = companies_query.distinct().all()
    companies = [c[0] for c in companies if c[0]]
    companies.sort()  # Sort alphabetically
    
    # Get unique branches for filter
    branches = Branch.query.filter_by(is_active=True).order_by(Branch.name).all()
    
    # Get request types based on selected department
    if department_filter:
        # If department is selected, show only request types for that department
        request_types = db.session.query(RequestType.name).filter(
            RequestType.department == department_filter,
            RequestType.is_active == True
        ).order_by(RequestType.name).all()
        request_types = [rt[0] for rt in request_types]
    else:
        # If no department selected, show all unique request types (remove duplicates)
        request_types = db.session.query(RequestType.name).filter(
            RequestType.is_active == True
        ).distinct().order_by(RequestType.name).all()
        request_types = [rt[0] for rt in request_types]
    
    return render_template('reports.html', 
                         requests=requests, 
                         pagination=pagination,
                         departments=departments,
                         companies=companies,
                         branches=branches,
                         request_types=request_types,
                         company_filter=company_filter,
                         branch_filter=branch_filter,
                         status_filter=status_filter,
                         payment_type_filter=payment_type_filter,
                         department_filter=department_filter,
                         request_type_filter=request_type_filter,
                         date_from=date_from,
                         date_to=date_to,
                         total_requests=total_requests,
                         completed_count=completed_count,
                         pending_count=pending_count,
                         total_amount=total_amount,
                         it_amount=it_amount,
                         user=current_user)


@app.route('/cheque-register')
@login_required
def cheque_register():
    """View cheque register page"""
    # Get filters from query parameters
    book_filter = request.args.get('book', '')
    status_filter = request.args.get('status', '')
    
    # Get all unique book numbers for the filter dropdown
    book_numbers = db.session.query(ChequeBook.book_no).distinct().order_by(ChequeBook.book_no).all()
    book_numbers = [book[0] for book in book_numbers]
    
    # Build query
    query = db.session.query(ChequeSerial).join(ChequeBook)
    
    # Apply book filter if provided
    if book_filter:
        try:
            book_no = int(book_filter)
            query = query.filter(ChequeBook.book_no == book_no)
        except ValueError:
            pass  # Invalid book number, ignore filter
    
    # Apply status filter if provided
    if status_filter:
        valid_statuses = ['Available', 'Reserved', 'Used', 'Cancelled']
        if status_filter in valid_statuses:
            query = query.filter(ChequeSerial.status == status_filter)
    
    # Fetch all cheque serials with their book information, ordered by book_no and serial_no
    cheque_serials = query.order_by(
        ChequeBook.book_no, ChequeSerial.serial_no
    ).all()
    
    return render_template('cheque_register.html', 
                          cheque_serials=cheque_serials, 
                          book_numbers=book_numbers,
                          book_filter=book_filter,
                          status_filter=status_filter)


@app.route('/cheque-register/reserve', methods=['GET', 'POST'])
@login_required
def reserve_cheque():
    """Reserve cheque serial numbers"""
    if request.method == 'GET':
        # Get book filter from query parameters
        book_filter = request.args.get('book', '')
        
        # Get all unique book numbers for the filter dropdown
        book_numbers = db.session.query(ChequeBook.book_no).distinct().order_by(ChequeBook.book_no).all()
        book_numbers = [book[0] for book in book_numbers]
        
        # Build query for available serials
        query = db.session.query(ChequeSerial).join(ChequeBook).filter(
            ChequeSerial.status == 'Available'
        )
        
        # Apply book filter if provided
        if book_filter:
            try:
                book_no = int(book_filter)
                query = query.filter(ChequeBook.book_no == book_no)
            except ValueError:
                pass  # Invalid book number, ignore filter
        
        # Fetch all available serial numbers
        available_serials = query.order_by(
            ChequeBook.book_no, ChequeSerial.serial_no
        ).all()
        
        return render_template('reserve_cheque.html', 
                             available_serials=available_serials,
                             book_numbers=book_numbers,
                             book_filter=book_filter)
    
    elif request.method == 'POST':
        # Handle reservation
        try:
            data = request.get_json()
            serial_ids = data.get('serial_ids', [])
            
            if not serial_ids:
                return jsonify({'success': False, 'error': 'No serial numbers selected'}), 400
            
            # Validate that all serials exist and are available
            serials = ChequeSerial.query.filter(
                ChequeSerial.id.in_(serial_ids),
                ChequeSerial.status == 'Available'
            ).all()
            
            if len(serials) != len(serial_ids):
                return jsonify({'success': False, 'error': 'Some selected serial numbers are not available'}), 400
            
            # Update status to Reserved
            reserved_count = 0
            for serial in serials:
                serial.status = 'Reserved'
                reserved_count += 1
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'reserved_count': reserved_count,
                'message': f'Successfully reserved {reserved_count} serial number(s)'
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/cheque-register/new-book', methods=['GET', 'POST'])
@login_required
def new_book():
    """Add new cheque book"""
    if request.method == 'POST':
        book_no = request.form.get('book_no')
        start_serial_no = request.form.get('start_serial_no')
        last_serial_no = request.form.get('last_serial_no')
        
        # Validation
        try:
            book_no = int(book_no)
            start_serial_no = int(start_serial_no)
            last_serial_no = int(last_serial_no)
            
            if start_serial_no > last_serial_no:
                flash('Starting serial number must be less than or equal to last serial number', 'error')
                return render_template('new_book.html')
            
            # Check if book number already exists
            existing_book = ChequeBook.query.filter_by(book_no=book_no).first()
            if existing_book:
                flash(f'Book number {book_no} already exists. Please use a different book number.', 'error')
                return render_template('new_book.html')
            
            # Create the cheque book
            cheque_book = ChequeBook(
                book_no=book_no,
                start_serial_no=start_serial_no,
                last_serial_no=last_serial_no,
                created_by_user_id=current_user.user_id
            )
            db.session.add(cheque_book)
            db.session.flush()  # Get the book ID
            
            # Generate all serial numbers from start to last (inclusive)
            serial_numbers = []
            for serial_no in range(start_serial_no, last_serial_no + 1):
                cheque_serial = ChequeSerial(
                    book_id=cheque_book.id,
                    serial_no=serial_no,
                    status='Available'
                )
                serial_numbers.append(cheque_serial)
            
            db.session.add_all(serial_numbers)
            db.session.commit()
            
            flash(f'Book {book_no} created successfully with {len(serial_numbers)} serial numbers ({start_serial_no} to {last_serial_no})', 'success')
            return redirect(url_for('cheque_register'))
        except ValueError:
            flash('Please enter valid numbers for all fields', 'error')
            return render_template('new_book.html')
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating book: {str(e)}', 'error')
            return render_template('new_book.html')
    
    return render_template('new_book.html')


@app.route('/api/request-types')
@login_required
def api_request_types():
    """API endpoint to get request types by department"""
    department = request.args.get('department', '')
    
    if department:
        # Get request types for specific department
        request_types = db.session.query(RequestType.name).filter(
            RequestType.department == department,
            RequestType.is_active == True
        ).order_by(RequestType.name).all()
        request_types = [rt[0] for rt in request_types]
    else:
        # Get all unique request types (remove duplicates)
        request_types = db.session.query(RequestType.name).filter(
            RequestType.is_active == True
        ).distinct().order_by(RequestType.name).all()
        request_types = [rt[0] for rt in request_types]
    
    return jsonify({
        'request_types': request_types
    })


@app.route('/reports/export/excel')
@login_required
@role_required('Finance Admin', 'Finance Staff', 'GM', 'CEO', 'IT Staff', 'Department Manager', 'Operation Manager')
def export_reports_excel():
    """Export filtered reports to an Excel file with frozen columns"""
    # Lazy imports to avoid hard dependency during app startup
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash('Excel export requires openpyxl. Install with: pip install openpyxl', 'warning')
        return redirect(url_for('reports', **request.args))
    except Exception as e:
        flash(f'Error importing Excel library: {str(e)}', 'error')
        return redirect(url_for('reports', **request.args))

    # Reuse the same filters as the reports() view
    department_filter = request.args.get('department', '')
    request_type_filter = request.args.get('request_type', '')
    company_filter = request.args.get('company', '')
    branch_filter = request.args.get('branch', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    status_filter = request.args.get('status', '')
    payment_type_filter = request.args.get('payment_type', '')

    # Show ALL statuses by default (consistent with reports() view), but exclude archived requests
    query = PaymentRequest.query.filter(PaymentRequest.is_archived == False)
    
    # Filter for Department Managers based on their department (same logic as reports() view)
    if current_user.role == 'Department Manager':
        if current_user.department == 'IT':
            # IT Department Manager can see ALL requests from ALL departments (all statuses)
            # No filtering needed - they see everything
            pass
        elif current_user.department == 'Auditing':
            # Auditing Department Manager can see:
            # 1. ALL requests from Auditing department (all statuses)
            # 2. Completed and Recurring requests from OTHER departments
            query = query.filter(
                db.or_(
                    PaymentRequest.department == 'Auditing',
                    PaymentRequest.status.in_(['Completed', 'Recurring'])
                )
            )
        else:
            # Other Department Managers (non-IT, non-Auditing) can ONLY see their own department's requests
            query = query.filter(PaymentRequest.department == current_user.department)
    
    if status_filter:
        if status_filter == 'All Pending':
            # Show both pending statuses
            query = query.filter(PaymentRequest.status.in_(['Pending Manager Approval', 'Pending Finance Approval']))
        else:
            query = query.filter_by(status=status_filter)
    if department_filter:
        query = query.filter_by(department=department_filter)
    if request_type_filter:
        # Special handling for "Others" to match both "Others" and "Others:..."
        if request_type_filter == 'Others':
            query = query.filter(PaymentRequest.request_type.like('Others%'))
        else:
            query = query.filter_by(request_type=request_type_filter)
    if company_filter:
        # Filter by person_company field only (company_name is no longer used)
        query = query.filter(PaymentRequest.person_company.ilike(f'%{company_filter}%'))
    if branch_filter:
        # Alias-aware branch filtering for Excel export
        selected_branch = Branch.query.filter_by(name=branch_filter).first()
        if selected_branch:
            alias_names = [a.alias_name for a in getattr(selected_branch, 'aliases', [])]
            names = [selected_branch.name] + alias_names
            query = query.filter(PaymentRequest.branch_name.in_(names))
        else:
            query = query.filter_by(branch_name=branch_filter)
    if payment_type_filter:
        if payment_type_filter == 'Recurring':
            query = query.filter(PaymentRequest.recurring == 'Recurring')
        elif payment_type_filter == 'One-Time':
            query = query.filter(
                db.or_(
                    PaymentRequest.recurring == None,
                    PaymentRequest.recurring == '',
                    PaymentRequest.recurring == 'One-Time'
                )
            )
    if date_from:
        query = query.filter(PaymentRequest.date >= datetime.strptime(date_from, '%Y-%m-%d').date())
    if date_to:
        query = query.filter(PaymentRequest.date <= datetime.strptime(date_to, '%Y-%m-%d').date())

    result_requests = query.order_by(PaymentRequest.date.desc()).all()

    # Helper function to convert to float
    def to_float(value):
        try:
            return float(value)
        except Exception:
            return 0.0

    # Sum all amounts (requests may be in any status)
    total_amount = sum(to_float(r.amount) for r in result_requests)

    try:
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payment Reports"

        # Add report header information
        ws['A1'] = 'Payment Reports'
        ws['A1'].font = Font(size=16, bold=True)
        
        generation_date = datetime.now().strftime('%Y-%m-%d %H:%M')
        ws['A2'] = f"Report Generated: {generation_date}"
        
        filters_line = f"Dept: {department_filter or 'All'} | Type: {request_type_filter or 'All'} | Branch: {branch_filter or 'All'} | Payment: {payment_type_filter or 'All'}"
        ws['A3'] = filters_line
        
        # Date scope
        if date_from and date_to:
            ws['A4'] = f"Date Range: {date_from} to {date_to}"
        elif date_from:
            ws['A4'] = f"Date From: {date_from} (no end date)"
        elif date_to:
            ws['A4'] = f"Date To: {date_to} (no start date)"
        else:
            ws['A4'] = "Date Range: All dates (no filter applied)"
        
        ws['A5'] = f"Total Amount: OMR {total_amount:.3f}"
        ws['A5'].font = Font(size=12, bold=True)

        # Add headers starting from row 7
        headers = ['ID', 'Type', 'Requestor', 'Department', 'Payment', 'Scheduled', 'Amount', 'Branch', 'Company', 'Submitted', 'Approved', 'Approver', 'Manager Duration', 'Finance Duration']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Helper function to format duration
        def format_duration(seconds):
            """Format duration in seconds to readable format (H:MM:SS)"""
            if not seconds:
                return ''
            
            hours = seconds // 3600
            minutes = (seconds % 3600) // 60
            secs = seconds % 60
            
            return f"{hours}:{minutes:02d}:{secs:02d}"

        # Add data rows
        for row_idx, r in enumerate(result_requests, 8):
            # For Approved column: show approval_date (matching reports page)
            approved_date = r.approval_date.strftime('%Y-%m-%d') if getattr(r, 'approval_date', None) else ''
            
            # For Person/Company column, show person_company field only
            company_display = r.person_company
            
            # Format duration columns
            manager_duration = format_duration(r.manager_approval_duration_minutes)
            finance_duration = format_duration(r.finance_approval_duration_minutes)
            
            scheduled_date = ''
            if (not getattr(r, 'recurring', None) or getattr(r, 'recurring', None) != 'Recurring') and getattr(r, 'payment_date', None):
                try:
                    scheduled_date = r.payment_date.strftime('%Y-%m-%d')
                except Exception:
                    scheduled_date = str(r.payment_date)

            row_data = [
                f"#{r.request_id}",
                str(r.request_type or ''),
                str(r.requestor_name or ''),
                str(r.department or ''),
                str(r.recurring or 'One-Time'),
                scheduled_date,
                f"OMR {to_float(r.amount):.3f}",
                str(r.branch_name or ''),
                str(company_display or ''),
                r.date.strftime('%Y-%m-%d') if getattr(r, 'date', None) else '',
                approved_date,
                str(r.approver or ''),
                manager_duration,
                finance_duration
            ]
            
            for col, data in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col, value=data)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Special handling for specific columns
            if column_letter == 'A':  # ID column - make it narrower
                adjusted_width = min(max_length + 2, 12)  # Cap ID column at 12 characters
            elif column_letter == 'D':  # Department column - make it wider
                adjusted_width = min(max_length + 2, 25)  # Cap Department column at 25 characters
            elif column_letter in ['L', 'M']:  # Duration columns - make them narrower
                adjusted_width = min(max_length + 2, 15)  # Cap duration columns at 15 characters
            else:
                adjusted_width = min(max_length + 2, 50)  # Cap other columns at 50 characters
            
            ws.column_dimensions[column_letter].width = adjusted_width

        # No frozen panes - all columns can be scrolled freely

        # Save to BytesIO
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        from flask import make_response
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ts = datetime.now().strftime('%Y-%m-%d_%H-%M')
        response.headers['Content-Disposition'] = f'attachment; filename=reports_{ts}.xlsx'
        response.headers['Content-Length'] = str(len(buffer.getvalue()))
        return response
        
    except Exception as e:
        flash(f'Error generating Excel: {str(e)}', 'error')
        return redirect(url_for('reports', **request.args))


@app.route('/reports/export/pdf')
@login_required
@role_required('Finance Admin', 'Finance Staff', 'GM', 'CEO', 'IT Staff', 'Department Manager', 'Operation Manager')
def export_reports_pdf():
    """Export filtered reports to a PDF including total amount and full list"""
    # Lazy imports to avoid hard dependency during app startup
    import io
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import mm
        from reportlab.lib.utils import simpleSplit
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        flash('PDF export requires reportlab. Install with: pip install reportlab', 'warning')
        return redirect(url_for('reports', **request.args))
    except Exception as e:
        flash(f'Error importing PDF library: {str(e)}', 'error')
        return redirect(url_for('reports', **request.args))

    # Optional Arabic shaping dependencies (used if available)
    try:
        import os
        import unicodedata
        import re
        import arabic_reshaper  # pip install arabic-reshaper
        from bidi.algorithm import get_display  # pip install python-bidi
    except Exception:
        arabic_reshaper = None
        get_display = None

    # Reuse the same filters as the reports() view
    department_filter = request.args.get('department', '')
    request_type_filter = request.args.get('request_type', '')
    company_filter = request.args.get('company', '')
    branch_filter = request.args.get('branch', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    status_filter = request.args.get('status', '')
    payment_type_filter = request.args.get('payment_type', '')

    # Show ALL statuses by default (consistent with reports() view), but exclude archived requests
    query = PaymentRequest.query.filter(PaymentRequest.is_archived == False)
    
    # Filter for Department Managers based on their department (same logic as reports() view)
    if current_user.role == 'Department Manager':
        if current_user.department == 'IT':
            # IT Department Manager can see ALL requests from ALL departments (all statuses)
            # No filtering needed - they see everything
            pass
        elif current_user.department == 'Auditing':
            # Auditing Department Manager can see:
            # 1. ALL requests from Auditing department (all statuses)
            # 2. Completed and Recurring requests from OTHER departments
            query = query.filter(
                db.or_(
                    PaymentRequest.department == 'Auditing',
                    PaymentRequest.status.in_(['Completed', 'Recurring'])
                )
            )
        else:
            # Other Department Managers (non-IT, non-Auditing) can ONLY see their own department's requests
            query = query.filter(PaymentRequest.department == current_user.department)
    
    if status_filter:
        if status_filter == 'All Pending':
            # Show both pending statuses
            query = query.filter(PaymentRequest.status.in_(['Pending Manager Approval', 'Pending Finance Approval']))
        else:
            query = query.filter_by(status=status_filter)
    if department_filter:
        query = query.filter_by(department=department_filter)
    if request_type_filter:
        # Special handling for "Others" to match both "Others" and "Others:..."
        if request_type_filter == 'Others':
            query = query.filter(PaymentRequest.request_type.like('Others%'))
        else:
            query = query.filter_by(request_type=request_type_filter)
    if company_filter:
        # Filter by person_company field only (company_name is no longer used)
        query = query.filter(PaymentRequest.person_company.ilike(f'%{company_filter}%'))
    if branch_filter:
        # Alias-aware branch filtering for PDF export
        selected_branch = Branch.query.filter_by(name=branch_filter).first()
        if selected_branch:
            alias_names = [a.alias_name for a in getattr(selected_branch, 'aliases', [])]
            names = [selected_branch.name] + alias_names
            query = query.filter(PaymentRequest.branch_name.in_(names))
        else:
            query = query.filter(PaymentRequest.branch_name == branch_filter)

    # Date filtering - use submission date (date field) which exists for all requests
    if date_from:
        query = query.filter(PaymentRequest.date >= datetime.strptime(date_from, '%Y-%m-%d').date())
    if date_to:
        query = query.filter(PaymentRequest.date <= datetime.strptime(date_to, '%Y-%m-%d').date())
    result_requests = query.order_by(PaymentRequest.date.desc()).all()

    # Helper function to convert to float
    def to_float(value):
        try:
            return float(value)
        except Exception:
            return 0.0

    # Sum all amounts (requests may be in any status)
    total_amount = sum(to_float(r.amount) for r in result_requests)

    try:
        # Build PDF in landscape orientation
        buffer = io.BytesIO()
        from reportlab.lib.pagesizes import landscape
        c = canvas.Canvas(buffer, pagesize=landscape(A4))
        width, height = landscape(A4)

        # Register and select a font that supports Arabic if available
        arabic_font_name = None
        try:
            # Candidate fonts: prefer bundled fonts, then system fonts (Windows/Linux/macOS)
            font_candidates = [
                # Project-bundled fonts
                os.path.join(app.root_path, 'static', 'fonts', 'Amiri-Regular.ttf'),
                os.path.join(app.root_path, 'static', 'fonts', 'NotoNaskhArabic-Regular.ttf'),
                os.path.join(app.root_path, 'static', 'fonts', 'NotoSansArabic-Regular.ttf'),
                os.path.join(app.root_path, 'static', 'fonts', 'Cairo-Regular.ttf'),
                os.path.join(app.root_path, 'static', 'fonts', 'Tahoma.ttf'),
                os.path.join(app.root_path, 'static', 'fonts', 'Arial.ttf'),
                # Windows common fonts
                r'C:\\Windows\\Fonts\\trado.ttf',            # Traditional Arabic
                r'C:\\Windows\\Fonts\\Tahoma.ttf',
                r'C:\\Windows\\Fonts\\arial.ttf',
                r'C:\\Windows\\Fonts\\arialuni.ttf',        # Arial Unicode MS (if present)
                r'C:\\Windows\\Fonts\\times.ttf',
                r'C:\\Windows\\Fonts\\segoeui.ttf',
                r'C:\\Windows\\Fonts\\segoeuib.ttf',
                # Linux
                '/usr/share/fonts/truetype/noto/NotoNaskhArabic-Regular.ttf',
                '/usr/share/fonts/truetype/noto/NotoSansArabic-Regular.ttf',
                '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                # macOS
                '/Library/Fonts/Arial Unicode.ttf',
                '/Library/Fonts/Tahoma.ttf',
                '/System/Library/Fonts/Supplemental/Times New Roman.ttf',
            ]
            for font_path in font_candidates:
                if os.path.exists(font_path):
                    try:
                        pdfmetrics.registerFont(TTFont('ArabicFont', font_path))
                        arabic_font_name = 'ArabicFont'
                        break
                    except Exception:
                        continue
        except Exception:
            # If anything goes wrong, we will fall back to built-in fonts
            arabic_font_name = None

        # Use Arabic-capable font for all body text if available
        body_font = arabic_font_name or 'Helvetica'

        # Helpers for Arabic detection and shaping
        def contains_arabic(text):
            if not text:
                return False
            for ch in str(text):
                code = ord(ch)
                if (
                    0x0600 <= code <= 0x06FF  # Arabic
                    or 0x0750 <= code <= 0x077F  # Arabic Supplement
                    or 0x08A0 <= code <= 0x08FF  # Arabic Extended-A
                    or 0xFB50 <= code <= 0xFDFF  # Arabic Presentation Forms-A
                    or 0xFE70 <= code <= 0xFEFF  # Arabic Presentation Forms-B
                ):
                    return True
            return False

        def prepare_text(value: object) -> str:
            s = '' if value is None else str(value)
            if arabic_reshaper and get_display and contains_arabic(s):
                try:
                    s = get_display(arabic_reshaper.reshape(s))
                except Exception:
                    # If shaping fails, return original text
                    pass
            return s

        # Margins (reduced for more space)
        left = 8 * mm
        right = width - 8 * mm
        top = height - 8 * mm
        y = top

        # Header
        c.setFont('Helvetica-Bold', 14)
        c.drawString(left, y, 'Payment Reports')
        c.setFont(body_font, 10)
        y -= 14
        
        # Report generation date
        generation_date = datetime.now().strftime('%Y-%m-%d %H:%M')
        c.drawString(left, y, prepare_text(f"Report Generated: {generation_date}"))
        y -= 12
        
        # Filters
        filters_line = f"Dept: {department_filter or 'All'} | Type: {request_type_filter or 'All'} | Branch: {branch_filter or 'All'} | Payment: {payment_type_filter or 'All'}"
        c.drawString(left, y, prepare_text(filters_line))
        y -= 12
        
        # Date scope
        if date_from and date_to:
            c.drawString(left, y, prepare_text(f"Date Range: {date_from} to {date_to}"))
        elif date_from:
            c.drawString(left, y, prepare_text(f"Date From: {date_from} (no end date)"))
        elif date_to:
            c.drawString(left, y, prepare_text(f"Date To: {date_to} (no start date)"))
        else:
            c.drawString(left, y, prepare_text("Date Range: All dates (no filter applied)"))
        y -= 12

        # Total amount
        c.setFont('Helvetica-Bold', 11)
        c.drawString(left, y, f"Total Amount: OMR {total_amount:.3f}")
        y -= 18

        # Helper function to wrap text manually
        def wrap_text(text, max_width, font_name=body_font, font_size=9):
            """Wrap text to fit within specified width using actual font metrics."""
            s = '' if text is None else str(text)
            s = prepare_text(s)
            if not s:
                return ['']

            words = s.split()
            lines = []
            current = ''

            def string_width(t: str) -> float:
                try:
                    return pdfmetrics.stringWidth(t, font_name, font_size)
                except Exception:
                    # Fallback estimate if metrics are unavailable
                    return len(t) * font_size * 0.5

            # Use a slightly smaller width to keep text away from column border
            effective_width = max(0, max_width - (1.5 * mm))
            for word in words:
                candidate = (current + ' ' + word).strip()
                if current and string_width(candidate) > effective_width:
                    lines.append(current)
                    current = word
                else:
                    current = candidate

            if current:
                lines.append(current)

            # Hard-break extremely long tokens that exceed width
            fixed_lines = []
            for line in lines:
                if string_width(line) <= effective_width:
                    fixed_lines.append(line)
                    continue
                # Break by characters
                buf = ''
                for ch in line:
                    if string_width(buf + ch) > effective_width and buf:
                        fixed_lines.append(buf)
                        buf = ch
                    else:
                        buf += ch
                if buf:
                    fixed_lines.append(buf)
            return fixed_lines or ['']
        
        # Table header
        c.setFont(body_font if arabic_font_name else 'Helvetica-Bold', 9)
        headers = ['ID', 'Type', 'Requestor', 'Department', 'Payment', 'Scheduled', 'Amount', 'Branch', 'Company', 'Submitted', 'Approved', 'Approver']
        # Column widths optimized for landscape A4 with proper spacing
        # [ID, Type, Requestor, Department, Payment, Scheduled, Amount, Branch, Company, Submitted, Approved, Approver]
        col_widths = [14*mm, 30*mm, 22*mm, 18*mm, 16*mm, 18*mm, 18*mm, 26*mm, 26*mm, 16*mm, 16*mm, 28*mm]
        # Calculate column positions based on widths to prevent overlapping
        column_gap = 4 * mm  # slightly larger inter-column gap
        col_x = [left]
        for i in range(1, len(col_widths)):
            col_x.append(col_x[i-1] + col_widths[i-1] + column_gap)
        for hx, text in zip(col_x, headers):
            c.drawString(hx, y, prepare_text(text))
        y -= 10
        c.line(left, y, right, y)
        y -= 8

        # Rows
        c.setFont(body_font, 9)
        row_height = 10
        bottom_margin = 15 * mm
        for r in result_requests:
            # Calculate the maximum height needed for this row across all columns first
            max_height = 0
            wrapped_lines_per_col = []
            
            # For Person/Company column, show person_company field only
            company_display = r.person_company
            
            # Build Submitted (date + manager approval START time if available)
            submitted_date = r.date.strftime('%Y-%m-%d') if getattr(r, 'date', None) else ''
            manager_start_time_str = ''
            if getattr(r, 'manager_approval_start_time', None):
                local_start = utc_to_local(r.manager_approval_start_time)
                manager_start_time_str = f" {local_start.strftime('%H:%M:%S')}"

            # Build Approved (date + finance approval END time if available)
            approved_date = r.approval_date.strftime('%Y-%m-%d') if getattr(r, 'approval_date', None) else ''
            finance_end_time_str = ''
            if getattr(r, 'finance_approval_end_time', None):
                local_end = utc_to_local(r.finance_approval_end_time)
                finance_end_time_str = f" {local_end.strftime('%H:%M:%S')}"

            scheduled_date = ''
            if (not getattr(r, 'recurring', None) or getattr(r, 'recurring', None) != 'Recurring') and getattr(r, 'payment_date', None):
                try:
                    scheduled_date = r.payment_date.strftime('%Y-%m-%d')
                except Exception:
                    scheduled_date = str(r.payment_date)

            row_data = [
                f"#{r.request_id}",
                str(r.request_type or ''),
                str(r.requestor_name or ''),
                str(r.department or ''),
                str(r.recurring or 'One-Time'),
                scheduled_date,
                f"OMR {to_float(r.amount):.3f}",
                str(r.branch_name or ''),
                str(company_display or ''),
                f"{submitted_date}{manager_start_time_str}",
                f"{approved_date}{finance_end_time_str}",
                str(r.approver or '')
            ]
            
            # Pre-wrap and measure to know if the row fits in the remaining space
            for i, (data, width) in enumerate(zip(row_data, col_widths)):
                if data:
                    lines = wrap_text(str(data), width, body_font, 9)
                else:
                    lines = ['']
                wrapped_lines_per_col.append(lines)
                max_height = max(max_height, len(lines) * 10)

            # If the row will overflow the page, create a new page and redraw the header
            if y - max_height < bottom_margin:
                c.showPage()
                y = top
                c.setFont(body_font if arabic_font_name else 'Helvetica-Bold', 9)
                for hx, text in zip(col_x, headers):
                    c.drawString(hx, y, prepare_text(text))
                y -= 10
                c.line(left, y, right, y)
                y -= 8
                c.setFont(body_font, 9)

            # Draw each column with the precomputed wrapping
            for i, lines in enumerate(wrapped_lines_per_col):
                for j, line in enumerate(lines):
                    c.drawString(col_x[i], y - (j * 10), line)
            
            y -= max_height + 3  # Reduced spacing between rows for more content

        c.showPage()
        c.save()
        pdf_value = buffer.getvalue()
        buffer.close()

        from flask import make_response
        response = make_response(pdf_value)
        response.headers['Content-Type'] = 'application/pdf'
        ts = datetime.now().strftime('%Y-%m-%d_%H-%M')
        response.headers['Content-Disposition'] = f'attachment; filename=reports_{ts}.pdf'
        response.headers['Content-Length'] = str(len(pdf_value))
        return response
        
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}', 'error')
        return redirect(url_for('reports', **request.args))


# ==================== USER MANAGEMENT ROUTES (IT ONLY) ====================

@app.route('/users')
@login_required
@role_required('IT Staff', 'Department Manager')
def manage_users():
    """Manage users (IT only)"""
    # Only IT or IT Department Manager
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    users = User.query.all()
    return render_template('manage_users.html', users=users, user=current_user)


@app.route('/users/new', methods=['GET', 'POST'])
@login_required
@role_required('IT Staff', 'Department Manager')
def new_user():
    """Create a new user - IT ONLY"""
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        name = request.form.get('name')
        username = request.form.get('username')  # This is now the email
        password = request.form.get('password')
        department = request.form.get('department')
        role = request.form.get('role')
        manager_id = request.form.get('manager_id')
        
        # Check if username (email) already exists
        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            flash('Email address already exists.', 'danger')
            return redirect(url_for('new_user'))
        
        # Department restriction removed - multiple accounts per department allowed
        
        # Determine manager assignment
        # If the new user IS a Department Manager, they are managed by the General Manager
        if role == 'Department Manager':
            gm_user = User.query.filter_by(role='GM').first()
            final_manager_id = gm_user.user_id if gm_user else None
        # If the new user IS a General Manager or CEO, they are managed by Abdalaziz (Finance Admin)
        elif role in ['GM', 'CEO']:
            abdalaziz_user = User.query.filter_by(name='Abdalaziz Al-Brashdi', department='Finance').first()
            final_manager_id = abdalaziz_user.user_id if abdalaziz_user else None
        else:
            # First preference: explicit manager selection from form
            if manager_id:
                final_manager_id = manager_id
            else:
                # Fallback: find the department's manager user
                dept_manager = User.query.filter_by(department=department, role='Department Manager').first()
                if dept_manager:
                    final_manager_id = dept_manager.user_id
                else:
                    # Special rules: Office → GM, Operation → Operation Manager, Project → Operation Manager, Finance → specific named manager
                    if department == 'Office':
                        gm_user = User.query.filter_by(role='GM').first()
                        final_manager_id = gm_user.user_id if gm_user else None
                    elif department == 'Operation':
                        op_manager = User.query.filter_by(role='Operation Manager').first()
                        final_manager_id = op_manager.user_id if op_manager else None
                    elif department == 'Project':
                        op_manager = User.query.filter_by(role='Operation Manager').first()
                        final_manager_id = op_manager.user_id if op_manager else None
                    elif department == 'Finance':
                        named_manager = User.query.filter_by(name='Abdalaziz Al-Brashdi', department='Finance').first()
                        final_manager_id = named_manager.user_id if named_manager else None
                    else:
                        # No fallback: manager assignment depends solely on users with 'Department Manager' role
                        final_manager_id = None
        
        new_user = User(
            name=name,
            username=username,  # Store email as username
            department=department,
            role=role,
            manager_id=final_manager_id,
            email=username  # Store email in both username and email fields
        )
        new_user.set_password(password)
        
        db.session.add(new_user)
        db.session.commit()
        
        log_action(f"Created new user: {username} ({role}) for department: {department}")
        
        # Notify IT Staff about user creation
        notify_users_by_role(
            request=None,  # No request for user management notifications
            notification_type="user_created",
            title="New User Created",
            message=f"New user {username} ({role}) has been created for {department} department by {current_user.name}",
            request_id=None
        )
        
        flash(f'User {username} created successfully for {department} department!', 'success')
        return redirect(url_for('manage_users'))
    
    # Get all users for manager selection
    all_users = User.query.all()
    return render_template('new_user.html', user=current_user, all_users=all_users)


@app.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
@role_required('IT Staff', 'Department Manager')
def edit_user(user_id):
    """Edit user information - IT ONLY"""
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    user_to_edit = User.query.get_or_404(user_id)
    
    if request.method == 'POST':
        new_name = request.form.get('name')
        new_username = request.form.get('username')  # This is the email
        new_password = request.form.get('password')
        new_department = request.form.get('department')
        new_role = request.form.get('role')
        new_manager_id = request.form.get('manager_id')
        
        # Check if email (username) is being changed and if it already exists
        if new_username != user_to_edit.username:
            existing_user = User.query.filter_by(username=new_username).first()
            if existing_user:
                flash('Email address already exists. Please choose a different email.', 'danger')
                return redirect(url_for('edit_user', user_id=user_id))
        
        # Department restriction removed - multiple accounts per department allowed
        
        # Determine manager assignment on edit
        if new_role == 'Department Manager':
            gm_user = User.query.filter_by(role='GM').first()
            final_manager_id = gm_user.user_id if gm_user else None
        elif new_role in ['GM', 'CEO']:
            abdalaziz_user = User.query.filter_by(name='Abdalaziz Al-Brashdi', department='Finance').first()
            final_manager_id = abdalaziz_user.user_id if abdalaziz_user else None
        else:
            if new_manager_id:
                final_manager_id = new_manager_id
            else:
                dept_manager = User.query.filter_by(department=new_department, role='Department Manager').first()
                if dept_manager:
                    final_manager_id = dept_manager.user_id
                else:
                    # Special rules: Office → GM, Operation → Operation Manager, Project → Operation Manager, Finance → specific named manager
                    if new_department == 'Office':
                        gm_user = User.query.filter_by(role='GM').first()
                        final_manager_id = gm_user.user_id if gm_user else None
                    elif new_department == 'Operation':
                        op_manager = User.query.filter_by(role='Operation Manager').first()
                        final_manager_id = op_manager.user_id if op_manager else None
                    elif new_department == 'Project':
                        op_manager = User.query.filter_by(role='Operation Manager').first()
                        final_manager_id = op_manager.user_id if op_manager else None
                    elif new_department == 'Finance':
                        named_manager = User.query.filter_by(name='Abdalaziz Al-Brashdi', department='Finance').first()
                        final_manager_id = named_manager.user_id if named_manager else None
                    else:
                        # No fallback: manager assignment depends solely on users with 'Department Manager' role
                        final_manager_id = None
        
        # Update user information
        user_to_edit.name = new_name
        user_to_edit.username = new_username  # Update email/username
        user_to_edit.email = new_username     # Update email field as well
        user_to_edit.department = new_department
        user_to_edit.role = new_role
        user_to_edit.manager_id = final_manager_id
        
        # Only update password if provided
        if new_password:
            user_to_edit.set_password(new_password)
        
        db.session.commit()
        
        log_action(f"Updated user: {user_to_edit.username} ({new_role}) - Department: {new_department}")
        
        # Notify IT Staff about user update
        notify_users_by_role(
            request=None,  # No request for user management notifications
            notification_type="user_updated",
            title="User Updated",
            message=f"User {user_to_edit.username} has been updated to {new_role} in {new_department} department by {current_user.name}",
            request_id=None
        )
        
        flash(f'User {user_to_edit.username} has been updated successfully!', 'success')
        return redirect(url_for('manage_users'))
    
    # Get all users for manager selection (excluding the user being edited)
    all_users = User.query.filter(User.user_id != user_id).all()
    return render_template('edit_user.html', user=current_user, user_to_edit=user_to_edit, all_users=all_users)


@app.route('/users/<int:user_id>/unlock', methods=['POST'])
@login_required
@role_required('IT Staff', 'Department Manager')
def unlock_user(user_id):
    """Unlock a locked user account - IT ONLY"""
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    
    user_to_unlock = User.query.get_or_404(user_id)
    
    if not user_to_unlock.is_account_locked():
        flash(f'Account {user_to_unlock.username} is not locked.', 'info')
        return redirect(url_for('manage_users'))
    
    user_to_unlock.unlock_account()
    log_action(f"IT Staff {current_user.username} unlocked account: {user_to_unlock.username}")
    
    # Create notification for the user
    create_notification(
        user_id=user_to_unlock.user_id,
        title="Account Unlocked",
        message=f"Your account has been unlocked by IT Staff. You can now log in again.",
        notification_type='account_unlocked',
        request_id=None
    )
    
    flash(f'Account {user_to_unlock.username} has been unlocked successfully!', 'success')
    return redirect(url_for('manage_users'))


@app.route('/users/<int:user_id>/reset_password', methods=['POST'])
@login_required
@role_required('IT Staff', 'Department Manager')
def reset_user_password(user_id):
    """Reset user password and unlock account - IT ONLY"""
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    
    user_to_reset = User.query.get_or_404(user_id)
    new_password = request.form.get('new_password')
    
    if not new_password or len(new_password) < 6:
        flash('Password must be at least 6 characters long.', 'danger')
        return redirect(url_for('manage_users'))
    
    # Reset password
    user_to_reset.set_password(new_password)
    
    # Unlock account if locked
    if user_to_reset.is_account_locked():
        user_to_reset.unlock_account()
    
    db.session.commit()
    log_action(f"IT Staff {current_user.username} reset password for: {user_to_reset.username}")
    
    # Create notification for the user
    create_notification(
        user_id=user_to_reset.user_id,
        title="Password Reset",
        message=f"Your password has been reset by IT Staff. Please log in with your new password and change it immediately.",
        notification_type='password_reset',
        request_id=None
    )
    
    flash(f'Password for {user_to_reset.username} has been reset successfully! The account has been unlocked.', 'success')
    return redirect(url_for('manage_users'))


@app.route('/users/<int:user_id>/delete', methods=['POST'])
@login_required
@role_required('IT Staff', 'Department Manager')
def delete_user(user_id):
    """Delete a user and handle related data"""
    if current_user.role == 'Department Manager' and current_user.department != 'IT':
        flash('You do not have permission to access this page.', 'danger')
        return redirect(url_for('dashboard'))
    user_to_delete = User.query.get_or_404(user_id)
    
    # Don't allow deleting yourself
    if user_to_delete.user_id == current_user.user_id:
        flash('You cannot delete your own account.', 'danger')
        return redirect(url_for('manage_users'))
    
    username = user_to_delete.username
    
    # Handle payment requests created by this user
    payment_requests = PaymentRequest.query.filter_by(user_id=user_id).all()
    if payment_requests:
        # Delete all payment requests by this user
        for req in payment_requests:
            # Delete associated receipt files if they exist (both requestor and finance admin receipts)
            import json
            if req.requestor_receipt_path:
                try:
                    requestor_receipts = json.loads(req.requestor_receipt_path)
                    for receipt_file in requestor_receipts:
                        filepath = os.path.join(app.config['UPLOAD_FOLDER'], receipt_file)
                        if os.path.exists(filepath):
                            os.remove(filepath)
                except (json.JSONDecodeError, TypeError):
                    filepath = os.path.join(app.config['UPLOAD_FOLDER'], req.requestor_receipt_path)
                    if os.path.exists(filepath):
                        os.remove(filepath)
            
            if req.finance_admin_receipt_path:
                try:
                    finance_receipts = json.loads(req.finance_admin_receipt_path)
                    for receipt_file in finance_receipts:
                        filepath = os.path.join(app.config['UPLOAD_FOLDER'], receipt_file)
                        if os.path.exists(filepath):
                            os.remove(filepath)
                except (json.JSONDecodeError, TypeError):
                    filepath = os.path.join(app.config['UPLOAD_FOLDER'], req.finance_admin_receipt_path)
                    if os.path.exists(filepath):
                        os.remove(filepath)
            
            # Also check legacy receipt_path for backwards compatibility
            if req.receipt_path:
                try:
                    legacy_receipts = json.loads(req.receipt_path)
                    for receipt_file in legacy_receipts:
                        filepath = os.path.join(app.config['UPLOAD_FOLDER'], receipt_file)
                        if os.path.exists(filepath):
                            os.remove(filepath)
                except (json.JSONDecodeError, TypeError):
                    filepath = os.path.join(app.config['UPLOAD_FOLDER'], req.receipt_path)
                    if os.path.exists(filepath):
                        os.remove(filepath)
            if req.proof_of_payment:
                proof_file = os.path.join(app.config['UPLOAD_FOLDER'], req.proof_of_payment)
                if os.path.exists(proof_file):
                    os.remove(proof_file)
            
            # Delete related records first
            InstallmentEditHistory.query.filter_by(request_id=req.request_id).delete()
            RecurringPaymentSchedule.query.filter_by(request_id=req.request_id).delete()
            LateInstallment.query.filter_by(request_id=req.request_id).delete()
            PaidNotification.query.filter_by(request_id=req.request_id).delete()
            Notification.query.filter_by(request_id=req.request_id).delete()
            FinanceAdminNote.query.filter_by(request_id=req.request_id).delete()
            
            db.session.delete(req)
    
    # Update audit logs to preserve history (user_id will be NULL, but username_snapshot kept)
    # No need to explicitly update - the nullable foreign key will handle this
    
    # Delete the user
    db.session.delete(user_to_delete)
    db.session.commit()
    
    log_action(f"Deleted user: {username} (and {len(payment_requests)} associated requests)")
    flash(f'User {username} has been deleted successfully.', 'success')
    return redirect(url_for('manage_users'))


# ==================== DEBUG ROUTES ====================

@app.route('/debug/requests')
@login_required
@role_required('Finance Admin', 'IT Staff', 'Department Manager')
def debug_requests():
    """Debug route to see all requests in the database"""
    all_requests = PaymentRequest.query.order_by(PaymentRequest.created_at.desc()).limit(20).all()
    debug_data = []
    for req in all_requests:
        debug_data.append({
            'request_id': req.request_id,
            'requestor_name': req.requestor_name,
            'department': req.department,
            'status': req.status,
            'user_id': req.user_id,
            'created_at': req.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'amount': float(req.amount)
        })
    return jsonify(debug_data)

# ==================== NOTIFICATION ROUTES ====================

@app.route('/notifications')
@login_required
@role_required('Finance Admin', 'Admin', 'Finance Staff', 'Project Staff', 'Operation Manager', 'IT Staff', 'Department Manager', 'GM', 'CEO', 'Operation Staff', 'HR Staff', 'Purchasing Staff', 'PR Staff', 'Auditing Staff', 'Customer Service Staff', 'Marketing Staff', 'Quality Control Staff', 'Research and Development Staff', 'Office Staff', 'Maintenance Staff', 'Procurement Staff', 'Logistic Staff')
def notifications():
    """View all notifications based on RBAC permissions with pagination"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 25, type=int)
    
    # Validate per_page to prevent abuse
    if per_page not in [10, 25, 50, 100]:
        per_page = 25
    
    pagination = get_notifications_for_user(current_user, page=page, per_page=per_page)
    return render_template('notifications.html', pagination=pagination, notifications=pagination.items, user=current_user)


@app.route('/notifications/mark_read/<int:notification_id>')
@login_required
@role_required('Finance Admin', 'Admin', 'Finance Staff', 'Project Staff', 'Operation Manager', 'IT Staff', 'Department Manager', 'GM', 'CEO', 'Operation Staff', 'HR Staff', 'Purchasing Staff', 'PR Staff', 'Auditing Staff', 'Customer Service Staff', 'Marketing Staff', 'Quality Control Staff', 'Research and Development Staff', 'Office Staff', 'Maintenance Staff', 'Procurement Staff', 'Logistic Staff')
def mark_notification_read(notification_id):
    """Mark a notification as read"""
    notification = Notification.query.filter_by(notification_id=notification_id, user_id=current_user.user_id).first()
    if notification:
        notification.is_read = True
        db.session.commit()
        return jsonify({'success': True})
    return jsonify({'success': False}), 404


@app.route('/notifications/mark_all_read')
@login_required
@role_required('Finance Admin', 'Admin', 'Finance Staff', 'Project Staff', 'Operation Manager', 'IT Staff', 'Department Manager', 'GM', 'CEO', 'Operation Staff', 'HR Staff', 'Purchasing Staff', 'PR Staff', 'Auditing Staff', 'Customer Service Staff', 'Marketing Staff', 'Quality Control Staff', 'Research and Development Staff', 'Office Staff', 'Maintenance Staff', 'Procurement Staff', 'Logistic Staff')
def mark_all_notifications_read():
    """Mark all notifications as read for current user"""
    Notification.query.filter_by(user_id=current_user.user_id, is_read=False).update({'is_read': True})
    db.session.commit()
    return jsonify({'success': True})

@app.route('/notifications/mark_paid/<int:notification_id>')
@login_required
@role_required('Finance Admin', 'Admin', 'Finance Staff', 'Project Staff', 'Operation Manager', 'IT Staff', 'Department Manager', 'GM', 'CEO', 'Operation Staff', 'HR Staff', 'Purchasing Staff', 'PR Staff', 'Auditing Staff', 'Customer Service Staff', 'Marketing Staff', 'Quality Control Staff', 'Research and Development Staff', 'Office Staff', 'Maintenance Staff', 'Procurement Staff', 'Logistic Staff')
def mark_notification_paid(notification_id):
    """Mark a recurring payment notification as paid and delete it"""
    notification = Notification.query.filter_by(
        notification_id=notification_id, 
        user_id=current_user.user_id
    ).first()
    
    if not notification:
        return jsonify({'success': False, 'message': 'Notification not found'}), 404
    
    # Only allow marking recurring due notifications as paid
    if notification.notification_type != 'recurring_due':
        return jsonify({'success': False, 'message': 'This notification cannot be marked as paid'}), 400
    
    # Record that this payment was marked as paid today
    paid_notification = PaidNotification(
        request_id=notification.request_id,
        user_id=current_user.user_id,
        paid_date=date.today()
    )
    db.session.add(paid_notification)
    
    # Delete the notification
    db.session.delete(notification)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Payment marked as paid. Notification will reappear on the next due date.'})


@app.route('/notifications/delete/<int:notification_id>')
@login_required
@role_required('Finance Admin', 'Admin', 'Finance Staff', 'Project Staff', 'Operation Manager', 'IT Staff', 'Department Manager', 'GM', 'CEO', 'Operation Staff', 'HR Staff', 'Purchasing Staff', 'PR Staff', 'Auditing Staff', 'Customer Service Staff', 'Marketing Staff', 'Quality Control Staff', 'Research and Development Staff', 'Office Staff', 'Maintenance Staff', 'Procurement Staff', 'Logistic Staff')
def delete_notification(notification_id):
    """Delete a specific notification"""
    notification = Notification.query.filter_by(notification_id=notification_id, user_id=current_user.user_id).first()
    if notification:
        db.session.delete(notification)
        db.session.commit()
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': 'Notification not found'}), 404

@app.route('/notifications/delete_all')
@login_required
@role_required('Finance Admin', 'Admin', 'Finance Staff', 'Project Staff', 'Operation Manager', 'IT Staff', 'Department Manager', 'GM', 'CEO', 'Operation Staff', 'HR Staff', 'Purchasing Staff', 'PR Staff', 'Auditing Staff', 'Customer Service Staff', 'Marketing Staff', 'Quality Control Staff', 'Research and Development Staff', 'Office Staff', 'Maintenance Staff', 'Procurement Staff', 'Logistic Staff')
def delete_all_notifications():
    """Delete all notifications for current user"""
    try:
        deleted_count = Notification.query.filter_by(user_id=current_user.user_id).delete()
        db.session.commit()
        return jsonify({'success': True, 'deleted_count': deleted_count})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/notifications/unread_count')
@login_required
@role_required('Finance Admin', 'Admin', 'Finance Staff', 'Project Staff', 'Operation Manager', 'IT Staff', 'Department Manager', 'GM', 'CEO', 'Operation Staff', 'HR Staff', 'Purchasing Staff', 'PR Staff', 'Auditing Staff', 'Customer Service Staff', 'Marketing Staff', 'Quality Control Staff', 'Research and Development Staff', 'Office Staff', 'Maintenance Staff', 'Procurement Staff', 'Logistic Staff')
def unread_notifications_count():
    """Get count of unread notifications based on RBAC"""
    count = get_unread_count_for_user(current_user)
    return jsonify({'count': count})


@app.route('/api/overdue-requests/count')
@login_required
@role_required('Finance Admin', 'Finance Staff')
def overdue_requests_count():
    """API endpoint to get overdue requests count"""
    count = get_overdue_requests_count()
    return jsonify({'count': count})


@app.route('/api/notifications/recent')
@login_required
@role_required('Finance Admin', 'Admin', 'Finance Staff', 'Project Staff', 'Operation Manager', 'IT Staff', 'Department Manager', 'GM', 'CEO', 'Operation Staff', 'HR Staff', 'Purchasing Staff', 'PR Staff', 'Auditing Staff', 'Customer Service Staff', 'Marketing Staff', 'Quality Control Staff', 'Research and Development Staff', 'Office Staff', 'Maintenance Staff', 'Procurement Staff', 'Logistic Staff')
def recent_notifications():
    """Get recent notifications for the user based on RBAC"""
    notifications = get_notifications_for_user(current_user)
    return jsonify([n.to_dict() for n in notifications])

@app.route('/debug/notifications')
@login_required
@role_required('Admin')
def debug_notifications():
    """Debug route to check all notifications in database"""
    all_notifications = Notification.query.order_by(Notification.created_at.desc()).limit(10).all()
    debug_data = []
    for notif in all_notifications:
        user = User.query.get(notif.user_id)
        debug_data.append({
            'notification_id': notif.notification_id,
            'user_id': notif.user_id,
            'username': user.username if user else 'Unknown',
            'user_role': user.role if user else 'Unknown',
            'title': notif.title,
            'message': notif.message,
            'notification_type': notif.notification_type,
            'is_read': notif.is_read,
            'created_at': notif.created_at.strftime('%Y-%m-%d %H:%M:%S')
        })
    return jsonify(debug_data)

@app.route('/debug/my-notifications')
@login_required
def debug_my_notifications():
    """Debug route to check current user's notifications"""
    all_user_notifications = Notification.query.filter_by(user_id=current_user.user_id).order_by(Notification.created_at.desc()).all()
    debug_data = []
    for notif in all_user_notifications:
        debug_data.append({
            'notification_id': notif.notification_id,
            'title': notif.title,
            'message': notif.message,
            'notification_type': notif.notification_type,
            'is_read': notif.is_read,
            'created_at': notif.created_at.strftime('%Y-%m-%d %H:%M:%S')
        })
    return jsonify({
        'user_id': current_user.user_id,
        'username': current_user.username,
        'role': current_user.role,
        'department': current_user.department,
        'total_notifications': len(all_user_notifications),
        'notifications': debug_data
    })

@app.route('/debug/test-notification')
@login_required
@role_required('Admin')
def debug_test_notification():
    """Test route to create a notification for current user"""
    try:
        test_notification = create_notification(
            user_id=current_user.user_id,
            title="Test Notification",
            message="This is a test notification to verify the system is working",
            notification_type="test",
            request_id=None
        )
        return jsonify({
            'success': True,
            'message': 'Test notification created successfully',
            'notification_id': test_notification.notification_id
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# Real-time dashboard API endpoints
@app.route('/api/dashboard/finance')
@login_required
@role_required('Finance Admin', 'Finance Staff')
def api_finance_dashboard():
    """API endpoint for finance dashboard data"""
    try:
        # Get the same data as the finance dashboard
        requests = PaymentRequest.query.order_by(PaymentRequest.created_at.desc()).all()
        
        # Render the dashboard template and return HTML
        from flask import render_template
        html = render_template('finance_dashboard.html', 
                              user=current_user, 
                              requests=requests,
                              status_filter=None,
                              department_filter=None,
                              search_query=None,
                              urgent_filter=None,
                              pagination=None)
        return html
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboard/admin')
@login_required
@role_required('Admin')
def api_admin_dashboard():
    """API endpoint for admin dashboard data"""
    try:
        requests = PaymentRequest.query.order_by(PaymentRequest.created_at.desc()).all()
        
        from flask import render_template
        html = render_template('admin_dashboard.html', 
                              user=current_user, 
                              requests=requests,
                              status_filter=None,
                              department_filter=None,
                              search_query=None,
                              urgent_filter=None,
                              pagination=None)
        return html
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/dashboard/it')
@login_required
@role_required('IT Staff')
def api_it_dashboard():
    """API endpoint for IT dashboard data"""
    try:
        requests = PaymentRequest.query.order_by(PaymentRequest.created_at.desc()).all()
        
        from flask import render_template
        html = render_template('it_dashboard.html', 
                              user=current_user, 
                              requests=requests,
                              status_filter=None,
                              department_filter=None,
                              search_query=None,
                              urgent_filter=None,
                              pagination=None)
        return html
    except Exception as e:
        return jsonify({'error': str(e)}), 500






# ==================== FILE UPLOAD ROUTES ====================

@app.route('/uploads/receipts/<filename>')
@login_required
def uploaded_file(filename):
    """Serve uploaded receipt files"""
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


@app.route('/uploads/cheque/<filename>')
@login_required
def uploaded_cheque_file(filename):
    """Serve uploaded cheque files"""
    return send_from_directory(app.config['CHEQUE_UPLOAD_FOLDER'], filename)


@app.route('/cheque-register/upload', methods=['POST'])
@login_required
def upload_cheque_file():
    """Upload a file for a cheque serial"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        serial_id = request.form.get('serial_id')
        
        if not serial_id:
            return jsonify({'success': False, 'error': 'Serial ID not provided'}), 400
        
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        # Get the cheque serial
        cheque_serial = ChequeSerial.query.get(serial_id)
        if not cheque_serial:
            return jsonify({'success': False, 'error': 'Cheque serial not found'}), 404
        
        # Validate file extension
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'pdf'}
        file_extension = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
        if file_extension not in allowed_extensions:
            return jsonify({'success': False, 'error': 'Invalid file type. Allowed: JPG, PNG, GIF, PDF'}), 400
        
        # Validate file size (50MB max)
        max_file_size = app.config.get('MAX_FILE_SIZE', 50 * 1024 * 1024)
        file.seek(0, 2)  # Seek to end
        file_size = file.tell()
        file.seek(0)  # Reset to beginning
        
        if file_size > max_file_size:
            return jsonify({'success': False, 'error': f'File too large. Maximum size is {max_file_size // (1024 * 1024)}MB'}), 400
        
        # Generate unique filename
        import uuid
        
        unique_id = str(uuid.uuid4())[:8]
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = secure_filename(f"cheque_{serial_id}_{timestamp}_{unique_id}_{file.filename}")
        filepath = os.path.join(app.config['CHEQUE_UPLOAD_FOLDER'], filename)
        
        # Save file
        file.save(filepath)
        
        # Update database
        cheque_serial.upload_path = filename
        db.session.commit()
        
        return jsonify({
            'success': True,
            'filename': filename,
            'message': 'File uploaded successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== ERROR HANDLERS ====================

@app.errorhandler(404)
def not_found(error):
    return render_template('404.html'), 404


@app.errorhandler(403)
def forbidden(error):
    return render_template('403.html'), 403


@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('500.html'), 500


@app.errorhandler(413)
def request_entity_too_large(error):
    """Handle request entity too large errors (file upload size exceeded)"""
    error_message = 'The uploaded file(s) exceed the maximum allowed size of 100MB total. Please reduce file sizes and try again.'
    
    # Check if it's a fetch request from /request/new (new payment request form)
    # Fetch requests don't set Accept: application/json for FormData, so we check the path
    if request.path == '/request/new' and request.method == 'POST':
        try:
            return jsonify({
                'error': 'Request too large',
                'message': error_message
            }), 413
        except:
            pass
    
    # Check if client accepts JSON (AJAX/fetch requests)
    accepts_json = 'application/json' in request.headers.get('Accept', '')
    if accepts_json or request.is_json:
        try:
            return jsonify({
                'error': 'Request too large',
                'message': error_message
            }), 413
        except:
            pass
    
    # For regular form submissions, redirect with flash message
    flash(f'File upload failed: {error_message}', 'error')
    if request.path == '/request/new':
        try:
            return redirect(url_for('new_request')), 413
        except:
            pass
    try:
        return redirect(request.referrer or url_for('dashboard')), 413
    except:
        # Fallback if redirect fails
        return f'File upload failed: {error_message}', 413



# ==================== MAIN ====================



@app.route('/admin/calendar')
@role_required('Admin', 'Project Staff', 'Finance Admin', 'Finance Staff', 'GM', 'CEO', 'Operation Manager', 'IT Staff', 'IT Department Manager')
def admin_calendar():
    """Calendar view for recurring payments (Admin and Project roles)"""
    return render_template('admin_calendar.html')

@app.route('/api/admin/recurring-events')
@role_required('Admin', 'Project Staff', 'Finance Admin', 'Finance Staff', 'GM', 'CEO', 'Operation Manager', 'IT Staff', 'IT Department Manager')
def api_admin_recurring_events():
    """API endpoint for calendar events (recurring + one-time scheduled)."""
    try:
        # Build query for recurring payment requests (exclude archived)
        # Include requests with Payment Type "Recurring" and status in allowed list
        query = PaymentRequest.query.filter(
            PaymentRequest.recurring == 'Recurring',  # Payment Type must be "Recurring"
            PaymentRequest.recurring_interval.isnot(None),
            PaymentRequest.recurring_interval != '',
            PaymentRequest.status.in_([
                'Pending Finance Approval',
                'Proof Pending',
                'Proof Sent',
                'Proof Rejected',
                'Completed',
                'Recurring'
            ]),
            PaymentRequest.is_archived == False
        )
        
        # Project users can only see their department's requests
        if current_user.role == 'Project Staff':
            query = query.filter(PaymentRequest.department == current_user.department)
        
        recurring_requests = query.all()
        
        # Group events by date
        events_by_date = {}
        today = date.today()
        
        for req in recurring_requests:
            # Check if this is a variable payment with installments
            schedule = RecurringPaymentSchedule.query.filter_by(request_id=req.request_id).all()
            
            if schedule:
                # For variable payments, calculate remaining amount
                total_paid_amount = 0
                paid_notifications = PaidNotification.query.filter_by(request_id=req.request_id).all()
                
                for installment in schedule:
                    # Check if this specific installment is paid
                    is_paid = any(
                        paid_notif.paid_date == installment.payment_date 
                        for paid_notif in paid_notifications
                    )
                    
                    # If this installment is paid, add its amount to total paid
                    if is_paid:
                        total_paid_amount += float(installment.amount)
                
                # Calculate remaining amount
                remaining_amount = float(req.amount) - total_paid_amount
                
                # If remaining amount is 0 or less, skip this request
                if remaining_amount <= 0:
                    continue
            else:
                # For regular recurring payments, check if there are any paid notifications
                paid_notifications_count = PaidNotification.query.filter_by(request_id=req.request_id).count()
                
                # If there are paid notifications, skip this request entirely
                if paid_notifications_count > 0:
                    continue
            
            if schedule:
                # For variable payments, show only the specific installment dates
                for installment in schedule:
                    # Check if this specific installment is paid using the RecurringPaymentSchedule.is_paid field
                    is_paid = installment.is_paid
                    
                    # Determine event color (red if marked late)
                    is_late = LateInstallment.query.filter_by(
                        request_id=req.request_id,
                        payment_date=installment.payment_date
                    ).first() is not None
                    
                    # Debug: Log payment status determination
                    print(f"Request {req.request_id}, Date {installment.payment_date}: Paid={is_paid}, Late={is_late}")
                    
                    event_color = '#2e7d32' if is_paid else ('#d32f2f' if is_late else '#8e24aa')
                    
                    # Calculate remaining amount
                    total_paid = sum(
                        pn.amount if hasattr(pn, 'amount') else 0 
                        for pn in PaidNotification.query.filter_by(request_id=req.request_id).all()
                    )
                    remaining_amount = req.amount - total_paid
                    
                    # Add to events_by_date
                    date_key = installment.payment_date.isoformat()
                    if date_key not in events_by_date:
                        events_by_date[date_key] = []
                    
                    events_by_date[date_key].append({
                        'title': f'OMR {installment.amount:.3f}',
                        'start': installment.payment_date.isoformat(),
                        'color': event_color,
                        'url': f'/request/{req.request_id}',
                        'extendedProps': {
                            'requestId': req.request_id,
                            'requestType': req.request_type,
                            'companyName': req.person_company or req.company_name or 'N/A',
                            'department': req.department,
                            'purpose': req.purpose,
                            'baseAmount': f'OMR {req.amount:.3f}',
                            'remainingAmount': f'OMR {remaining_amount:.3f}'
                        }
                    })
            else:
                # For regular recurring payments, generate future due dates
                start_date = today
                end_date = today + timedelta(days=365)
                
                due_dates = generate_future_due_dates(req, start_date, end_date)
                
                for due_date, amount in due_dates:
                    # Check if this payment is already marked as paid
                    paid_notification = PaidNotification.query.filter_by(
                        request_id=req.request_id,
                        paid_date=due_date
                    ).first()
                    
                    # Determine event color (red if marked late)
                    is_late = LateInstallment.query.filter_by(
                        request_id=req.request_id,
                        payment_date=due_date
                    ).first() is not None
                    
                    # Debug: Log payment status determination
                    print(f"Request {req.request_id}, Date {due_date}: Paid={paid_notification is not None}, Late={is_late}")
                    
                    event_color = '#2e7d32' if paid_notification else ('#d32f2f' if is_late else '#8e24aa')
                    
                    # Add to events_by_date
                    date_key = due_date.isoformat()
                    if date_key not in events_by_date:
                        events_by_date[date_key] = []
                    
                    events_by_date[date_key].append({
                        'title': f'OMR {amount:.3f}',
                        'start': due_date.isoformat(),
                        'color': event_color,
                        'url': f'/request/{req.request_id}',
                        'extendedProps': {
                            'requestId': req.request_id,
                            'requestType': req.request_type,
                            'companyName': req.person_company or req.company_name or 'N/A',
                            'department': req.department,
                            'purpose': req.purpose,
                            'baseAmount': None,
                            'remainingAmount': None
                        }
                    })
        
        # Also include ONE-TIME scheduled payments that the current user is authorized to see
        def is_authorized_for_one_time(req, user):
            # Requestor
            if req.user_id == user.user_id:
                return True
            # Temporary manager
            if getattr(req, 'temporary_manager_id', None) == user.user_id:
                return True
            # GM and Operation Manager (global)
            if user.role in ['GM', 'Operation Manager']:
                return True
            # IT Staff and IT Department Manager (global IT access)
            if user.department == 'IT' and user.role in ['IT Staff', 'Department Manager']:
                return True
            # Assigned manager
            if getattr(req.user, 'manager_id', None) == user.user_id:
                return True
            # Department Manager of same department
            if user.role == 'Department Manager' and user.department == req.department:
                return True
            # Finance Admin special rules
            if user.role == 'Finance Admin':
                if req.status == 'Pending Finance Approval':
                    if user.name == 'Abdalaziz Al-Brashdi':
                        return req.user_id == user.user_id or getattr(req.user, 'manager_id', None) == user.user_id
                    else:
                        return req.user_id == user.user_id
            return False

        one_time_query = PaymentRequest.query.filter(
            (PaymentRequest.recurring.is_(None)) | (PaymentRequest.recurring != 'Recurring'),
            PaymentRequest.payment_date.isnot(None),
            PaymentRequest.status.in_([
                'Pending Finance Approval',
                'Proof Pending',
                'Proof Sent',
                'Proof Rejected',
                'Completed',
                'Recurring'
            ]),
            PaymentRequest.is_archived == False
        )
        # Project Staff can only see their department's one-time requests as well
        if current_user.role == 'Project Staff':
            one_time_query = one_time_query.filter(PaymentRequest.department == current_user.department)
        one_time_requests = one_time_query.all()

        for req in one_time_requests:
            if not is_authorized_for_one_time(req, current_user):
                continue
            date_key = req.payment_date.isoformat()
            if date_key not in events_by_date:
                events_by_date[date_key] = []
            events_by_date[date_key].append({
                'title': f'OMR {float(req.amount):.3f}',
                'start': req.payment_date.isoformat(),
                'color': '#0d6efd',  # blue for scheduled one-time
                'url': f'/request/{req.request_id}',
                'extendedProps': {
                    'requestId': req.request_id,
                    'requestType': req.request_type,
                    'companyName': req.person_company or req.company_name or 'N/A',
                    'department': req.department,
                    'purpose': req.purpose,
                    'baseAmount': f'OMR {float(req.amount):.3f}',
                    'remainingAmount': None,
                    'oneTime': True
                }
            })

        # Convert grouped events to calendar format
        calendar_events = []
        for date_key, day_events in events_by_date.items():
            # Create a summary event for the day
            total_amount = sum(float(str(event['title']).replace('OMR ', '')) for event in day_events)
            count = len(day_events)
            
            # Determine the overall color for this date
            # Check if all payments are paid (green), all are late (red), or mixed/due (purple)
            paid_count = sum(1 for event in day_events if event.get('color') == '#2e7d32')
            late_count = sum(1 for event in day_events if event.get('color') == '#d32f2f')
            
            if paid_count == count:
                # All payments are paid - green
                date_color = '#2e7d32'
                status_text = 'paid'
            elif late_count > 0:
                # Some or all payments are late - red
                date_color = '#d32f2f'
                status_text = 'late'
            else:
                # All payments are due - purple
                date_color = '#8e24aa'
                status_text = 'due'
            
            calendar_events.append({
                'title': f'{count} payment{"s" if count != 1 else ""} {status_text}',
                'start': date_key,
                'color': date_color,
                'extendedProps': {
                    'count': count,
                    'totalAmount': total_amount,
                    'date': date_key,
                    'status': status_text,
                    'paidCount': paid_count,
                    'lateCount': late_count,
                    'events': day_events  # Store all events for this date
                }
            })
        
        return jsonify(calendar_events)
        
    except Exception as e:
        print(f"Error generating calendar events: {e}")
        return jsonify([])

def add_months(source_date, months):
    """Add months to a date, handling month overflow correctly"""
    month = source_date.month - 1 + months
    year = source_date.year + month // 12
    month = month % 12 + 1
    day = min(source_date.day, [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month - 1])
    return date(year, month, day)

def generate_future_due_dates(req, start_date, end_date):
    """Generate future due dates for a recurring request"""
    if not req.recurring_interval:
        return []
    
    try:
        # Parse the recurring interval
        parts = req.recurring_interval.split(':')
        if len(parts) < 2:
            return []
        
        frequency = parts[0]
        interval = int(parts[1])
        
        if frequency == 'monthly':
            due_dates = []
            
            # Check for new :date: format (e.g., monthly:1:date:2025-10-18:end:2026-01-30)
            if len(parts) > 2 and parts[2] == 'date':
                # Parse the start date
                try:
                    base_date_str = parts[3]
                    base_date = datetime.strptime(base_date_str, '%Y-%m-%d').date()
                    
                    # Parse end date if present
                    config_end_date = None
                    if 'end' in parts:
                        end_index = parts.index('end')
                        if end_index + 1 < len(parts):
                            config_end_date = datetime.strptime(parts[end_index + 1], '%Y-%m-%d').date()
                    
                    # Use the earlier of the two end dates
                    effective_end_date = end_date
                    if config_end_date:
                        effective_end_date = min(end_date, config_end_date)
                    
                    # Generate recurring dates starting from base_date
                    current_date = base_date
                    if current_date < start_date:
                        # If base_date is before start_date, advance to the first occurrence after start_date
                        months_diff = (start_date.year - base_date.year) * 12 + (start_date.month - base_date.month)
                        cycles = (months_diff // interval) + 1
                        current_date = add_months(base_date, cycles * interval)
                    
                    # Generate dates
                    while current_date <= effective_end_date:
                        if current_date >= start_date:
                            due_dates.append((current_date, req.amount))
                        # Move to next occurrence
                        current_date = add_months(current_date, interval)
                    
                except (ValueError, IndexError) as e:
                    print(f"Error parsing date format: {e}")
                    return []
            
            # Check if old specific days format is configured
            elif len(parts) > 2 and parts[2] == 'days':
                # Parse specific days from the interval
                days = [int(day) for day in parts[3].split(',')]
                year = int(parts[4]) if len(parts) > 4 else start_date.year
                month = int(parts[5]) if len(parts) > 5 else start_date.month
                
                # Generate dates for the specific days
                current_date = start_date
                month_count = 0
                
                while current_date <= end_date:
                    if month_count % interval == 0:
                        # Generate dates for each specific day in this month
                        for day in days:
                            try:
                                # Create date for this specific day
                                due_date = date(current_date.year, current_date.month, day)
                                if due_date >= start_date and due_date <= end_date:
                                    due_dates.append((due_date, req.amount))
                            except ValueError:
                                # Skip invalid dates (like Feb 30)
                                continue
                    
                    # Move to next month
                    if current_date.month == 12:
                        current_date = date(current_date.year + 1, 1, 1)
                    else:
                        current_date = date(current_date.year, current_date.month + 1, 1)
                    month_count += 1
            else:
                # Fallback to original logic for simple monthly intervals
                current_date = start_date
                month_count = 0
                while current_date <= end_date:
                    if month_count % interval == 0:
                        due_dates.append((current_date, req.amount))
                    current_date = current_date + timedelta(days=1)
                    if current_date.day == 1:  # New month
                        month_count += 1
            
            return due_dates[:12]  # Limit to 12 months
        
        return []
        
    except Exception as e:
        print(f"Error generating due dates: {e}")
        return []


@app.route('/api/requests/mark_paid', methods=['POST'])
@role_required('Admin')
def mark_request_paid():
    """Mark an entire request as paid"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No data provided'}), 400
        
        request_id = data.get('request_id')
        amount = data.get('amount')
        
        if not request_id or not amount:
            return jsonify({'success': False, 'message': 'Missing required fields'}), 400
        
        # Get the payment request
        payment_request = PaymentRequest.query.get(request_id)
        if not payment_request:
            return jsonify({'success': False, 'message': 'Request not found'}), 404
        
        # Check if request is approved
        if payment_request.status != 'Approved':
            return jsonify({'success': False, 'message': 'Request must be approved before marking as paid'}), 400
        
        # Check if request is recurring
        if not payment_request.recurring_interval:
            return jsonify({'success': False, 'message': 'This endpoint is only for recurring payments'}), 400
        
        # Check if this request is already paid
        existing_paid = PaidNotification.query.filter_by(request_id=request_id).first()
        
        if existing_paid:
            return jsonify({'success': False, 'message': 'This request is already marked as paid'}), 400
        
        # Create paid notification for the entire request
        paid_notification = PaidNotification(
            request_id=request_id,
            user_id=current_user.user_id,
            paid_date=date.today()
        )
        
        db.session.add(paid_notification)
        db.session.commit()
        
        # Log the action
        audit_log = AuditLog(
            user_id=current_user.user_id,
            action='Mark Request Paid',
            details=f'Marked entire request #{request_id} (OMR {amount}) as paid',
            ip_address=request.remote_addr
        )
        db.session.add(audit_log)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Request marked as paid successfully'})
        
    except Exception as e:
        print(f"Error marking request as paid: {e}")
        db.session.rollback()
        return jsonify({'success': False, 'message': 'An error occurred while marking request as paid'}), 500


@app.route('/api/installments/mark_paid', methods=['POST'])
@role_required('Admin', 'Finance Staff')
def mark_installment_paid():
    """Mark a specific installment as paid"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No data provided'}), 400
        
        request_id = data.get('request_id')
        payment_date = data.get('payment_date')
        amount = data.get('amount')
        
        if not request_id or not payment_date or not amount:
            return jsonify({'success': False, 'message': 'Missing required fields'}), 400
        
        # Convert payment_date string to date object
        from datetime import datetime
        try:
            payment_date_obj = datetime.strptime(payment_date, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'message': 'Invalid date format'}), 400
        
        # Get the payment request
        payment_request = PaymentRequest.query.get(request_id)
        if not payment_request:
            return jsonify({'success': False, 'message': 'Request not found'}), 404
        
        # Check if request is approved
        if payment_request.status != 'Approved':
            return jsonify({'success': False, 'message': 'Request must be approved before marking as paid'}), 400
        
        # Check if this installment is already paid
        existing_paid = PaidNotification.query.filter_by(
            request_id=request_id,
            paid_date=payment_date_obj
        ).first()
        
        if existing_paid:
            return jsonify({'success': False, 'message': 'This installment is already marked as paid'}), 400
        
        # Create paid notification
        paid_notification = PaidNotification(
            request_id=request_id,
            user_id=current_user.user_id,
            paid_date=payment_date_obj
        )
        
        db.session.add(paid_notification)
        db.session.commit()
        
        # Log the action
        audit_log = AuditLog(
            user_id=current_user.user_id,
            action='Mark Installment Paid',
            details=f'Marked installment of OMR {amount} due on {payment_date} as paid for request #{request_id}',
            ip_address=request.remote_addr
        )
        db.session.add(audit_log)
        db.session.commit()
        
        # Check if all installments are now paid and mark as completed if so
        check_recurring_payment_completion(request_id)
        
        return jsonify({'success': True, 'message': 'Installment marked as paid successfully'})
        
    except Exception as e:
        print(f"Error marking installment as paid: {e}")
        db.session.rollback()
        return jsonify({'success': False, 'message': 'An error occurred while marking installment as paid'}), 500


@app.route('/api/installments/mark_late', methods=['POST'])
@role_required('Admin')
def mark_installment_late():
    """Mark a specific installment as late (Admin only)"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No data provided'}), 400
        
        request_id = data.get('request_id')
        payment_date = data.get('payment_date')
        
        if not request_id or not payment_date:
            return jsonify({'success': False, 'message': 'Missing required fields'}), 400
        
        # Convert payment_date string to date object
        from datetime import datetime
        try:
            payment_date_obj = datetime.strptime(payment_date, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'message': 'Invalid date format'}), 400
        
        # Get the payment request
        payment_request = PaymentRequest.query.get(request_id)
        if not payment_request:
            return jsonify({'success': False, 'message': 'Request not found'}), 404
        
        # Check if request is approved
        if payment_request.status != 'Approved':
            return jsonify({'success': False, 'message': 'Request must be approved before marking installments as late'}), 400
        
        # If already paid, cannot be marked late
        existing_paid = PaidNotification.query.filter_by(
            request_id=request_id,
            paid_date=payment_date_obj
        ).first()
        if existing_paid:
            return jsonify({'success': False, 'message': 'This installment is already marked as paid'}), 400
        
        # Check if already marked late
        from models import LateInstallment
        existing_late = LateInstallment.query.filter_by(
            request_id=request_id,
            payment_date=payment_date_obj
        ).first()
        if existing_late:
            return jsonify({'success': True, 'message': 'Installment already marked as late'})
        
        # Create late installment record
        late = LateInstallment(
            request_id=request_id,
            payment_date=payment_date_obj,
            marked_by_user_id=current_user.user_id
        )
        db.session.add(late)
        db.session.commit()
        
        # Log the action
        log_action(f"Marked installment due on {payment_date} as LATE for request #{request_id}")
        
        return jsonify({'success': True, 'message': 'Installment marked as late'})
    except Exception as e:
        print(f"Error marking installment as late: {e}")
        db.session.rollback()
        return jsonify({'success': False, 'message': 'An error occurred while marking installment as late'}), 500


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        # Check for timing alerts on startup
        try:
            check_finance_approval_timing_alerts()
            print("Timing alerts check completed on startup")
        except Exception as e:
            print(f"Error checking timing alerts on startup: {e}")
        
        # Start background scheduler in a separate thread
        scheduler_thread = threading.Thread(target=background_scheduler, daemon=True)
        scheduler_thread.start()
        print("Background scheduler started")
    
    socketio.run(app, debug=True, host='0.0.0.0', port=5005)

